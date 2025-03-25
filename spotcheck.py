#!/usr/bin/python3
################################### IMPORT MODULES - START ########################################
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk
from functools import partial
from time import sleep, gmtime, strftime
from datetime import *
from picamera import PiCamera
import cv2
import numpy as np
from matplotlib import pyplot
from PIL import ImageTk, Image
import serial
from fractions import Fraction
import os
import shutil
import awesometkinter as atk
import math
from enum import Enum
from ftplib import FTP
# ~ import Pmw
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as Img
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders
import re
import dns.resolver
import socket
import board
import adafruit_ds1307
import time
import RPi.GPIO as GPIO
################################### IMPORT MODULES - END ########################################
################################### CHECK APP INSTANCE - START #############################################
if not os.path.exists('/home/pi/Spotcheck/.instance.txt'):
	fw_instance = open('/home/pi/Spotcheck/.instance.txt', 'x')
	fw_instance.writelines('0\n')
	number_of_instance = 0
	fw_instance.close()
else:
	fr_instance = open('/home/pi/Spotcheck/.instance.txt')
	number_of_instance = int(fr_instance.readline().strip('\n'))
	fr_instance.close()
	
print("number_of_instance",number_of_instance)

if(number_of_instance == 0):
	fw_instance = open('/home/pi/Spotcheck/.instance.txt', 'w')
	fw_instance.writelines('1\n')
	fw_instance.close()
	number_of_instance = 1
else:
	os._exit(0)

#################################### CHECK APP INSTANCE - END #############################################
##################################### GPIO INIT - START ###################################
BLUELIGHT_PIN = 26
GPIO.setwarnings(False)
GPIO.setmode(GPIO.BCM)
GPIO.setup(BLUELIGHT_PIN, GPIO.OUT, initial=GPIO.LOW)
###################################### GPIO INIT - END ####################################
#################################### GUI RULES - START #########################################
APP_BGD_COLOR = "white smoke"

MAIN_MENU_BUTTON_BGD_COLOR = "grey80"
MAIN_MENU_BUTTON_TXT_COLOR = "black"
MAIN_MENU_BUTTON_ACTIVE_BGD_COLOR = "lawn green"
MAIN_MENU_BUTTON_FRAME_BGD_COLOR = "white smoke"
MAIN_MENU_BUTTON_WIDTH = 15
MAIN_MENU_BUTTON_HEIGHT = 4
MAIN_FUNCTION_FRAME_BGD_COLOR = "white"
MAIN_TITLE_FRAME_BGD_COLOR = "dodger blue"
MAIN_BUTTON_FRAME_BGD_COLOR = "dodger blue"

LABEL_FRAME_TXT_COLOR = "black"
LABEL_FRAME_BGD_COLOR = "white"
LABEL_TXT_COLOR = "black"
LABEL_BGD_COLOR = "white"

MAIN_FUNCTION_BUTTON_WIDTH = 12
MAIN_FUNCTION_BUTTON_HEIGHT = 3
MAIN_FUNCTION_BUTTON_BGD_COLOR = "grey80"
MAIN_FUNCTION_BUTTON_TXT_COLOR = "black"

TITILE_FRAME_BGD_COLOR = "dodger blue"
TITILE_FRAME_TXT_COLOR = "white"

SWITCH_PAGE_BUTTON_WIDTH = 10
SWITCH_PAGE_BUTTON_HEIGHT = 2
SWITCH_PAGE_BUTTON_BGD_COLOR = "grey80"
SWITCH_PAGE_BUTTON_TXT_COLOR = "black"

CONFIRM_BUTTON_WIDTH = 10
CONFIRM_BUTTON_HEIGHT = 2
CONFIRM_BUTTON_BGD_COLOR = "grey80"
CONFIRM_BUTTON_TXT_COLOR = "black"

SAMPLE_BUTTON_WIDTH = 2
SAMPLE_BUTTON_HEIGHT = 2
SAMPLE_BUTTON_FRAME_BDG_COLOR = "grey96"
SAMPLE_BUTTON_BGD_COLOR = "lavender"
SAMPLE_BUTTON_TXT_COLOR = "black"
SAMPLE_BUTTON_ACTIVE_BGD_COLOR = "white"
SAMPLE_BUTTON_CHOOSE_BGD_COLOR = "turquoise2"
SAMPLE_BUTTON_DONE_BGD_COLOR = "lawn green"
SAMPLE_BUTTON_TMP_BGD_COLOR = "grey99"

RESULT_TABLE_FRAME_BGD_COLOR = "grey96"
RESULT_LABEL_BGD_COLOR = "lawn green"
RESULT_LABEL_ERROR_BGD_COLOR = "firebrick2"
RESULT_LABEL_TXT_COLOR = "black"

PROGRAM_BUTTON_BGD_COLOR = "cyan2"
PROGRAM_BUTTON_TXT_COLOR = "black"
PROGRAM_BUTTON_ACTIVE_BGD_COLOR = "lawn green"

LOGIN_BUTTON_BGD_COLOR = "grey85"
LOGIN_BUTTON_TXT_COLOR = "black"

NA_COLOR = "grey82"
ERROR_COLOR = "DodgerBLue2"
NEGATIVE_COLOR = "green3"
POSITIVE_COLOR = "red"
LOW_COPY_COLOR = "pink"

QS_FIRSTWELL_COLOR = "tomato1"
QS_LASTWELL_COLOR = "deep sky blue"

MAIN_MENU_BUTTON_FONT = ('Helvetica', 10, 'bold')
LABELFRAME_TXT_FONT = ('Helvetica', 12)
MAIN_FUCNTION_BUTTON_FONT = ('Helvetica', 10)
TITLE_TXT_FONT = ('Helvetica', 10, 'bold')
SWITCH_PAGE_BUTTON_FONT = ('Helvetica', 10)
LABEL_FRAME_TXT_FONT = ('Helvetica', 10)
LABEL_TXT_FONT = ('Helvetica', 10)
ENTRY_TXT_FONT = ('Helvetica', 11)
CONFIRM_BUTTON_TXT_FONT = ('Helvetica', 10)
SAMPLE_BUTTON_TXT_FONT = ('Helvetica', 10)
SAMPLE_LABEL_TXT_FONT = ('Helvetica', 13)
RESULT_LABEL_TXT_FONT = ('Helvetica', 10)
RESULT_LABEL_TXT_FONT_1 = ('Helvetica', 10, 'bold')
PROGRAM_BUTTON_TXT_FONT = ('Helvetica', 10)
LOGIN_LABEL_TXT_FONT = ('Helvetica', 15)
LOGIN_BUTTON_TXT_FONT = ('Helvetica', 10)

RESULT_CELL_START = 18


MainScreen_Language = {
	"Screening Button": ["SCREENING MODE", "ĐỊNH TÍNH"],
	"Quantitative Button": ["QUANTITATIVE MODE", "ĐỊNH LƯỢNG"],
	"Environment Button": ["Environment", "Môi trường"],
	"Host Button": ["   Host   ", "Vật chủ"],
	"Analysis Button": ["Analysis", "Phân tích"],
	"Setting Button 2": ["Setting", "Cài đặt"],
	"ViewResult Button": ["View result", "Kết quả"],
	"CreateFile Button": ["Create samples file", "Tạo tệp các mẫu"],
	"Setting Button 1": ["Connectivity", "Kết nối"],
	"Exit Button": ["Exit", "Thoát"],
	"Screening LabelFrame": ["Screening Mode", "Định tính"],
	"Quantitative LabelFrame": ["Quantitative Mode", "Định lượng"],

	##### Messagebox #####
	"Exit Confirm": ["Do you want to close the app ?","Bạn có muốn đóng ứng dụng ?"],
	"SystemCheck Ask": ["It's been a while since the last system check, would you like to check again ?", "Bạn có muốn kiểm tra lại hệ thống ?"]
}

ViewResult_Language = {
	"Title Label": ["VIEW RESULT", "XEM KẾT QUẢ"],
	"Back Button": ["Back", "Quay lại"],
	"Open Button": ["Open", "Mở"],
	"Information LabelFrame": ["Information", "Thông tin"],
	"TestName Label": ["Test name", "Tên thí nghiệm"],
	"TechnicianName Label": ["Technician name", "Người thực hiện"],
	"Date Label": ["Date", "Ngày thực hiện"],
	"SampleName Button": ["Sample name", "Tên mẫu"],
	"SamplePosition Button": ["Sample name", "Tên mẫu"],
	"Result Button": ["Result", "Kết quả"],
}

CreateFile_Language = {
	"Title Label": ["SAMPLES FILE SETUP", "CÀI ĐẶT TỆP MẪU"],
	"Back Button": ["Back", "Quay lại"],
	"Load Button": ["Load", "Mở"],
	"Create Button": ["Create", "Tạo"],
	"SampleProperties LabelFrame": ["Sample properties", "Đặc tính mẫu"],
	"SampleName Label": ["Sample name", "Tên mẫu"],
	"QuickSetup LabelFrame": ["Quick Setup", "Thiết lập nhanh"],
	"FirstWell Button": ["First well", "Giếng đầu"],
	"LastWell Button": ["Last well", "Giếng cuối"],
	"OK Button": ["Set", "Đặt"],
	"Set Button": ["Apply", "Xác nhận"],

	##### Messagebox #####
	"SampleName Empty": ["Please enter sample name.", "Xin nhập tên mẫu."],
	"CreateFile Confirm": ["Do you want to create file ?", "Bạn muốn khởi tạo tệp ?"],
	"CreateFile Done": ["File has been created.\n Do you want to go back to the previous screen ?", "Tệp đã được tạo.\n Bạn có muốn quay về màn hình trước ?"],
	"FileName OverLength": ["File name cannot be more than 30 characters.", "Tên tệp không thể nhiều hơn 30 ký tự."],
	"AllowPutSample Inform": ["Now you can place the sample on the plate before press [Next].", "Bạn đã có thể cho mẫu vào plate sau đó ấn [Kế tiếp]."],
	"QuickSetup Confirm": ["Do you want to automatically name the samples ?","Bạn có muốn tự động đặt tên các mẫu ?"]
}

Connect_Language = {
	"Title Label": ["CONNECTIVITY","KẾT NỐI"],
	"Back Button": ["Back", "Quay lại"]
}

Screening0_Language = {
	"Title Label": ["SCREENING", "ĐỊNH TÍNH"],
	"ExperimentName Label": ["File name :", "Tên tệp :"],
	"TechnicianName Label": ["Technician: ", "Người thực hiện: "],
	"TemplateName Label": ["Template: ", "Template: "],
	"Back Button": ["Back", "Quay lại"],
	"Next Button": ["Next", "Kế tiếp"],

	##### Messagebox #####
	"Folder Exists": ["This folder already exists, do you want to overwrite it ?", "Thư mục này đã tồn tại, bạn có muốn ghi đè ?"]
}

Screening2_Language = {
	"Title Label": ["SCREENING", "ĐỊNH TÍNH"],
	"SamplesFile LabelFrame": ["Samples file", "Tệp mẫu"],
	"Load Button": ["Load", "Mở"],
	"Create Button": ["Create", " Tạo"],
	"Back Button": ["Back", "Quay lại"],
	"Next Button": ["Next", "Tiếp tục"],
	"OK Button" : ["OK", "Xác nhận"],
	"Cancel Button": ["Cancel", "Huỷ"],

	##### Messagebox #####
	"SamplesFile Empty": ["You haven't loaded the samples file.", "Bạn chưa tải lên tệp mẫu."],
	"Email Confirm": ["Do you want the device to automatically email the results ?", "Bạn có muốn thiết bị tự động email kết quả ?"],
	"Email Empty": ["Please enter recipient email.", "Bạn chưa nhập email người nhận."],
	"AllowPutSample Inform": ["Now you can place the sample on the plate before press [Next].", "Bạn đã có thể cho mẫu vào plate sau đó ấn [Kế tiếp]."],
	"ServerCheck Error": ["There was an error while syncing the server", "Có lỗi xảy ra khi đồng bộ với server"]
}

Screening3_Language = {
	"Title Label": ["SCREENING", "ĐỊNH TÍNH"],
	"Process Label": ["Processing...", "Đang xử lý..."],
	"Result Tab": ["Result", "Kết quả"],
	"Report Tab": ["Report", "Báo cáo"],
	"Images Tab": ["Images", "Hình ảnh"],
	"Note Negative Label": ["NEGATIVE (N)", "ÂM TÍNH (N)"],
	"Note LowCopy Label": ["LOW COPY (P_L)", "BẢN SAO THẤP (P_L)"],
	"Note Positive Label": ["POSITIVE (N)", "DƯƠNG TÍNH (P)"],
	"Note NoSample Label": ["NO SAMPLE (N/A)", "KHÔNG MẪU (N/A)"],

	"RawImage LabelFrame": ["Raw Image", "Ảnh chụp"],
	"AnalyzedImage LabelFrame": ["Analyzed Image", "Ảnh phân tích"],
	"Environment RadioButton": ["Screening", "Định tính"],
	"Host RadioButton": ["Host", "Vật chủ"],
	"Quantitative RadioButton": ["Quantitative", "Định lượng"],
	"Finish Button": ["Finish", "Hoàn thành"],
	"Save Button": ["Save", "Lưu"],
	"Result Title Text": ["ANALYSIS RESULTS", "KẾT QUẢ"],
	"Result ExperimentName Text": ["Test :", "Xét nghiệm :"],
	"Result TechnicianName Text": ["Technician: ", "Người thực hiện: "],
	"Result Date Text": ["Date: ", "Ngày thực hiện: "],
	"Result Note Text": ["Note: ", "Chú ý: "],
	"Result NoSample Text": ["+ N/A: No sample", "+ N/A: Không mẫu"],
	"Result Negative Text": ["+ N: Negative", " + N: Âm tính"],
	"Result LowCopy Text": ["+ P_L: Low copy", "+ P_L: Bản sao thấp"],
	"Result Positive Text": ["+ P_H: Positive", "+ P_H: Dương tính"],
	"Result TechnicianNameSign Text": ["Technician", "Người thực hiện"],
	"Result HeadOfDivisionSign Text": ["Head of Division", "Trưởng bộ phận"],
	"Result SampleName Text": ["Sample name", "Tên mẫu"],
	"Result SamplePosition Text": ["Position", "Vị trí"],
	"Result SCResult Text": ["Spotcheck Result", "Kết quả Spotcheck"],
	"Result GelResult Text": ["Gel Result", "Kết quả gel"],
	"Result FinalResult Text": ["Final Result", "Kết luận"],
	"FinalTitle Label": ["ANALYSIS RESULTS", "KẾT QUẢ PHÂN TÍCH"],

	#---- Only for Tam Viet C----#
	"Result CTResult Text": ["CT Result", "Kết quả CT"],
	"Result NoteTable Text": ["Conversion Table", "Bảng quy đổi kết quả Spotcheck sang kết quả CT"],
	"Result NoteLabel Text": ["Note: ", "Lưu ý: "],
	"Result NoteContent Text": ["Conversion results from Spotcheck to CT are only relative.", "Kết quả quy đổi từ Spotcheck sang CT chỉ mang tính chất tương đối."],
	#---- Only for Tam Viet C----#

	
	##### Messagebox #####
	"Complete Inform": ["COMPLETED", "Hoàn thành"],
	"SaveFile Success": ["File saved", "Đã lưu"],
	"Server Error": ["There was an error while syncing the server", "Có lỗi xảy ra trong quá trình đồng bộ"],
	"Finish Question": ["Do you want to go back to the main screen ?", "Bạn có muốn trở lại màn hình chính ?"]
}

QuantitativeProgramList_Language = {
	"Title Label": ["QUANTITATIVE", "ĐỊNH LƯỢNG"],
	"Information LabelFrame": ["Information", "Thông tin"],
	"KitName Label": ["Kit name :", "Tên Kit :"],
	"Parameters Label": ["Parameters :", "Thông số :"],
	"Back Button": ["Back", "Quay lại"],
	"Choose Button": ["Choose", "Lựa chọn"],

	### Messagebox ###
	"Kit Empty": ["Please select a kit.", "Bạn hãy chọn 1 kit."]
}

Quantitative0_Language = { 
	"Title Label": ["QUANTITATIVE", "ĐỊNH LƯỢNG"],
	"ExperimentName Label": ["File name:", "Tên tệp:"],
	"TechnicianName Label": ["Technician: ", "Người thực hiện: "],
	"TemplateName Label": ["Template: ", "Template: "],
	"Back Button": ["Back", "Quay lại"],
	"Next Button": ["Next", "Kế tiếp"],

	### Messagebox ###
	"Folder Exists": ["This folder already exists, do you want to overwrite it ?", "Thư mục này đã tồn tại, bạn có muốn ghi đè ?"]
}

Quantitative1_Language = {
	"Title Label": ["QUANTITATIVE", "ĐỊNH LƯỢNG"],
	"Information LabelFrame": ["Information", "Thông tin"],
	"KitName Label": ["Kit name :", "Tên Kit :"],
	"Parameters Label": ["Parameters :", "Thông số :"],
	"Back Button": ["Back", "Quay lại"],
	"Next Button": ["Next", "Kế tiếp"],

	### Messagebox ###
	"Kit Empty": ["Please select a kit.", "Bạn hãy chọn 1 kit."]
}

Quantitative2_Language = {
	"Title Label": ["SCREENING", "ĐỊNH LƯỢNG"],
	"SamplesFile LabelFrame": ["Samples file", "Tệp mẫu"],
	"Load Button": ["Load", "Mở"],
	"Create Button": ["Create", " Tạo"],
	"Back Button": ["Back", "Quay lại"],
	"Next Button": ["Next", "Tiếp tục"],
	"OK Button" : ["OK", "Xác nhận"],
	"Cancel Button": ["Cancel", "Huỷ"],

	##### Messagebox #####
	"SamplesFile Empty": ["You haven't loaded the samples file.", "Bạn chưa tải lên tệp mẫu."],
	"Email Confirm": ["Do you want the device to automatically email the results ?", "Bạn có muốn thiết bị tự động email kết quả ?"],
	"Email Empty": ["Please enter recipient email.", "Bạn chưa nhập email người nhận."],
	"AllowPutSample Inform": ["Now you can place the sample on the plate before press [Next].", "Bạn đã có thể cho mẫu vào plate sau đó ấn [Kế tiếp]."],
	"ServerCheck Error": ["There was an error while syncing the server", "Có lỗi xảy ra khi đồng bộ với server"]
}

Quantitative3_Language = {
	"Title Label": ["QUANTITATIVE", "ĐỊNH LƯỢNG"],
	"Process Label": ["Processing...", "Đang xử lý..."],
	"Result Tab": ["Result", "Kết quả"],
	"Report Tab": ["Report", "Báo cáo"],
	"Images Tab": ["Images", "Hình ảnh"],

	"RawImage LabelFrame": ["Raw Image", "Ảnh chụp"],
	"AnalyzedImage LabelFrame": ["Analyzed Image", "Ảnh phân tích"],

	"Finish Button": ["Finish", "Hoàn thành"],
	"Result Title Text": ["ANALYSIS RESULTS", "KẾT QUẢ"],
	"Result ExperimentName Text": ["Test :", "Xét nghiệm :"],
	"Result TechnicianName Text": ["Technician: ", "Người thực hiện: "],
	"Result Date Text": ["Date: ", "Ngày thực hiện: "],

	"Result TechnicianNameSign Text": ["Technician", "Người thực hiện"],
	"Result HeadOfDivisionSign Text": ["Head of Division", "Trưởng bộ phận"],
	"Result SampleName Text": ["Sample name", "Tên mẫu"],
	"Result SamplePosition Text": ["Position", "Vị trí"],
	"Result SCResult Text": ["Spotcheck Result", "Kết quả Spotcheck"],
	"Result GelResult Text": ["Gel Result", "Kết quả gel"],
	"Result FinalResult Text": ["Final Result", "Kết luận"],
	"FinalTitle Label": ["ANALYSIS RESULTS", "KẾT QUẢ PHÂN TÍCH"],
	
	##### Messagebox #####
	"Complete Inform": ["COMPLETED", "Hoàn thành"],
	"SaveFile Success": ["File saved", "Đã lưu"],
	"Server Error": ["There was an error while syncing the server", "Có lỗi xảy ra trong quá trình đồng bộ"],
	"Finish Question": ["Do you want to go back to the main screen ?", "Bạn có muốn trở lại màn hình chính ?"]
}

QuantitativeKit_Language = {
	"Title Label": ["KITS MANAGEMENT", "QUẢN LÝ KIT"],
	"KitName Label": ["Kit name:", "Tên kit:"],
	"Concentration Label": ["Concentration", "Nồng độ"],
	"Value Label": ["Value", "Giá trị"],
	"NValue Label": ["N Value", "Giá trị N"],
	"Save Button": [" Save ", " Lưu "],
	"Delete Button": ["Delete", " Xoá "],
	"Clear Button": ["Clear", " Huỷ "],
	"Back Button": ["Back", "Quay lại"],
	"Analysis Button": ["Analysis", "Phân tích"],

	##### Messagebox #####
	"FileExists Error": ["File already exists", "Tệp đã tồn tại"],
	"FileExists Ask": ["Do you want to overwrite this file ?", "Bạn có muốn ghi đè ?"],
	"SaveFile Success": ["File saved", "Đã lưu"],
	"NValue Empty": ["Please enter [N value]", "Bạn chưa nhập giá trị N"],
	"Concentration Empty": ["At least 3 pairs of [Concentrations & Values] ​are required.", "Cần ít nhất 3 cặp nồng độ và giá trị."],
	"Kit Empty": ["Please enter [Kit name]", "Bạn chưa nhập tên kit !"],
	"Delete Confirm": ["Do you want to delete this file ?", "Bạn có muốn xoá tệp ?"],
	"FileNotExists Error": ["File does not exist", "Tệp không tồn tại"],
	"Delete Done": ["Deleted.", "Đã xoá."]
}

SystemCheck_Language = {
	"Title Label": ["SYSTEM CHECK", "KIỂM TRA HỆ THỐNG"],
	"Process Label": ["Checking background...", "Đang kiểm tra..."],

	##### Messagebox #####
	"CheckAgain Ask": ["Do you want to check again ?", "Bạn có muốn kiểm tra lại ?"],

}
#################################### GUI RULES - END #########################################

#################################### ERROR LIST - START #####################################
class ERROR_LIST(Enum):
	CAMERA_ERROR = 1
	SERIAL_TIMEOUT_ERROR = 2
	SYSTEM_ERROR_1 = 3
	SYSTEM_ERROR_2 = 4
##################################### ERROR LIST - END ######################################

################################### CAMERA PARAMETERS - START ###################################
CAM_FRAMERATE_NUMERATOR = 1
CAM_FRAMERATE_DENOMINATOR = 6
CAM_SENSOR_MODE = 3
CAM_ROTATION = 180
CAM_ISO = 200
CAM_SLEEP_TIME = 2
CAM_SHUTTER_SPEED = 6000000
CAM_EXPOSURE_MODE = "off"

def camera_capture(output):
	camera = PiCamera()
	camera.framerate = Fraction(CAM_FRAMERATE_NUMERATOR, CAM_FRAMERATE_DENOMINATOR)
	camera.sensor_mode = CAM_SENSOR_MODE
	camera.rotation = CAM_ROTATION
	camera.iso = CAM_ISO
	sleep(CAM_SLEEP_TIME)
	camera.shutter_speed = CAM_SHUTTER_SPEED
	camera.exposure_mode = CAM_EXPOSURE_MODE
	camera.capture(output)
	camera.close()
###################################  CAMERA PARAMETERS - END #####################################

##################################### SERIAL CONFIG - START ######################################
ser = serial.Serial(
	port = '/dev/serial0',
	baudrate = 115200,
	parity = serial.PARITY_NONE,
	stopbits = serial.STOPBITS_ONE,
	bytesize = serial.EIGHTBITS,
	timeout = 1
)
###################################### SERIAL CONFIG - END #######################################

######################################## REAL TIME INIT - START #######################################
i2c = board.I2C()
rtc = adafruit_ds1307.DS1307(i2c)
######################################### REAL TIME INIT - END ########################################

######################################### DIRECTORIES INIT - START ##########################################
# ~ BUILD = 1
# ~ if(BUILD==1):
	# ~ dist_dir = os.path.abspath(os.getcwd())
	# ~ spotcheck_dist_dir = os.path.dirname(dist_dir)
	# ~ working_dir = os.path.dirname(spotcheck_dist_dir)
	# ~ parent_dir = os.path.dirname(working_dir)
# ~ else:
	# ~ working_dir = os.path.abspath(os.getcwd())
	# ~ parent_dir = os.path.dirname(working_dir)
working_dir = '/home/pi/Spotcheck'
parent_dir = '/home/pi'

#-------------------- In Spotcheck Fodler -----------------------#
# Programs path
if not os.path.exists(working_dir + "/Programs"):
	f = os.path.join(working_dir + '/', "Programs")
	os.mkdir(f)
programs_path = working_dir + "/Programs/"

# Programs Screening
# ~ if not os.path.exists(programs_path + "Screening"):
	# ~ f = os.path.join(programs_path, "Screening")
	# ~ os.mkdir(f)
# ~ programs_qualitative_path = programs_path + "Screening/"

# Programs Quantitaitve
if not os.path.exists(programs_path + "/Quantitative"):
	f = os.path.join(programs_path + '/', "Quantitative")
	os.mkdir(f)
programs_quantitative_path = programs_path + "Quantitative/"

#------------------------- In Desktop ---------------------------#
# Spotcheck result path
if not os.path.exists(parent_dir + "/Desktop/Spotcheck_Results"):
	f = os.path.join(parent_dir + "/Desktop/", "Spotcheck_Results")
	os.mkdir(f)
results_path = parent_dir + "/Desktop/Spotcheck_Results/"

# ~ # Spotcheck results screening path
if not os.path.exists(results_path + "Screening"):
	f = os.path.join(results_path, "Screening")
	os.mkdir(f)
results_qualitative_path = results_path + "Screening/"

# ~ # Spotcheck results quantitative path
if not os.path.exists(results_path + "Quantitative"):
	f = os.path.join(results_path, "Quantitative")
	os.mkdir(f)
results_quantitative_path = results_path + "Quantitative/"

# Programs results path
# if not os.path.exists(parent_dir + "/Desktop/Spotcheck_Programs"):
# 	f = os.path.join(parent_dir + "/Desktop/", "Spotcheck_Programs")
# 	os.mkdir(f)
# results_programs_path = parent_dir + "/Desktop/Spotcheck_Programs/"

# Programs results screening path
# if not os.path.exists(results_programs_path + "Screening"):
# 	f = os.path.join(results_programs_path, "Screening")
# 	os.mkdir(f)
# results_programs_qualitative_path = results_programs_path + "Screening/"

# Programs results quantitative path
# if not os.path.exists(results_programs_path + "Quantitative"):
# 	f = os.path.join(results_programs_path, "Quantitative")
# 	os.mkdir(f)
# results_programs_quantitative_path = results_programs_path + "Quantitative/"

# Spotcheck ID path
if not os.path.exists(parent_dir + "/Desktop/Spotcheck_ID"):
	f = os.path.join(parent_dir + "/Desktop/", "Spotcheck_ID")
	os.mkdir(f)
id_path = parent_dir + "/Desktop/Spotcheck_ID/"

# Spotchek ID screening path
# if not os.path.exists(id_path + "Screening"):
	# f = os.path.join(id_path, "Screening")
	# os.mkdir(f)
# id_qualitative_path = id_path + "Screening/"

# Spotchek ID quantitative path
# if not os.path.exists(id_path + "Quantitative"):
	# f = os.path.join(id_path, "Quantitative")
	# os.mkdir(f)
# id_quantitative_path = id_path + "Quantitative/"

# Spotcheck ID old path
if not os.path.exists(parent_dir + "/Desktop/Spotcheck_ID/Spotcheck_ID_Old"):
	f = os.path.join(parent_dir + "/Desktop/Spotcheck_ID/", "Spotcheck_ID_Old")
	os.mkdir(f)
id_old_path = parent_dir + "/Desktop/Spotcheck_ID/Spotcheck_ID_Old/"

# System Check Path
if not os.path.exists(parent_dir + "/Desktop/System_Check"):
	f = os.path.join(parent_dir + '/Desktop/', "System_Check")
	os.mkdir(f)
system_check_path = parent_dir + "/Desktop/System_Check/"
########################################## DIRECTORIES INIT - END ###########################################

############################################ FILES INIT - START #############################################
# Check Version
if not os.path.exists(working_dir + "/version.txt"):
	fw = open(working_dir + "/version.txt",'w')
	fw.writelines(["48\n","1\n"])
	fw.close()

fr = open(working_dir + "/version.txt","r")
SC_VERSION = int(fr.readline())
SERIAL_COMUNICATION = int(fr.readline())

if(SC_VERSION == 48):
	WELL_COLUMN = 6
	WELL_ROW = 8
else: 
	WELL_COLUMN = 4
	WELL_ROW = 4

# Cofficient.xlsx
if not os.path.exists(working_dir + "/coefficient.xlsx"):
	wb = Workbook()
	sheet = wb.active
	for r in range(0, WELL_ROW):
		for c in range(0, WELL_COLUMN):
			pos1 = str(chr(66+c)) + str(r+2)    # bat dau tu B2
			pos2 = str(chr(66+c)) + str(r+11)   # bat dau tu B11
			sheet[pos1] = 1
			sheet[pos2] = 20
	wb.save(working_dir + "/coefficient.xlsx")
	wb.close()

coefficient = list(range(SC_VERSION))
base_intensity = list(range(SC_VERSION))
wb = load_workbook(working_dir + "/coefficient.xlsx")
sheet = wb.active
index=0
for r in range(0, WELL_ROW):
	for c in range(0, WELL_COLUMN):
		pos1 = str(chr(66+c)) + str(r+2)    # bat dau tu B2
		pos2 = str(chr(66+c)) + str(r+11)   # bat dau tu B11
		coefficient[index] = float(sheet[pos1].value)
		base_intensity[index] = float(sheet[pos2].value)
		index += 1
index=0
wb.close()

# Coordinates.txt & .coordinates.txt
if not os.path.exists(working_dir + "/coordinates.txt"):
	fw = open(working_dir + "/coordinates.txt",'w')
	fw.writelines(["0\n","0\n", "1\n","1\n"])
	fw.close()
fr = open(working_dir + "/coordinates.txt","r")
x1 = int(fr.readline())
y1 = int(fr.readline())
x2 = int(fr.readline())
y2 = int(fr.readline())

if not os.path.exists(working_dir + "/.coordinates.txt"):
	fw = open(working_dir + "/.coordinates.txt",'w')
	fw.close()

# config.txt
if not os.path.exists(working_dir + "/config.txt"):
	fw = open(working_dir + "/config.txt",'w')
	fw.writelines(["1\n","0\n"])
	fw.close()
fr = open(working_dir + "/config.txt","r")
a = float(fr.readline())
b = float(fr.readline())

# multiplier.txt
if not os.path.exists(working_dir + "/multiplier.txt"):
	fw = open(working_dir + "/multiplier.txt",'w')
	fw.writelines(["1\n","1\n","1\n"])
	fw.close()
fr = open(working_dir + "/multiplier.txt","r")
NUM_1 = float(fr.readline())
NUM_2 = float(fr.readline())
NUM_3 = float(fr.readline())

# .email.txt
if not os.path.exists(working_dir + "/.email.txt"):
	fw = open(working_dir + "/.email.txt",'w')
	fw.writelines(["0\n"])
	fw.close()

# .server.txt
if not os.path.exists(working_dir + "/.server.txt"):
	fw = open(working_dir + "/.server.txt",'w')
	fw.writelines(["0\n"])
	fw.close()    
fr = open(working_dir + "/.server.txt","r")
server_active_0 = int(fr.readline().strip())
fr.close()

# .system.txt
if not os.path.exists(working_dir + "/.system.txt"):
	fw = open(working_dir + "/.system.txt",'w')
	fw.writelines(["01\n","01\n", "24\n","00\n", "00\n", "00\n", "1.11\n"])
	fw.close()
fr = open(working_dir + "/.system.txt","r")
last_checkDay = int(fr.readline())
last_checkMonth = int(fr.readline())
last_checkYear = int(fr.readline())
last_checkHour = int(fr.readline())
last_checkMinute = int(fr.readline())
last_checkSecond = int(fr.readline())
last_checkValue = float(fr.readline())

# .oldinfo.txt
if not os.path.exists('/home/pi/Spotcheck/.oldinfo.txt'):
	fw_info = open('/home/pi/Spotcheck/.oldinfo.txt', 'x')
	fw_info.writelines('@gmail.com\n')
	fw_info.writelines('user\n')
	fw_info.close()
	autofill_email = "@gmail.com"
	autofill_user = "user"
else:
	fr_info = open('/home/pi/Spotcheck/.oldinfo.txt')
	autofill_email = fr_info.readline().strip('\n')
	autofill_user = fr_info.readline().strip('\n')

# active_code.txt
try:
	fr = open(working_dir + "/active_code.txt","r")
	active_code = fr.readline().strip('\n')
except:
	active_code = 'phusa@full'

# .trial_info.txt
try:
	fr = open("/var/tmp/.trial_info.txt","r")
	trial_date = int(fr.readline())
	trial_month = int(fr.readline())
	trial_year = int(fr.readline())
	trial_30days_extend_code = fr.readline().strip('\n')
	trial_full_active_code = fr.readline().strip('\n')
except:
	trial_date = 0
	trial_month = 0
	trial_year = 0
	trial_30days_extend_code = 'phusa@30'
	trial_full_active_code = 'phusa@full'

# language.txt 
if not os.path.exists(working_dir + "/language.txt"):
	fw = open(working_dir + "/language.txt",'w')
	fw.writelines(["0\n"])
	fw.close()
fr = open(working_dir + "/language.txt","r")
language = int(fr.readline())


############################################# FILES INIT - END ##############################################

##################################### ScrollableFrame Class - Start #########################################
class ScrollableFrame(Frame):
	def __init__(self, container, *args, **kwargs):
		super().__init__(container, *args, **kwargs)
		canvas = Canvas(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
		self.scrollable_frame = Frame(canvas, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)

		self.scrollable_frame.bind(
			"<Configure>",
			lambda e: canvas.configure(
				scrollregion=canvas.bbox("all")
			)
		)

		canvas.create_window((0, 0), window=self.scrollable_frame)
		canvas.configure(yscrollcommand=scrollbar.set)
		canvas.pack(side="left", expand=TRUE)
		scrollbar.pack(side="right", fill="y")

class ScrollableFrame1(Frame):
	def __init__(self, container, *args, **kwargs):
		super().__init__(container, *args, **kwargs)
		canvas = Canvas(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR, height=370)
		scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
		self.scrollable_frame = Frame(canvas, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)

		self.scrollable_frame.bind(
			"<Configure>",
			lambda e: canvas.configure(
				scrollregion=canvas.bbox("all")
			)
		)

		canvas.create_window((0, 0), window=self.scrollable_frame)
		canvas.configure(yscrollcommand=scrollbar.set)
		canvas.pack(side="left", expand=TRUE)
		scrollbar.pack(side="right", fill="y")

class ScrollableFrame2(Frame):
	def __init__(self, container, *args, **kwargs):
		super().__init__(container, *args, **kwargs)
		canvas = Canvas(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR, height=370, width=519)
		scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
		self.scrollable_frame = Frame(canvas, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)

		self.scrollable_frame.bind(
			"<Configure>",
			lambda e: canvas.configure(
				scrollregion=canvas.bbox("all")
			)
		)

		canvas.create_window((0, 0), window=self.scrollable_frame)
		canvas.configure(yscrollcommand=scrollbar.set)
		canvas.pack(side="left", expand=TRUE)
		scrollbar.pack(side="right", fill="y")
##################################### ScrollableFrame Class - End #########################################

######################################## EMAIL CLASS - START ###########################################
class AutoMail():
	def __init__(self, sender, password, recipient, subject, content, zip_file, zip_file_name):
		self.SMTP_SERVER = 'smtp.gmail.com'
		self.SMTP_PORT = 587
		self.sender = sender
		self.password = password
		self.recipient = recipient
		self.subject = subject
		self.content = content
		self.zip_file = zip_file
		self.zip_file_name = zip_file_name

	def send(self):
		emailData = MIMEMultipart()
		emailData['Subject'] = self.subject
		emailData['To'] = self.recipient
		emailData['From'] = self.sender

		emailData.attach(MIMEText(self.content))

	#     imageData = MIMEImage(open(image, 'rb').read(), 'jpg')
	#     imageData.add_header('Content-Disposition', 'attachment; filename="image.jpg"')
	#     emailData.attach(imageData)

		with open(self.zip_file,'rb') as file:
			# emailData.attach(MIMEApplication(file.read(), Name= self.zip_file_name + '.zip'))
			emailData.attach(MIMEApplication(file.read(), Name= self.zip_file_name + '.xlsx'))

		session = smtplib.SMTP(self.SMTP_SERVER, self.SMTP_PORT)
		session.ehlo()
		session.starttls()
		session.ehlo()

		#session.login(mail_address, password)
		session.login(self.sender, self.password)

		session.sendmail(self.sender, self.recipient.split(','), emailData.as_string())
		session.quit
######################################### EMAIL CLASS - END ############################################

######################################### IMAGE PROCESSING CLASS - START ############################################
class Process_Image():
	def __init__(self, image_path, start_point=(x1,y1), end_point=(x2,y2)):
		self.image = cv2.imread(image_path)
		self.start_point = start_point
		self.end_point = end_point

	def process(self, coefficient, mode=0, well_list=[]):
		blur_img = cv2.GaussianBlur(self.image.copy(), (35,35), 0)
		gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)

		thresh, binary_img = cv2.threshold(gray_img.copy(), 30, maxval=255, type=cv2.THRESH_BINARY)
		contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
		print("Number of contours: " + str(len(contours)))

		contours = sorted(contours, key=lambda data:self.sorting_xy(data))

		if(mode==0):
			try:
				bounding_rect_first = cv2.boundingRect(contours[0])
				bounding_rect_last = cv2.boundingRect(contours[len(contours)-1])
				new_start_point = (bounding_rect_first[0] - 12, bounding_rect_first[1] - 12)
				new_end_point = (bounding_rect_last[0] + bounding_rect_last[2] + 12, bounding_rect_last[1] + bounding_rect_last[3] + 12)
				print('New start point:', new_start_point)
				print('New end point:', new_end_point)
				fw= open(working_dir + '/.coordinates.txt','w')
				fw.writelines("Start Point: " + str(new_start_point) + "\n")
				fw.writelines("End Point: " + str(new_end_point))
			except:
				print("Can't find new coordinates")
				pass

		contour_img = np.zeros_like(gray_img)
		contour_img = cv2.rectangle(contour_img, self.start_point, self.end_point, (255,255,255), -1)
		rect_w = self.end_point[0] - self.start_point[0]
		rect_h = self.end_point[1] - self.start_point[1]
		cell_w = round(rect_w / WELL_COLUMN)
		cell_h = round(rect_h / WELL_ROW)
		for i in range(1, WELL_COLUMN):
			contour_img = cv2.line(contour_img, (self.start_point[0] + i*cell_w, self.start_point[1]), (self.start_point[0] + i*cell_w, self.end_point[1]),(0,0,0), 4)
		for i in range(1, WELL_ROW):
			contour_img = cv2.line(contour_img, (self.start_point[0], self.start_point[1] + i*cell_h), (self.end_point[0], self.start_point[1] + i*cell_h),(0,0,0), 4)

		thresh1 , binary1_img = cv2.threshold(contour_img, 250, maxval=255, type=cv2.THRESH_BINARY)
		contours1, hierarchy1 = cv2.findContours(binary1_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

		contours1 = sorted(contours1, key=lambda data:self.sorting_y(data))
		if(SC_VERSION == 48):
			contours1_h1 = contours1[0:6]
			contours1_h2 = contours1[6:12]
			contours1_h3 = contours1[12:18]
			contours1_h4 = contours1[18:24]
			contours1_h5 = contours1[24:30]
			contours1_h6 = contours1[30:36]
			contours1_h7 = contours1[36:42]
			contours1_h8 = contours1[42:48]
			contours1_h1 = sorted(contours1_h1, key=lambda data:self.sorting_x(data))
			contours1_h2 = sorted(contours1_h2, key=lambda data:self.sorting_x(data))
			contours1_h3 = sorted(contours1_h3, key=lambda data:self.sorting_x(data))
			contours1_h4 = sorted(contours1_h4, key=lambda data:self.sorting_x(data))
			contours1_h5 = sorted(contours1_h5, key=lambda data:self.sorting_x(data))
			contours1_h6 = sorted(contours1_h6, key=lambda data:self.sorting_x(data))
			contours1_h7 = sorted(contours1_h7, key=lambda data:self.sorting_x(data))
			contours1_h8 = sorted(contours1_h8, key=lambda data:self.sorting_x(data))
			sorted_contours1 = contours1_h1 + contours1_h2 + contours1_h3 + contours1_h4 + contours1_h5 + contours1_h6 + contours1_h7 + contours1_h8
		else:
			contours1_h1 = contours1[0:4]
			contours1_h2 = contours1[4:8]
			contours1_h3 = contours1[8:12]
			contours1_h4 = contours1[12:16]
			contours1_h1 = sorted(contours1_h1, key=lambda data:self.sorting_x(data))
			contours1_h2 = sorted(contours1_h2, key=lambda data:self.sorting_x(data))
			contours1_h3 = sorted(contours1_h3, key=lambda data:self.sorting_x(data))
			contours1_h4 = sorted(contours1_h4, key=lambda data:self.sorting_x(data))
			sorted_contours1 = contours1_h1 + contours1_h2 + contours1_h3 + contours1_h4
		
		list_intensities = []
		sum_intensities = []
		result_list = list(range(SC_VERSION))
		area = list(range(SC_VERSION))

		blur1_img = cv2.GaussianBlur(self.image.copy(), (25,25), 0)
		tmp_list = list(range(SC_VERSION))
		list_bgrvalue = []
		list_index = list(range(SC_VERSION))
		for i in range(len(sorted_contours1)):
			list_index[i] = []
			cimg = np.zeros_like(gray_img)
			cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
			pts = np.where(cimg == 255)
			list_bgrvalue.append(blur1_img[pts[0], pts[1]])
			for j in range(len(list_bgrvalue[i])):
				 list_index[i].append(round((list_bgrvalue[i][j][1]*3 + list_bgrvalue[i][j][2])))
			list_index[i].sort()
			list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
			area[i]= cv2.contourArea(sorted_contours1[i])
			tmp_list[i] = list_intensities[i]/1000
			result_list[i] = round(tmp_list[i],1)

		# for i in range(1,7):
		# 	result_list[6*i+1]=round(result_list[6*i+1]*(1-(0.02*round(result_list[6*i]/70) + 0.02*round(result_list[6*i+2]/70) + 0.02*round(result_list[6*i-5]/70) + 0.02*round(result_list[6*i+7]/70) + 0.003*round(result_list[6*i-6]/70) + 0.003*round(result_list[6*i-4]/70) + 0.003*round(result_list[6*i+6]/70) + 0.003*round(result_list[6*i+8]/70)+ 0.006*round(result_list[6*i+3]/70))),1)
		# 	result_list[6*i+2]=round(result_list[6*i+2]*(1-(0.02*round(result_list[6*i+1]/70) + 0.02*round(result_list[6*i+3]/70) + 0.02*round(result_list[6*i-4]/70) + 0.02*round(result_list[6*i+8]/70) + 0.003*round(result_list[6*i-5]/70) + 0.003*round(result_list[6*i-3]/70) + 0.003*round(result_list[6*i+7]/70) + 0.003*round(result_list[6*i+9]/70)+ 0.006*round(result_list[6*i+4]/70)+ 0.006*round(result_list[6*i]/70))),1)
		# 	result_list[6*i+3]=round(result_list[6*i+3]*(1-(0.02*round(result_list[6*i+2]/70) + 0.02*round(result_list[6*i+4]/70) + 0.02*round(result_list[6*i-3]/70) + 0.02*round(result_list[6*i+9]/70) + 0.003*round(result_list[6*i-4]/70) + 0.003*round(result_list[6*i-2]/70) + 0.003*round(result_list[6*i+8]/70) + 0.003*round(result_list[6*i+10]/70)+ 0.006*round(result_list[6*i+5]/70)+ 0.006*round(result_list[6*i+1]/70))),1)
		# 	result_list[6*i+4]=round(result_list[6*i+4]*(1-(0.02*round(result_list[6*i+3]/70) + 0.02*round(result_list[6*i+5]/70) + 0.02*round(result_list[6*i-2]/70) + 0.02*round(result_list[6*i+10]/70) + 0.003*round(result_list[6*i-3]/70) + 0.003*round(result_list[6*i-1]/70) + 0.003*round(result_list[6*i+9]/70) + 0.003*round(result_list[6*i+11]/70)+ 0.006*round(result_list[6*i+2]/70))),1)

		# 	result_list[6*i]=round(result_list[6*i]*(1-(0.02*round(result_list[6*i+1]/70) + 0.015*round(result_list[6*i-6]/70) + 0.015*round(result_list[6*i+6]/70) + 0.003*round(result_list[6*i-5]/70) + 0.003*round(result_list[6*i+7]/70)+0.006*round(result_list[6*i+2]/70))),1)
		# 	result_list[6*i+5]=round(result_list[6*i+5]*(1-(0.02*round(result_list[6*i+4]/70) + 0.015*round(result_list[6*i-1]/70) + 0.015*round(result_list[6*i+11]/70) + 0.003*round(result_list[6*i-2]/70) + 0.003*round(result_list[6*i+10]/70)+0.006*round(result_list[6*i+3]/70))),1)

		# for i in range(1,5):
		# 	result_list[i]=round(result_list[i]*(1-(0.02*round(result_list[i-1]/70) + 0.02*round(result_list[i+1]/70) + 0.015*round(result_list[i+6]/70)+ 0.003*round(result_list[i+5]/70) + 0.003*round(result_list[i+7]/70))),1)
		# 	result_list[i+42]=round(result_list[i+42]*(1-(0.02*round(result_list[i+41]/70) + 0.02*round(result_list[i+43]/70) + 0.015*round(result_list[i+36]/70)+ 0.003*round(result_list[i+35]/70) + 0.003*round(result_list[i+37]/70))),1)

		# result_list[0]=round(result_list[0]*(1-(0.015*round(result_list[1]/70) + 0.015*round(result_list[6]/70))),1)
		# result_list[5]=round(result_list[5]*(1-(0.015*round(result_list[4]/70) + 0.015*round(result_list[11]/70))),1)
		# result_list[42]=round(result_list[42]*(1-(0.015*round(result_list[43]/70) + 0.015*round(result_list[36]/70))),1)
		# result_list[47]=round(result_list[47]*(1-(0.015*round(result_list[46]/70) + 0.015*round(result_list[41]/70))),1)

		for i in range(len(sorted_contours1)):
			result_list[i] = round(result_list[i]*coefficient[i],1)

		for i in range(len(sorted_contours1)):
			if(result_list[i]>99):
				result_list[i]=99

		for i in range(len(sorted_contours1)):
			if ((i!=0) and ((i+1)%6==0)):
				print('%.1f'%(result_list[i]))
			else:
				print('%.1f'%(result_list[i]), end = ' | ')

		blurori_img = cv2.GaussianBlur(self.image.copy(), (25,25), 0)
		if(mode):
			for i in range(len(sorted_contours1)):
				if(well_list[i]=='N/A'):
					cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,0), thickness = -1)
				else:
					cv2.drawContours(blurori_img, sorted_contours1, i, (255,255,0), thickness = 2)
					# if(well_list[i] < float(thr_set)):
					#     cv2.drawContours(blurori_img, sorted_contours1, i, (0,255,0), thickness = 2)
					# else:
					#     cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,255), thickness = 2)
		else:
			for i in range(len(sorted_contours1)):
				cv2.drawContours(blurori_img, sorted_contours1, i, (255,255,0), thickness = 2)
		return (result_list, blurori_img)

	def sorting_y(self, contour):
		rect_y = cv2.boundingRect(contour)
		return rect_y[1]
	def sorting_x(self, contour):
		rect_x = cv2.boundingRect(contour)
		return rect_x[0]
	def sorting_xy(self, contour):
		rect_xy = cv2.boundingRect(contour)
		return math.sqrt(math.pow(rect_xy[0],2) + math.pow(rect_xy[1],2))
########################################## IMAGE PROCESSING CLASS - END #############################################

############################################## GUI DESIGN _ START ################################################
class SystemCheckFrame(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.err = 0
		self.mode_check = 0 # 0 --> first check when open app -> back to main menu if check ok
							# 1 --> check when start analysis -> back to qualitative analysis 0 if check ok
							# 2 --> check when start analysis -> back to quantitative analysis 0 if check ok
							
		# 3 main frame
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = SystemCheck_Language["Title Label"][language],
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame

		# In button frame
		# ~ self.back_button = Button(self.button_frame,
								# ~ text = "Back",
								# ~ font = SWITCH_PAGE_BUTTON_FONT,
								# ~ # width = SWITCH_PAGE_BUTTON_WIDTH,
								# ~ # height = SWITCH_PAGE_BUTTON_HEIGHT,
								# ~ bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								# ~ fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								# ~ borderwidth = 0,
								# ~ command = self.back_clicked)
		# ~ self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

		# ~ self.next_button = Button(self.button_frame,
								# ~ text = "Next",
								# ~ font = SWITCH_PAGE_BUTTON_FONT,
								# ~ bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								# ~ fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								# ~ borderwidth = 0,
								# ~ command = self.next_clicked)
		# ~ self.next_button.pack(ipadx=30, ipady=10, side=RIGHT)

	def back_clicked(self):
		try:
			self.check_result_frame.destroy()
		except:
			pass

		try:
			self.progressbar.pack_forget()
			self.process_label.pack_forget()
		except:
			pass

		self.base_window.forget_page()

		if(self.mode_check == 0):
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_1)
		elif(self.mode_check == 1):
			#self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_1)
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
		elif(self.mode_check == 2):
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_0)
		self.base_window.switch_page()

	def next_clicked(self):
		if(self.err == 2):
			msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SYSTEM_ERROR_2'].value),
												ERROR_LIST(ERROR_LIST['SYSTEM_ERROR_2'].value).name,
												icon = "error")
		else:
			print("mode_check:", self.mode_check)
			if(self.mode_check == 0):
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
				self.base_window.switch_page()
			elif(self.mode_check == 1):
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
				self.base_window.switch_page()
			elif(self.mode_check == 2):
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_0)
				self.base_window.switch_page()

	def serial_handle(self):
		try:
			self.check_result_frame.destroy()
		except:
			pass
		
		style = ttk.Style()
		style.theme_use('clam')
		style.configure("1.Horizontal.TProgressbar", troughcolor ='grey85', background='green3')
		self.progressbar = ttk.Progressbar(self.work_frame,
									style="1.Horizontal.TProgressbar",
									length = 200,
									mode = 'determinate')
		self.progressbar.pack(ipadx=2, ipady=2)

		self.process_label = Label(self.work_frame,
						bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
						fg = 'black',
						text = SystemCheck_Language["Process Label"][language],
						font = LABEL_TXT_FONT)
		self.process_label.pack(ipadx=2, ipady=2, anchor=N)


		if(SERIAL_COMUNICATION):
			ser.flushInput()
			ser.flushOutput()

			self.progressbar['value']=5
			self.base_window.update_idletasks()
			sleep(0.5)

			data_send = 'P'
			print("Data send:", data_send)
			ser.write(data_send.encode())

			receive_data = StringVar()
			count = 0
			bled_ready = 0
			while(receive_data != 'C'):
				if(ser.in_waiting>0):
					receive_data = ser.readline().decode('utf-8').rstrip()
					print("Data received:", receive_data)

					self.progressbar['value']=10
					self.base_window.update_idletasks()
					sleep(0.5)

					if(receive_data == 'C'):
						self.progressbar['value']=20
						self.base_window.update_idletasks()
						sleep(0.5)

						bled_ready = 1
						break;
				else:
					sleep(1)
					count += 1
					if(count > 15):
						msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value),
													ERROR_LIST(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value).name,
													icon = "error")
						break;
		else:
			bled_ready = 1

		if(bled_ready):
			GPIO.output(BLUELIGHT_PIN, GPIO.HIGH)
			
			self.progressbar['value']=45
			self.base_window.update_idletasks()
			print("self.mode_check: ", self.mode_check)

			try:
				camera_capture(system_check_path + 'system_check_raw.jpg')
				self.progressbar['value']=65
				self.base_window.update_idletasks()
			except Exception as e :
				msg = messagebox.showerror("ERR "+ str(ERROR_LIST['CAMERA_ERROR'].value),
										ERROR_LIST(ERROR_LIST['CAMERA_ERROR'].value).name,
										icon = "error")
				if(msg=='ok'):
					global number_of_instance
					number_of_instance = 0
					fw_instance = open('/home/pi/Spotcheck/.instance.txt', 'w')
					fw_instance.writelines('0\n')
					fw_instance.close()
					os._exit(0)
					self.base_window.destroy()

			if(self.mode_check==0):
				self.result, self.image = Process_Image(system_check_path + 'system_check_raw.jpg').process(coefficient)
				cv2.imwrite(system_check_path + 'system_check_process.jpg', self.image)
			elif(self.mode_check==1):
				self.result, self.image = Process_Image(system_check_path + 'system_check_raw.jpg').process(coefficient)
				cv2.imwrite(system_check_path + 'system_check_process.jpg', self.image)
			elif(self.mode_check==2):
				self.result, self.image = Process_Image(system_check_path + 'system_check_raw.jpg').process(coefficient)
				cv2.imwrite(system_check_path + 'system_check_process.jpg', self.image)

			self.progressbar['value']=80
			self.base_window.update_idletasks()
			sleep(0.5)

			self.progressbar['value']=95
			self.base_window.update_idletasks()
			sleep(0.5)
			
			GPIO.output(BLUELIGHT_PIN, GPIO.LOW)

			wb = Workbook()
			sheet = wb.active

			if(SC_VERSION == 48):
				sheet["A2"] = "A"
				sheet["A3"] = "B"
				sheet["A4"] = "C"
				sheet["A5"] = "D"
				sheet["A6"] = "E"
				sheet["A7"] = "F"
				sheet["A8"] = "G"
				sheet["A9"] = "H"
				sheet["B1"] = "1"
				sheet["C1"] = "2"
				sheet["D1"] = "3"
				sheet["E1"] = "4"
				sheet["F1"] = "5"
				sheet["G1"] = "6"
			else:
				sheet["A2"] = "A"
				sheet["A3"] = "B"
				sheet["A4"] = "C"
				sheet["A5"] = "D"
				sheet["B1"] = "1"
				sheet["C1"] = "2"
				sheet["D1"] = "3"
				sheet["E1"] = "4"
			
			index = 0
			for r in range(0, WELL_ROW):
				for c in range(0, WELL_COLUMN):
					pos = str(chr(66+c)) + str(r+2)
					sheet[pos] = self.result[index]
					index += 1

			self.average_current_intensity = round(sum(self.result)/len(self.result), 1)
			self.threshold = round(self.average_current_intensity * a + b, 1)
			sheet['I2'] = "Average: " + str(self.average_current_intensity)
			sheet['I3'] = "Threshold: " + str(self.threshold)

			wb.save(system_check_path + "nonsample_value.xlsx")
			wb.close()

			self.progressbar['value']=100
			self.base_window.update_idletasks()
			sleep(0.5)

			self.progressbar.destroy()
			self.process_label.destroy()

			average_base_intensity = round(sum(base_intensity)/len(base_intensity), 1)
			tmp_value = round(average_base_intensity/self.average_current_intensity, 2)

			self.check_result_frame = Frame(self.work_frame, bg=RESULT_TABLE_FRAME_BGD_COLOR)
			self.check_result_frame.pack()

			result_label = list(range(SC_VERSION))
			index = 0
			for r in range(0, WELL_ROW):
				for c in range(0, WELL_COLUMN):
					t = str(chr(65 + c)) + str(index + 1)
					result_label[index] = Label(self.check_result_frame,
										text = t,
										width = 6,
										height = 3,
										bg = RESULT_LABEL_BGD_COLOR,
										font = RESULT_LABEL_TXT_FONT)
					result_label[index].grid(row=r,column=c, padx=2, pady=2)
					
					# Check system error 1
					if(self.result[index] > base_intensity[index] + base_intensity[index]*30/100 or self.result[index] < base_intensity[index] - base_intensity[index]*30/100):
						result_label[index]['bg'] = RESULT_LABEL_ERROR_BGD_COLOR
						self.err = 1

					index += 1

			# Check system error 2
			if(self.average_current_intensity > average_base_intensity + average_base_intensity*30/100 or
				self.average_current_intensity < average_base_intensity - average_base_intensity*30/100):
					for i in range(0, SC_VERSION):
						result_label[i]['bg'] = RESULT_LABEL_ERROR_BGD_COLOR
					self.err = 2

			# Save time and value check
			now = datetime.now()
			time1 = last_checkYear*31536000 + last_checkMonth*2419200 + last_checkDay*86400 + last_checkHour*3600 + last_checkMinute*60 + last_checkSecond
			time2 = now.year*31536000 + last_checkMonth*2419200 + last_checkDay*86400 + last_checkHour*3600 + last_checkMinute*60 + last_checkSecond  
			fw = open(working_dir + "/.system.txt",'w')
			fw.writelines([str(now.day) + "\n",str(now.month) + "\n", str(now.year) + "\n", str(now.hour) + "\n", str(now.minute) + "\n", str(now.second) + "\n", str(self.threshold) + "\n"])
			fw.close()

			# Error handle
			if(self.err == 1):
				msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SYSTEM_ERROR_1'].value),
										ERROR_LIST(ERROR_LIST['SYSTEM_ERROR_1'].value).name,
										icon = "error")
				err_msg = messagebox.askquestion("", SystemCheck_Language["CheckAgain Ask"][language])
				if(err_msg == "yes"):
					self.check_result_frame.destroy()
					self.base_window.system_check.mode_check = 0
					self.base_window.forget_page()
					self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
					self.base_window.switch_page()
					self.base_window.update_idletasks()
					self.base_window.system_check.serial_handle()
				else:
					self.next_clicked()

			elif(self.err == 2):
				msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SYSTEM_ERROR_2'].value),
										ERROR_LIST(ERROR_LIST['SYSTEM_ERROR_2'].value).name,
										icon = "error")
				err_msg = messagebox.askquestion("", SystemCheck_Language["CheckAgain Ask"][language])
				if(err_msg == "yes"):
					self.check_result_frame.destroy()
					self.base_window.system_check.mode_check = 0
					self.base_window.forget_page()
					self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
					self.base_window.switch_page()
					self.base_window.update_idletasks()
					self.base_window.system_check.serial_handle()
				else:
					number_of_instance = 0
					fw_instance = open('/home/pi/Spotcheck/.instance.txt', 'w')
					fw_instance.writelines('0\n')
					fw_instance.close()
					os._exit(0)
					self.base_window.destroy()
			else:
				# ~ msg = messagebox.showinfo("", "Finished checking !")
				self.next_clicked()

		else:
			# ~ self.back_clicked()
			self.base_window.destroy()


class NewQualitativeFrame3(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		# 3 main frame
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = "CALIBRATION",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

	def serial_handle(self):
		style = ttk.Style()
		style.theme_use('clam')
		style.configure("1.Horizontal.TProgressbar", troughcolor ='grey85', background='green3')
		self.progressbar = ttk.Progressbar(self.work_frame,
									style="1.Horizontal.TProgressbar",
									length = 200,
									mode = 'determinate')
		self.progressbar.pack(ipadx=2, ipady=2)

		self.process_label = Label(self.work_frame,
						bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
						fg = 'black',
						text = 'Processing...',
						font = LABEL_TXT_FONT)
		self.process_label.pack(ipadx=2, ipady=2, anchor=N)

		ser.flushInput()
		ser.flushOutput()

		self.progressbar['value']=5
		self.base_window.update_idletasks()
		sleep(0.5)

		data_send = 'P'
		print("Data send:", data_send)
		ser.write(data_send.encode())

		receive_data = StringVar()
		count = 0
		bled_ready = 0
		while(receive_data != 'C'):
			if(ser.in_waiting>0):
				receive_data = ser.readline().decode('utf-8').rstrip()
				print("Data received:", receive_data)

				self.progressbar['value']=10
				self.base_window.update_idletasks()
				sleep(0.5)

				if(receive_data == 'C'):
					self.progressbar['value']=20
					self.base_window.update_idletasks()
					sleep(0.5)

					bled_ready = 1
					break;
			else:
				sleep(1)
				count += 1
				if(count > 15):
					msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value),
												ERROR_LIST(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value).name,
												icon = "error")
					break;

		if(bled_ready):
			self.progressbar['value']=45
			self.base_window.update_idletasks()
			try:
				camera_capture(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + '/raw.jpg')

				self.progressbar['value']=65
				self.base_window.update_idletasks()
			except Exception as e :
				msg = messagebox.showerror("ERR "+ str(ERROR_LIST['CAMERA_ERROR'].value),
										ERROR_LIST(ERROR_LIST['CAMERA_ERROR'].value).name,
										icon = "error")
				if(msg=='ok'):
					self.base_window.destroy()

			self.result, self.image = Process_Image(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + '/raw.jpg').process(coefficient, mode=1, well_list=self.base_window.new_qualitative_2.well_button)

			self.progressbar['value']=80
			self.base_window.update_idletasks()
			sleep(0.5)

			cv2.imwrite(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + '/process.jpg', self.image)

			self.progressbar['value']=95
			self.base_window.update_idletasks()
			sleep(0.5)

			sum_value = 0
			active_well_number = 0

			# Save analysis file
			wb = Workbook()
			sheet = wb.active
			sheet["A2"] = "A"
			sheet["A3"] = "B"
			sheet["A4"] = "C"
			sheet["A5"] = "D"
			sheet["A6"] = "E"
			sheet["A7"] = "F"
			sheet["A8"] = "G"
			sheet["A9"] = "H"
			sheet["B1"] = "1"
			sheet["C1"] = "2"
			sheet["D1"] = "3"
			sheet["E1"] = "4"
			sheet["F1"] = "5"
			sheet["G1"] = "6"
			for i in range(0,48):
				if(i<6):
					pos = str(chr(65+i+1)) + "2"
				if(i>=6 and i<12):
					pos = str(chr(65+i-5)) + "3"
				if(i>=12 and i<18):
					pos = str(chr(65+i-11)) + "4"
				if(i>=18 and i<24):
					pos = str(chr(65+i-17)) + "5"
				if(i>=24 and i<30):
					pos = str(chr(65+i-23)) + "6"
				if(i>=30 and i<36):
					pos = str(chr(65+i-29)) + "7"
				if(i>=36 and i<42):
					pos = str(chr(65+i-35)) + "8"
				if(i>=42):
					pos = str(chr(65+i-41)) + "9"
				if(self.base_window.new_qualitative_2.well_button[i]['text'][0] != '#'):
					active_well_number += 1
					sum_value += round(self.result[i]/self.base_window.system_check.threshold,2)
					sheet[pos] = round(self.result[i]/self.base_window.system_check.threshold,2)
					print(self.result[i])
				else:
					sheet[pos] = "N/A"
			wb.save(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + "/analysis_value.xlsx")
			wb.close()

			self.final_value = round(sum_value/active_well_number,2)
			print("active_well_number:", active_well_number)
			print("sum_value:", sum_value)
			print("final_value:", self.final_value)

			# Save program file
			wb = Workbook()
			sheet = wb.active
			sheet["B2"] = "User name:"
			sheet["B3"] = "Comment:"
			sheet["D5"] = "A"
			sheet["D6"] = "B"
			sheet["D7"] = "C"
			sheet["D8"] = "D"
			sheet["D9"] = "E"
			sheet["D10"] = "F"
			sheet["D11"] = "G"
			sheet["D12"] = "H"
			sheet["E4"] = "1"
			sheet["F4"] = "2"
			sheet["G4"] = "3"
			sheet["H4"] = "4"
			sheet["I4"] = "5"
			sheet["J4"] = "6"

			sheet["C2"] = self.base_window.new_qualitative_1.user_name
			sheet["C3"] = self.base_window.new_qualitative_1.comments

			sheet["D2"] = "Base Value:"
			sheet["E2"] = self.final_value

			count = -1
			for i in range(0,48):
				if(i<6):
					pos = str(chr(68+i+1)) + "5"
				if(i>=6 and i<12):
					pos = str(chr(68+i-5)) + "6"
				if(i>=12 and i<18):
					pos = str(chr(68+i-11)) + "7"
				if(i>=18 and i<24):
					pos = str(chr(68+i-17)) + "8"
				if(i>=24 and i<30):
					pos = str(chr(68+i-23)) + "9"
				if(i>=30 and i<36):
					pos = str(chr(68+i-29)) + "10"
				if(i>=36 and i<42):
					pos = str(chr(68+i-35)) + "11"
				if(i>=42):
					pos = str(chr(68+i-41)) + "12"

				if(self.base_window.new_qualitative_2.well_button[i]['text'][0] != '#'):
					count += 1
					sheet[pos] = self.base_window.new_qualitative_2.well_button[i]['text']
					sheet["B" + str(4 + count)] = self.base_window.new_qualitative_2.well_button[i]['text']
					sheet["C" + str(4 + count)] = round(self.result[i]/self.base_window.system_check.threshold,2)
				else:
					sheet[pos] = "N/A"
			wb.save(programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + ".xlsx")
			wb.close()

			self.progressbar['value']=100
			self.base_window.update_idletasks()
			sleep(0.5)

			self.progressbar.destroy()
			self.process_label.destroy()

			self.check_result_frame = Frame(self.work_frame, bg=RESULT_TABLE_FRAME_BGD_COLOR)
			self.check_result_frame.pack(side=LEFT, anchor=E)

			result_label = list(range(48))
			r=0
			c=-1
			for i in range(0,48):
				c+=1
				if(c>5):
					c=0
					r+=1
				result_label[i] = Label(self.check_result_frame,
										width=6,
										height=3,
										# ~ bg = RESULT_LABEL_BGD_COLOR,
										font = RESULT_LABEL_TXT_FONT)
				if(self.base_window.new_qualitative_2.well_button[i]['text'][0] != '#'):
					result_label[i]['text'] = round(self.result[i]/self.base_window.system_check.threshold,2)
					result_label[i]['bg'] = "cyan"
				else:
					result_label[i]['text'] = "N/A"
					result_label[i]['bg'] = "grey80"
				result_label[i].grid(row=r,column=c, padx=2, pady=2)

			self.pfi_label = Label(self.work_frame,
									bg = LABEL_BGD_COLOR,
									text = 'PFi value: ' + str(self.final_value),
									fg = 'blue',
									font = ("Helvetica", 13, 'bold'))
			self.pfi_label.pack(side=LEFT, ipadx=5, ipady=5, padx=100, pady=50, anchor=NE, expand=TRUE)

			self.title_label['text'] = "CALIBRATION RESULT"

			# In button frame
			self.finish_button = Button(self.button_frame,
									text = "FINISH",
									font = SWITCH_PAGE_BUTTON_FONT,
									# width = SWITCH_PAGE_BUTTON_WIDTH,
									# height = SWITCH_PAGE_BUTTON_HEIGHT,
									bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
									fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.finish_clicked)
			self.finish_button.pack(ipady=10)

			msg = messagebox.showinfo("","COMPLETED")

		else:
			self.back_clicked()

	def finish_clicked(self):
		msg = messagebox.askquestion("","Do you want to go back to the main screen ?")
		if(msg=="yes"):
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_option)
			# ~ self.base_window.page_num = PAGE_LIST.index('MAIN_MENU')
			self.base_window.switch_page()
			self.base_window.main_menu.reset()

class NewQuantitativeFrame3(NewQualitativeFrame3):
	def __init__(self, container):
		super().__init__(container)
		self.title_label['text'] = "CALIBRATION"

	def serial_handle(self):
		style = ttk.Style()
		style.theme_use('clam')
		style.configure("1.Horizontal.TProgressbar", troughcolor ='grey85', background='green3')
		self.progressbar = ttk.Progressbar(self.work_frame,
									style="1.Horizontal.TProgressbar",
									length = 200,
									mode = 'determinate')
		self.progressbar.pack(ipadx=2, ipady=2)

		self.process_label = Label(self.work_frame,
						bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
						fg = 'black',
						text = 'Processing...',
						font = LABEL_TXT_FONT)
		self.process_label.pack(ipadx=2, ipady=2, anchor=N)

		ser.flushInput()
		ser.flushOutput()

		self.progressbar['value']=5
		self.base_window.update_idletasks()
		sleep(0.5)

		data_send = 'P'
		print("Data send:", data_send)
		ser.write(data_send.encode())

		receive_data = StringVar()
		count = 0
		bled_ready = 0
		while(receive_data != 'C'):
			if(ser.in_waiting>0):
				receive_data = ser.readline().decode('utf-8').rstrip()
				print("Data received:", receive_data)

				self.progressbar['value']=10
				self.base_window.update_idletasks()
				sleep(0.5)

				if(receive_data == 'C'):
					self.progressbar['value']=20
					self.base_window.update_idletasks()
					sleep(0.5)

					bled_ready = 1
					break;
			else:
				sleep(1)
				count += 1
				if(count > 15):
					msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value),
												ERROR_LIST(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value).name,
												icon = "error")
					break;

		if(bled_ready):
			self.progressbar['value']=45
			self.base_window.update_idletasks()
			try:
				camera_capture(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + '/raw.jpg')

				self.progressbar['value']=65
				self.base_window.update_idletasks()
			except Exception as e :
				msg = messagebox.showerror("ERR "+ str(ERROR_LIST['CAMERA_ERROR'].value),
										ERROR_LIST(ERROR_LIST['CAMERA_ERROR'].value).name,
										icon = "error")
				if(msg=='ok'):
					self.base_window.destroy()

			self.result, self.image = Process_Image(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + '/raw.jpg').process(coefficient, mode=1, well_list=self.base_window.new_quantitative_2.well_button)

			self.progressbar['value']=80
			self.base_window.update_idletasks()
			sleep(0.5)

			cv2.imwrite(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + '/process.jpg', self.image)

			self.progressbar['value']=95
			self.base_window.update_idletasks()
			sleep(0.5)

			sum_value = 0
			active_well_number = 0

			sum_concen_0 = 0
			sum_concen_1 = 0
			sum_concen_2 = 0
			sum_concen_3 = 0
			sum_concen_4 = 0
			sum_concen_5 = 0
			sum_concen_6 = 0
			sum_concen_7 = 0
			sum_concen_8 = 0
			sum_concen_9 = 0

			concen_0_number = 0
			concen_1_number = 0
			concen_2_number = 0
			concen_3_number = 0
			concen_4_number = 0
			concen_5_number = 0
			concen_6_number = 0
			concen_7_number = 0
			concen_8_number = 0
			concen_9_number = 0


			# Save analysis file
			wb = Workbook()
			sheet = wb.active
			sheet["A2"] = "A"
			sheet["A3"] = "B"
			sheet["A4"] = "C"
			sheet["A5"] = "D"
			sheet["A6"] = "E"
			sheet["A7"] = "F"
			sheet["A8"] = "G"
			sheet["A9"] = "H"
			sheet["B1"] = "1"
			sheet["C1"] = "2"
			sheet["D1"] = "3"
			sheet["E1"] = "4"
			sheet["F1"] = "5"
			sheet["G1"] = "6"
			for i in range(0,48):
				if(i<6):
					pos = str(chr(65+i+1)) + "2"
				if(i>=6 and i<12):
					pos = str(chr(65+i-5)) + "3"
				if(i>=12 and i<18):
					pos = str(chr(65+i-11)) + "4"
				if(i>=18 and i<24):
					pos = str(chr(65+i-17)) + "5"
				if(i>=24 and i<30):
					pos = str(chr(65+i-23)) + "6"
				if(i>=30 and i<36):
					pos = str(chr(65+i-29)) + "7"
				if(i>=36 and i<42):
					pos = str(chr(65+i-35)) + "8"
				if(i>=42):
					pos = str(chr(65+i-41)) + "9"
				if(self.base_window.new_quantitative_2.well_button[i]['text'][0] != '#'):
					if(self.base_window.new_quantitative_2.concentration[i] == 0):
						sum_concen_0 += self.result[i]/self.base_window.system_check.threshold
						concen_0_number += 1
					elif(self.base_window.new_quantitative_2.concentration[i] == 1):
						sum_concen_1 += self.result[i]/self.base_window.system_check.threshold
						concen_1_number += 1
					elif(self.base_window.new_quantitative_2.concentration[i] == 2):
						sum_concen_2 += self.result[i]/self.base_window.system_check.threshold
						concen_2_number += 1
					elif(self.base_window.new_quantitative_2.concentration[i] == 3):
						sum_concen_3 += self.result[i]/self.base_window.system_check.threshold
						concen_3_number += 1
					elif(self.base_window.new_quantitative_2.concentration[i] == 4):
						sum_concen_4 += self.result[i]/self.base_window.system_check.threshold
						concen_4_number += 1
					elif(self.base_window.new_quantitative_2.concentration[i] == 5):
						sum_concen_5 += self.result[i]/self.base_window.system_check.threshold
						concen_5_number += 1
					elif(self.base_window.new_quantitative_2.concentration[i] == 6):
						sum_concen_6 += self.result[i]/self.base_window.system_check.threshold
						concen_6_number += 1
					elif(self.base_window.new_quantitative_2.concentration[i] == 7):
						sum_concen_7 += self.result[i]/self.base_window.system_check.threshold
						concen_7_number += 1
					elif(self.base_window.new_quantitative_2.concentration[i] == 8):
						sum_concen_8 += self.result[i]/self.base_window.system_check.threshold
						concen_8_number += 1
					else:
						sum_concen_9 += self.result[i]/self.base_window.system_check.threshold
						concen_9_number += 1

					sheet[pos] = round(self.result[i]/self.base_window.system_check.threshold,2)
					print(self.result[i])
				else:
					sheet[pos] = "N/A"

			print("sum_concen_0: ", sum_concen_0)
			print("concen_0_number: ",concen_0_number)
			print("sum_concen_1: ", sum_concen_1)
			print("concen_1_number: ",concen_1_number)
			print("sum_concen_2: ", sum_concen_2)
			print("concen_2_number: ",concen_2_number)
			print("sum_concen_3: ", sum_concen_3)
			print("concen_3_number: ",concen_3_number)
			print("sum_concen_4: ", sum_concen_4)
			print("concen_4_number: ",concen_4_number)
			print("sum_concen_5: ", sum_concen_5)
			print("concen_5_number: ",concen_5_number)
			print("sum_concen_6: ", sum_concen_6)
			print("concen_6_number: ",concen_6_number)
			print("sum_concen_7: ", sum_concen_7)
			print("concen_7_number: ",concen_7_number)
			print("sum_concen_8: ", sum_concen_8)
			print("concen_8_number: ",concen_8_number)
			print("sum_concen_9: ", sum_concen_9)
			print("concen_9_number: ",concen_9_number)
			wb.save(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + "/analysis_value.xlsx")
			wb.close()

			number_concentration = 0
			pts_list = []
			if(concen_0_number != 0):
				number_concentration += 1
				avg_concen_0  = round(sum_concen_0/concen_0_number,2)
				print("avg_concen_0: ",avg_concen_0)
				# ~ concen_0_pt = [avg_concen_0,0]
				# ~ pts_list.append(concen_0_pt)
			if(concen_1_number != 0):
				number_concentration += 1
				avg_concen_1  = round(sum_concen_1/concen_1_number,2)
				print("avg_concen_1: ",avg_concen_1)
				concen_1_pt = [1, avg_concen_1]
				# ~ concen_1_pt = [1, 0.79]
				pts_list.append(concen_1_pt)
			if(concen_2_number != 0):
				number_concentration += 1
				avg_concen_2  = round(sum_concen_2/concen_2_number,2)
				print("avg_concen_2: ",avg_concen_2)
				concen_2_pt = [2, avg_concen_2]
				# ~ concen_2_pt = [2, 0.96]
				pts_list.append(concen_2_pt)
			if(concen_3_number != 0):
				number_concentration += 1
				avg_concen_3  = round(sum_concen_3/concen_3_number,2)
				print("avg_concen_3: ",avg_concen_3)
				concen_3_pt = [3, avg_concen_3]
				# ~ concen_3_pt = [3, 1.19]
				pts_list.append(concen_3_pt)
			if(concen_4_number != 0):
				number_concentration += 1
				avg_concen_4  = round(sum_concen_4/concen_4_number,2)
				print("avg_concen_4: ",avg_concen_4)
				concen_4_pt = [4, avg_concen_4]
				# ~ concen_4_pt = [4, 1.42]
				pts_list.append(concen_4_pt)
			if(concen_5_number != 0):
				number_concentration += 1
				avg_concen_5  = round(sum_concen_5/concen_5_number,2)
				print("avg_concen_5: ",avg_concen_5)
				concen_5_pt = [5, avg_concen_5]
				# ~ concen_5_pt = [5, 1.62]
				pts_list.append(concen_5_pt)
			if(concen_6_number != 0):
				number_concentration += 1
				avg_concen_6  = round(sum_concen_6/concen_6_number,2)
				print("avg_concen_6: ",avg_concen_6)
				concen_6_pt = [6, avg_concen_6]
				pts_list.append(concen_6_pt)
			if(concen_7_number != 0):
				number_concentration += 1
				avg_concen_7  = round(sum_concen_7/concen_7_number,2)
				print("avg_concen_7: ",avg_concen_7)
				concen_7_pt = [7, avg_concen_7]
				pts_list.append(concen_7_pt)
			if(concen_8_number != 0):
				number_concentration += 1
				avg_concen_8  = round(sum_concen_8/concen_8_number,2)
				print("avg_concen_8: ",avg_concen_8)
				concen_8_pt = [8, avg_concen_8]
				pts_list.append(concen_8_pt)
			if(concen_9_number != 0):
				number_concentration += 1
				avg_concen_9  = round(sum_concen_9/concen_9_number,2)
				print("avg_concen_9: ",avg_concen_9)
				concen_9_pt = [9, avg_concen_9]
				pts_list.append(concen_9_pt)

			pts_list = np.array(pts_list)
			print("pts_list: ", pts_list)

			x = pts_list[:,0]
			y = pts_list[:,1]

			a_value, b_value = np.polyfit(x,y,1)
			print("a_value: ", a_value)
			print("b_value: ", b_value)

			# Save program file
			wb = Workbook()
			sheet = wb.active
			sheet["B2"] = "User name:"
			sheet["B3"] = "Comment:"
			sheet["E5"] = "A"
			sheet["E6"] = "B"
			sheet["E7"] = "C"
			sheet["E8"] = "D"
			sheet["E9"] = "E"
			sheet["E10"] = "F"
			sheet["E11"] = "G"
			sheet["E12"] = "H"
			sheet["F4"] = "1"
			sheet["G4"] = "2"
			sheet["H4"] = "3"
			sheet["I4"] = "4"
			sheet["J4"] = "5"
			sheet["K4"] = "6"

			sheet["C2"] = self.base_window.new_quantitative_1.user_name
			sheet["C3"] = self.base_window.new_quantitative_1.comments

			sheet["D2"] = "a:"
			sheet["E2"] = a_value
			sheet["F2"] = "b:"
			sheet["G2"] = b_value
			sheet["H2"] = "N base value:"
			sheet["I2"] = avg_concen_0

			count = -1
			for i in range(0,48):
				if(i<6):
					pos = str(chr(69+i+1)) + "5"
				if(i>=6 and i<12):
					pos = str(chr(69+i-5)) + "6"
				if(i>=12 and i<18):
					pos = str(chr(69+i-11)) + "7"
				if(i>=18 and i<24):
					pos = str(chr(69+i-17)) + "8"
				if(i>=24 and i<30):
					pos = str(chr(69+i-23)) + "9"
				if(i>=30 and i<36):
					pos = str(chr(69+i-29)) + "10"
				if(i>=36 and i<42):
					pos = str(chr(69+i-35)) + "11"
				if(i>=42):
					pos = str(chr(68+i-41)) + "12"

				if(self.base_window.new_quantitative_2.well_button[i]['text'][0] != '#'):
					count += 1
					sheet[pos] = self.base_window.new_quantitative_2.well_button[i]['text']
					sheet["B" + str(4 + count)] = self.base_window.new_quantitative_2.well_button[i]['text']
					if(self.base_window.new_quantitative_2.concentration[i] == 0):
						sheet["C" + str(4 + count)] = "0 copies"
					if(self.base_window.new_quantitative_2.concentration[i] == 1):
						sheet["C" + str(4 + count)] = "10e1 copies"
					if(self.base_window.new_quantitative_2.concentration[i] == 2):
						sheet["C" + str(4 + count)] = "10e2 copies"
					if(self.base_window.new_quantitative_2.concentration[i] == 3):
						sheet["C" + str(4 + count)] = "10e3 copies"
					if(self.base_window.new_quantitative_2.concentration[i] == 4):
						sheet["C" + str(4 + count)] = "10e4 copies"
					if(self.base_window.new_quantitative_2.concentration[i] == 5):
						sheet["C" + str(4 + count)] = "10e5 copies"
					if(self.base_window.new_quantitative_2.concentration[i] == 6):
						sheet["C" + str(4 + count)] = "10e6 copies"
					if(self.base_window.new_quantitative_2.concentration[i] == 7):
						sheet["C" + str(4 + count)] = "10e7 copies"
					if(self.base_window.new_quantitative_2.concentration[i] == 8):
						sheet["C" + str(4 + count)] = "10e8 copies"
					if(self.base_window.new_quantitative_2.concentration[i] == 9):
						sheet["C" + str(4 + count)] = "10e9 copies"
					sheet["D" + str(4 + count)] = round(self.result[i]/self.base_window.system_check.threshold,2)
				else:
					sheet[pos] = "N/A"
			wb.save(programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + ".xlsx")
			wb.close()

			self.progressbar['value']=100
			self.base_window.update_idletasks()
			sleep(0.5)

			self.progressbar.destroy()
			self.process_label.destroy()

			self.check_result_frame = Frame(self.work_frame, bg=RESULT_TABLE_FRAME_BGD_COLOR)
			self.check_result_frame.pack(side=LEFT, anchor=E)

			result_label = list(range(48))
			r=0
			c=-1
			for i in range(0,48):
				c+=1
				if(c>5):
					c=0
					r+=1
				result_label[i] = Label(self.check_result_frame,
										width=6,
										height=3,
										# ~ bg = RESULT_LABEL_BGD_COLOR,
										font = RESULT_LABEL_TXT_FONT)
				if(self.base_window.new_quantitative_2.well_button[i]['text'][0] != '#'):
					result_label[i]['text'] = round(self.result[i]/self.base_window.system_check.threshold,2)
					result_label[i]['bg'] = "cyan"
				else:
					result_label[i]['text'] = "N/A"
					result_label[i]['bg'] = "grey80"
				result_label[i].grid(row=r,column=c, padx=2, pady=2)


			self.pfi_label = Label(self.work_frame,
									bg = LABEL_BGD_COLOR,
									text = 'PFi value: ' + str(avg_concen_0),
									fg = 'blue',
									font = ("Helvetica", 13, 'bold'))
			self.pfi_label.pack(side=LEFT, ipadx=5, ipady=5, padx=100, pady=50, anchor=NE, expand=TRUE)

			self.title_label['text'] = "CALIBRATION RESULT"

			# In button frame
			self.finish_button = Button(self.button_frame,
									text = "FINISH",
									font = SWITCH_PAGE_BUTTON_FONT,
									# width = SWITCH_PAGE_BUTTON_WIDTH,
									# height = SWITCH_PAGE_BUTTON_HEIGHT,
									bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
									fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.finish_clicked)
			self.finish_button.pack(ipady=10, side=RIGHT, ipadx=20, padx=40, anchor=W)

			def show_chart():
				print("x: ", x)
				print("y: ", y)
				pyplot.scatter(x,y)
				pyplot.plot(x, a_value*x + b_value,'r')
				pyplot.title("SC48 fit line")
				pyplot.xlabel('Concentration')
				pyplot.ylabel('Fi')
				pyplot.show()

			self.show_chart_button = Button(self.button_frame,
									text = "Chart",
									font = SWITCH_PAGE_BUTTON_FONT,
									# width = SWITCH_PAGE_BUTTON_WIDTH,
									# height = SWITCH_PAGE_BUTTON_HEIGHT,
									bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
									fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = show_chart)
			self.show_chart_button.pack(ipady=10, ipadx=20, padx=40, side=LEFT, anchor=E)

			msg = messagebox.showinfo("","COMPLETED")

		else:
			self.back_clicked()

	def finish_clicked(self):
		msg = messagebox.askquestion("","Do you want to go back to the main screen ?")
		if(msg=="yes"):
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
			# ~ self.base_window.page_num = PAGE_LIST.index('MAIN_MENU')
			self.base_window.switch_page()
			self.base_window.main_menu.reset()

class NewQualitativeFrame2(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		# 3 main frame
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = "CALIBRATION",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		# Sample button frame
		self.well_button_table_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR)
		self.well_button_table_frame.pack(side=LEFT)
		self.well_button = list(range(48))
		r=0
		c=-1
		for i in range(0,48):
			c+=1
			if(c>5):
				c=0
				r+=1
			self.well_button[i] = Button(self.well_button_table_frame,
										bg = SAMPLE_BUTTON_BGD_COLOR,
										fg = SAMPLE_BUTTON_TXT_COLOR,
										activebackground = SAMPLE_BUTTON_ACTIVE_BGD_COLOR,
										justify = 'left',
										borderwidth = 0,
										text = '#',
										width = SAMPLE_BUTTON_WIDTH,
										height = SAMPLE_BUTTON_HEIGHT)
			self.well_button[i]['command'] = partial(self.well_button_clicked, i)
			self.well_button[i].grid(row=r, column=c, padx=2, pady=2)

		# Properties frame
		self.property_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR, width=495)
		self.property_frame.pack(fill=BOTH, expand=TRUE, side=RIGHT)

		self.property_labelframe = LabelFrame(self.property_frame,
										text = "Sample Properties",
										font  = LABEL_FRAME_TXT_FONT,
										bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
										fg = LABEL_FRAME_TXT_COLOR)
		self.property_labelframe.pack(fill=BOTH, expand=TRUE, padx=10, pady=10)

		self.property_labelframe.rowconfigure(0, weight=1)
		self.property_labelframe.rowconfigure(1, weight=1)
		self.property_labelframe.rowconfigure(2, weight=1)
		self.property_labelframe.rowconfigure(3, weight=4)

		self.well_name_label = Label(self.property_labelframe,
						bg = SAMPLE_BUTTON_CHOOSE_BGD_COLOR,
						fg = LABEL_TXT_COLOR,
						font = SAMPLE_LABEL_TXT_FONT)
		self.well_name_label.grid(row=0, column=0, columnspan=2, sticky=EW)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = "Back",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

		self.note_label = Label(self.button_frame,
							text='Please set and input primer mix kit',
							fg = 'red',
							bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR,
							font = LABEL_TXT_FONT)
		self.note_label.pack(side=LEFT, ipadx=175)

		self.next_button = Button(self.button_frame,
								text = "Next",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.next_clicked)
		self.next_button.pack(ipadx=30, ipady=10, side=RIGHT)

	def well_button_clicked(self,n):
		if(self.well_button[n]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR):
			for k in range (0,48):
				if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
					self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
				else:
					self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
			self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
		else:
			for k in range (0,48):
				if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
					self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
				if(self.well_button[k]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
					self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
			self.well_button[n]['bg'] = SAMPLE_BUTTON_TMP_BGD_COLOR


		def ok_clicked(event=None):
			if(self.sample_name_entry.get()==''):
				self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
				self.well_button[n]['text'] = '#'
				msgbox = messagebox.showwarning("","Please enter sample name !")
			else:
				self.well_button[n]['text'] = self.sample_name_entry.get()
				self.well_button[n]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
				try:
					if(n==42):
						self.well_button_clicked(1)
					elif(n==43):
						self.well_button_clicked(2)
					elif(n==44):
						self.well_button_clicked(3)
					elif(n==45):
						self.well_button_clicked(4)
					elif(n==46):
						self.well_button_clicked(5)
					elif(n==47):
						self.well_button_clicked(0)
					else:
						self.well_button_clicked(n+6)
				except:
					self.well_button_clicked(0)


		sample_name_label = Label(self.property_labelframe,
									text = "Sample Name",
									font = LABEL_TXT_FONT,
									bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
									fg = LABEL_TXT_COLOR)
		sample_name_label.grid(row=1, column=0, padx=78, pady=2, sticky=SE)

		self.sample_name_entry = Entry(self.property_labelframe, width=20, font=ENTRY_TXT_FONT)
		if(self.well_button[n]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
			self.sample_name_entry.insert(0, self.well_button[n]['text'])
		#id_entry.bind("<Button-1>", enter_entry)
		self.sample_name_entry.bind("<Return>", ok_clicked)
		self.sample_name_entry.grid(row=2, column=0, padx=30, pady=0)
		self.sample_name_entry.focus_set()

		if(n<6):
			self.well_name_label['text'] = "A" + str(n+1)
		elif(n<12):
			self.well_name_label['text'] = "B" + str(n+1-6)
		elif(n<18):
			self.well_name_label['text'] = "C" + str(n+1-12)
		elif(n<24):
			self.well_name_label['text'] = "D" + str(n+1-18)
		elif(n<30):
			self.well_name_label['text'] = "E" + str(n+1-24)
		elif(n<36):
			self.well_name_label['text'] = "F" + str(n+1-30)
		elif(n<42):
			self.well_name_label['text'] = "G" + str(n+1-36)
		else:
			self.well_name_label['text'] = "H" + str(n+1-42)

		# ~ if(n<8):
			# ~ self.well_name_label['text'] = str(chr(65+n)) + '1'
		# ~ if(n>=8 and n<16):
			# ~ self.well_name_label['text'] = str(chr(65+n-8)) + '2'
		# ~ if(n>=16 and n<24):
			# ~ self.well_name_label['text'] = str(chr(65+n-16)) + '3'
		# ~ if(n>=24 and n<32):
			# ~ self.well_name_label['text'] = str(chr(65+n-24)) + '4'
		# ~ if(n>=32 and n<40):
			# ~ self.well_name_label['text'] = str(chr(65+n-32)) + '5'
		# ~ if(n>=40):
			# ~ self.well_name_label['text'] = str(chr(65+n-40)) + '6'

		self.ok_button = Button(self.property_labelframe,
								text = "OK",
								bg = CONFIRM_BUTTON_BGD_COLOR,
								fg = CONFIRM_BUTTON_TXT_COLOR,
								font = CONFIRM_BUTTON_TXT_FONT,
								borderwidth = 0,
								command = ok_clicked)
		self.ok_button.grid(row=3, column=0, columnspan=2, ipadx=30, ipady=10)

	def back_clicked(self):
		self.base_window.frame_list.remove(self.base_window.system_check)
		del self.base_window.system_check
		self.base_window.system_check = SystemCheckFrame(self.base_window)
		self.base_window.frame_list.append(self.base_window.system_check)

		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_1)
		self.base_window.switch_page()


	def next_clicked(self):
		self.number_well_active = 0
		for i in range(0,48):
			if(self.well_button[i]['text'] != '#'):
				self.number_well_active += 1

		if(self.number_well_active == 0):
			msg = messagebox.showwarning("","You need to enter at least 1 sample name !")
		else:
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_3)
			self.base_window.switch_page()
			self.base_window.update_idletasks()
			self.base_window.new_qualitative_3.serial_handle()


class NewQuantitativeFrame2(NewQualitativeFrame2):
	def __init__(self, container):
		super().__init__(container)

		# In title frame
		self.title_label['text'] = "CALIBRATION"

		self.concentration = list(range(48))

	def well_button_clicked(self,n):
		if(self.well_button[n]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR):
			for k in range (0,48):
				if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
					self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
				else:
					self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
			self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
		else:
			for k in range (0,48):
				if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
					self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
				if(self.well_button[k]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
					self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
			self.well_button[n]['bg'] = SAMPLE_BUTTON_TMP_BGD_COLOR


		def ok_clicked(event=None):
			if(self.sample_name_entry.get()==''):
				self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
				self.well_button[n]['text'] = '#'+str(n+1)
				msgbox = messagebox.showwarning("","Please enter sample name !")
			else:
				self.well_button[n]['text'] = self.sample_name_entry.get()
				self.well_button[n]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
				self.concentration[n] = self.concentration_combobox.current()
				print("concentration[%d] = %s"%(n,self.concentration[n]))
				try:
					if(n==42):
						self.well_button_clicked(1)
					elif(n==43):
						self.well_button_clicked(2)
					elif(n==44):
						self.well_button_clicked(3)
					elif(n==45):
						self.well_button_clicked(4)
					elif(n==46):
						self.well_button_clicked(5)
					elif(n==47):
						self.well_button_clicked(0)
					else:
						self.well_button_clicked(n+6)
				except:
					self.well_button_clicked(0)


		sample_name_label = Label(self.property_labelframe,
									text = "Sample Name",
									font = LABEL_TXT_FONT,
									bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
									fg = LABEL_TXT_COLOR)
		sample_name_label.grid(row=1, column=0, padx=78, pady=2, sticky=SE)

		concentration_label = Label(self.property_labelframe,
									text = "Concentration (copies)",
									font = LABEL_TXT_FONT,
									bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
									fg = LABEL_TXT_COLOR)
		concentration_label.grid(row=1, column=1, padx=78, pady=2, sticky=S)

		self.sample_name_entry = Entry(self.property_labelframe, width=20, font=ENTRY_TXT_FONT)
		if(self.well_button[n]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
			self.sample_name_entry.insert(0, self.well_button[n]['text'])
		#id_entry.bind("<Button-1>", enter_entry)
		self.sample_name_entry.bind("<Return>", ok_clicked)
		self.sample_name_entry.grid(row=2, column=0, padx=0, pady=0, sticky=NE)
		self.sample_name_entry.focus_set()


		self.concentration_value = StringVar()
		self.concentration_combobox = ttk.Combobox(self.property_labelframe,
												state = "readonly",
												width = 8,
												textvariable = self.concentration_value)
		self.concentration_combobox['values'] = ('0',
												'10e1',
												'10e2',
												'10e3',
												'10e4',
												'10e5',
												'10e6',
												'10e7',
												'10e8',
												'10e9')
		if(self.well_button[n]['text'][0] != '#'):
			self.concentration_combobox.current(self.concentration[n])
		else:
			self.concentration_combobox.current(0)

		self.concentration_combobox.grid(row=2, column=1, padx=0, pady=0, sticky=N)


		if(n<6):
			self.well_name_label['text'] = "A" + str(n+1)
		elif(n<12):
			self.well_name_label['text'] = "B" + str(n+1-6)
		elif(n<18):
			self.well_name_label['text'] = "C" + str(n+1-12)
		elif(n<24):
			self.well_name_label['text'] = "D" + str(n+1-18)
		elif(n<30):
			self.well_name_label['text'] = "E" + str(n+1-24)
		elif(n<36):
			self.well_name_label['text'] = "F" + str(n+1-30)
		elif(n<42):
			self.well_name_label['text'] = "G" + str(n+1-36)
		else:
			self.well_name_label['text'] = "H" + str(n+1-42)


		self.ok_button = Button(self.property_labelframe,
								text = "OK",
								bg = CONFIRM_BUTTON_BGD_COLOR,
								fg = CONFIRM_BUTTON_TXT_COLOR,
								font = CONFIRM_BUTTON_TXT_FONT,
								borderwidth = 0,
								command = ok_clicked)
		self.ok_button.grid(row=3, column=0, columnspan=2, ipadx=30, ipady=10)


	def back_clicked(self):
		self.base_window.frame_list.remove(self.base_window.system_check)
		del self.base_window.system_check
		self.base_window.system_check = SystemCheckFrame(self.base_window)
		self.base_window.frame_list.append(self.base_window.system_check)

		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_1)
		self.base_window.switch_page()

	def next_clicked(self):
		self.number_well_active = 0
		self.number_concentration = 0
		concen0_flag = 0
		concen1_flag = 0
		concen2_flag = 0
		concen3_flag = 0
		concen4_flag = 0
		concen5_flag = 0
		concen6_flag = 0
		concen7_flag = 0
		concen8_flag = 0
		concen9_flag = 0
		for i in range(0,48):
			if(self.well_button[i]['bg'] == SAMPLE_BUTTON_DONE_BGD_COLOR):
				self.number_well_active += 1
				if(self.concentration[i] == 0):
					if(concen0_flag == 0):
						self.number_concentration += 1
						concen0_flag = 1
				if(self.concentration[i] == 1):
					if(concen1_flag == 0):
						self.number_concentration += 1
						concen1_flag = 1
				if(self.concentration[i] == 2):
					if(concen2_flag == 0):
						self.number_concentration += 1
						concen2_flag = 1
				if(self.concentration[i] == 3):
					if(concen3_flag == 0):
						self.number_concentration += 1
						concen3_flag = 1
				if(self.concentration[i] == 4):
					if(concen4_flag == 0):
						self.number_concentration += 1
						concen4_flag = 1
				if(self.concentration[i] == 5):
					if(concen5_flag == 0):
						self.number_concentration += 1
						concen5_flag = 1
				if(self.concentration[i] == 6):
					if(concen6_flag == 0):
						self.number_concentration += 1
						concen6_flag = 1
				if(self.concentration[i] == 7):
					if(concen7_flag == 0):
						self.number_concentration += 1
						concen7_flag = 1
				if(self.concentration[i] == 8):
					if(concen8_flag == 0):
						self.number_concentration += 1
						concen8_flag = 1
				if(self.concentration[i] == 9):
					if(concen9_flag == 0):
						self.number_concentration += 1
						concen9_flag = 1

		if(self.number_well_active == 0):
			msg = messagebox.showwarning("","Please enter sample name !")
		elif(self.number_concentration <= 1 or concen0_flag == 0): # it nhat phai co 2 concentratiobn khac nhau
			msg = messagebox.showwarning("","Must have at least 3 different concentrations (include 0 copy)!")
		else:
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_3)
			self.base_window.switch_page()
			self.base_window.update_idletasks()
			self.base_window.new_quantitative_3.serial_handle()



class NewQualitativeFrame1(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.experiment_name = StringVar()
		self.user_name = StringVar()
		self.comments = StringVar()

		# 3 main frame
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = "CALIBRATION",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		setup_labelframe = LabelFrame(self.work_frame,
									text = "Properties",
									font = LABEL_FRAME_TXT_FONT,
									bg = LABEL_FRAME_BGD_COLOR,
									fg = LABEL_FRAME_TXT_COLOR)
		setup_labelframe.pack(expand=TRUE)

		experiment_name_label = Label(setup_labelframe,
									text = "Kit Name",
									font = LABEL_TXT_FONT,
									bg = LABEL_BGD_COLOR,
									fg = LABEL_TXT_COLOR)
		experiment_name_label.grid(row=0, column=1, sticky=E, pady=20, padx=30)

		user_name_label = Label(setup_labelframe,
								text = "User Name",
								font = LABEL_TXT_FONT,
								bg = LABEL_BGD_COLOR,
								fg = LABEL_TXT_COLOR)
		user_name_label.grid(row=1, column=1, sticky=E, pady=20, padx=30)

		comment_label = Label(setup_labelframe,
								text = "Comment",
								font = LABEL_TXT_FONT,
								bg = LABEL_BGD_COLOR,
								fg = LABEL_TXT_COLOR)
		comment_label.grid(row=2, column=1, sticky=NE, pady=20, padx=30)

		self.experiment_name_entry = Entry(setup_labelframe, width=30, font=ENTRY_TXT_FONT)
		self.experiment_name_entry.grid(row=0, column=2, sticky=W, pady=20, padx=35)
		self.user_name_entry = Entry(setup_labelframe, width=30, font=ENTRY_TXT_FONT)
		self.user_name_entry.grid(row=1, column=2, sticky=W, pady=20, padx=35)
		self.comments_text = Text(setup_labelframe, width=30, height=9, font=ENTRY_TXT_FONT)
		self.comments_text.grid(row=2, column=2, sticky=E, pady=20, padx=35)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = "Back",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

		self.next_button = Button(self.button_frame,
								text = "Next",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.next_clicked)
		self.next_button.pack(ipadx=30, ipady=10, side=RIGHT)

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_calib_list)
		self.base_window.switch_page()

	def next_clicked(self):
		msg = messagebox.askokcancel("","Please make sure no sample is placed in the device !")
		if(msg == True):
			self.experiment_name = self.experiment_name_entry.get()
			self.user_name = self.user_name_entry.get()
			self.comments = self.comments_text.get("1.0",'end-1c')
			if(self.experiment_name==''):
				messagebox.showwarning("","Plese enter Experiment Name !")
			else:
				if (os.path.exists(programs_qualitative_path + self.experiment_name + ".xlsx")):
					msg = messagebox.askquestion("","This Experiment Name already exists.\n Do you want to replace it?")
					if(msg == 'yes'):
						if os.path.exists(results_programs_qualitative_path + self.experiment_name):
							f = results_programs_qualitative_path + self.experiment_name
							shutil.rmtree(f)
							os.mkdir(f)
						else:
							f = os.path.join(results_programs_qualitative_path + self.base_window.experiment_name)
							os.mkdir(f)

						self.base_window.system_check.mode_check = 0
						self.base_window.forget_page()
						self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
						self.base_window.switch_page()
						self.base_window.update_idletasks()
						self.base_window.system_check.serial_handle()
				else:
					if os.path.exists(results_programs_qualitative_path + self.experiment_name):
						f = results_programs_qualitative_path + self.experiment_name
						shutil.rmtree(f)
						os.mkdir(f)
					else:
						f = os.path.join(results_programs_qualitative_path + self.experiment_name)
						os.mkdir(f)
					self.base_window.system_check.mode_check = 0
					self.base_window.forget_page()
					self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
					self.base_window.switch_page()
					self.base_window.update_idletasks()
					self.base_window.system_check.serial_handle()


class NewQuantitativeFrame1(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.experiment_name = StringVar()
		self.user_name = StringVar()
		self.comments = StringVar()

		# 3 main frame
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = "KITS MANAGEMENT",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		setup_labelframe = LabelFrame(self.work_frame,
									text = "Properties",
									font = LABEL_FRAME_TXT_FONT,
									bg = LABEL_FRAME_BGD_COLOR,
									fg = LABEL_FRAME_TXT_COLOR)
		setup_labelframe.pack(expand=TRUE)

		experiment_name_label = Label(setup_labelframe,
									text = "Kit Name",
									font = LABEL_TXT_FONT,
									bg = LABEL_BGD_COLOR,
									fg = LABEL_TXT_COLOR)
		experiment_name_label.grid(row=0, column=1, sticky=E, pady=20, padx=30)

		info_label = Label(setup_labelframe,
								text = "Information",
								font = LABEL_TXT_FONT,
								bg = LABEL_BGD_COLOR,
								fg = LABEL_TXT_COLOR)
		info_label.grid(row=1, column=1, sticky=NE, pady=20, padx=30)

		self.experiment_name_entry = Entry(setup_labelframe, width=30, font=ENTRY_TXT_FONT)
		self.experiment_name_entry.grid(row=0, column=2, sticky=W, pady=20, padx=35)
		self.properties_text = Text(setup_labelframe, width=30, height=9, font=ENTRY_TXT_FONT)
		self.properties_text.grid(row=1, column=2, sticky=E, pady=20, padx=35)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = "Back",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

		self.next_button = Button(self.button_frame,
								text = "Next",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.next_clicked)
		self.next_button.pack(ipadx=30, ipady=10, side=RIGHT)

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_calib_list)
		self.base_window.switch_page()

	def next_clicked(self):
		msg = messagebox.askokcancel("","Please make sure no sample is placed in the device !")
		if(msg == True):
			self.experiment_name = self.experiment_name_entry.get()
			self.user_name = self.user_name_entry.get()
			self.comments = self.comments_text.get("1.0",'end-1c')
			if(self.experiment_name==''):
				messagebox.showwarning("","Plese enter Experiment Name !")
			else:
				if (os.path.exists(programs_quantitative_path + self.experiment_name + ".xlsx")):
					msg = messagebox.askquestion("","This Experiment Name already exists.\n Do you want to replace it?")
					if(msg == 'yes'):
						if os.path.exists(results_programs_quantitative_path + self.experiment_name):
							f = results_programs_quantitative_path + self.experiment_name
							shutil.rmtree(f)
							os.mkdir(f)
						else:
							f = os.path.join(results_programs_quantitative_path + self.base_window.experiment_name)
							os.mkdir(f)

						self.base_window.system_check.mode_check = 2
						self.base_window.forget_page()
						self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
						self.base_window.switch_page()
						self.base_window.update_idletasks()
						self.base_window.system_check.serial_handle()
				else:
					if os.path.exists(results_programs_quantitative_path + self.experiment_name):
						f = results_programs_quantitative_path + self.experiment_name
						shutil.rmtree(f)
						os.mkdir(f)
					else:
						f = os.path.join(results_programs_quantitative_path + self.experiment_name)
						os.mkdir(f)
					self.base_window.system_check.mode_check = 2
					self.base_window.forget_page()
					self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
					self.base_window.switch_page()
					self.base_window.update_idletasks()
					self.base_window.system_check.serial_handle()


class QualitativeAnalysisFrame3(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = Screening3_Language["Title Label"][language],
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

	def serial_handle(self):
		style = ttk.Style()
		style.theme_use('clam')
		style.configure("1.Horizontal.TProgressbar", troughcolor ='grey85', background='green3')
		self.progressbar = ttk.Progressbar(self.work_frame,
									style="1.Horizontal.TProgressbar",
									length = 200,
									mode = 'determinate')
		self.progressbar.pack(ipadx=2, ipady=2)

		self.process_label = Label(self.work_frame,
						bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
						fg = 'black',
						text = Screening3_Language["Process Label"][language],
						font = LABEL_TXT_FONT)
		self.process_label.pack(ipadx=2, ipady=2, anchor=N)

		if(SERIAL_COMUNICATION):
			ser.flushInput()
			ser.flushOutput()

			self.progressbar['value']=5
			self.base_window.update_idletasks()
			sleep(0.5)

			data_send = 'P'
			print("Data send:", data_send)
			ser.write(data_send.encode())

			receive_data = StringVar()
			count = 0
			bled_ready = 0
			while(receive_data != 'C'):
				if(ser.in_waiting>0):
					receive_data = ser.readline().decode('utf-8').rstrip()
					print("Data received:", receive_data)

					self.progressbar['value']=10
					self.base_window.update_idletasks()
					sleep(0.5)

					if(receive_data == 'C'):
						self.progressbar['value']=20
						self.base_window.update_idletasks()
						sleep(0.5)

						bled_ready = 1
						break;
				else:
					sleep(1)
					count += 1
					if(count > 15):
						msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value),
													ERROR_LIST(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value).name,
													icon = "error")
						break;
		else:
			bled_ready = 1
		
		if(bled_ready):
			GPIO.output(BLUELIGHT_PIN, GPIO.HIGH)
			
			self.progressbar['value']=45
			self.base_window.update_idletasks()
			try:
				#camera_capture(self.base_window.qualitative_analysis_1.analysis_result_folder + '/raw.jpg')
				camera_capture(self.base_window.qualitative_analysis_0.analysis_result_folder + '/raw.jpg')
				
				self.progressbar['value']=65
				self.base_window.update_idletasks()
			except Exception as e :
				msg = messagebox.showerror("ERR "+ str(ERROR_LIST['CAMERA_ERROR'].value),
										ERROR_LIST(ERROR_LIST['CAMERA_ERROR'].value).name,
										icon = "error")
				if(msg=='ok'):
					self.base_window.destroy()

			self.result, self.image = Process_Image(self.base_window.qualitative_analysis_0.analysis_result_folder + '/raw.jpg').process(coefficient, mode=1, well_list=self.base_window.qualitative_analysis_2.id_list)            
			cv2.imwrite(self.base_window.qualitative_analysis_0.analysis_result_folder + '/process.jpg', self.image)
			
			self.progressbar['value']=80
			self.base_window.update_idletasks()
			sleep(0.5)
			
			GPIO.output(BLUELIGHT_PIN, GPIO.LOW)

			sum_value = 0
			active_well_number = 0

			# Save analysis file
			wb = Workbook()
			sheet = wb.active

			sheet["A2"] = "A"
			sheet["A3"] = "B"
			sheet["A4"] = "C"
			sheet["A5"] = "D"
			sheet["B1"] = "1"
			sheet["C1"] = "2"
			sheet["D1"] = "3"
			sheet["E1"] = "4"
			if(SC_VERSION == 48):
				sheet["A6"] = "E"
				sheet["A7"] = "F"
				sheet["A8"] = "G"
				sheet["A9"] = "H"
				sheet["F1"] = "5"
				sheet["G1"] = "6"

			index = 0
			for r in range(0, WELL_ROW):
				for c in range(0, WELL_COLUMN):
					pos = str(chr(66+c)) + str(r+2)
					if(self.base_window.qualitative_analysis_2.id_list[index] != 'N/A'):
						sheet[pos] = round(self.result[index]/self.base_window.system_check.threshold, 2)
					else:
						sheet[pos] = "N/A"

					index += 1

			wb.save(self.base_window.qualitative_analysis_0.analysis_result_folder + "/analysis_value.xlsx")
			wb.close()


			self.progressbar['value']=95
			self.base_window.update_idletasks()
			sleep(0.5)

			self.progressbar['value']=100
			self.base_window.update_idletasks()
			sleep(0.5)
			
			

			self.progressbar.destroy()
			self.process_label.destroy()

			self.tab_control = ttk.Notebook(self.work_frame)
			self.result_tab = Frame(self.tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
			self.report_tab = Frame(self.tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
			self.image_tab = Frame(self.tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)

			self.tab_control.add(self.result_tab, text = Screening3_Language["Result Tab"][language])
			self.tab_control.add(self.report_tab, text = Screening3_Language["Report Tab"][language])
			self.tab_control.add(self.image_tab, text= Screening3_Language["Images Tab"][language])
			self.tab_control.grid(row=0, column=0, padx=0, pady=1, sticky=EW)

			self.check_result_frame = Frame(self.result_tab, bg=RESULT_TABLE_FRAME_BGD_COLOR)
			self.check_result_frame.grid(row=0, column=0, padx=76)

			# ~ Pmw.initialise(self.base_window)
			# ~ self.tooltip = list(range(48))
			
			#self.pfi_value = round(self.result[47]/self.base_window.system_check.threshold,2)
			self.pfi_value = round(self.base_window.system_check.average_current_intensity/self.base_window.system_check.threshold,2)

			self.result_label = list(range(SC_VERSION))
			index = 0 
			for r in range(0, WELL_ROW):
				for c in range(0, WELL_COLUMN):
					self.result_label[index] = Label(self.check_result_frame,
											width = 6, 
											height = 3,
											font = RESULT_LABEL_TXT_FONT_1)

					# ~ self.tooltip[i] = Pmw.Balloon(self.base_window)
					# ~ self.tooltip[i].bind(result_label[i], self.base_window.qualitative_analysis_2.id_list[i])

					# neu chon do nhay thap ([Tam Viet only] - Chỉ quan tâm đến độ nhạy thấp)
					if(self.base_window.main_menu.threshold_value == 0):
						if(self.base_window.qualitative_analysis_2.id_list[index] != 'N/A'):
							self.result_label[index]['text'] = round(self.result[index]/self.base_window.system_check.threshold,2)
							if(float(self.result_label[index]['text']) <= self.base_window.main_menu.num1):
								self.result_label[index]['bg'] = NEGATIVE_COLOR
							elif(float(self.result_label[index]['text']) <= self.base_window.main_menu.num2):
								self.result_label[index]['bg'] = LOW_COPY_COLOR
							else:
								self.result_label[index]['bg'] = POSITIVE_COLOR
						else:
							self.result_label[index]['text'] = "N/A"
							self.result_label[index]['bg'] = NA_COLOR


					# neu chon do nhay cao
					else:
						if(self.base_window.qualitative_analysis_2.id_list[index] != 'N/A'):
							self.result_label[index]['text'] = round(self.result[index]/self.base_window.system_check.threshold,2)
							if(float(self.result_label[index]['text']) <= self.base_window.main_menu.num2):
								self.result_label[index]['bg'] = NEGATIVE_COLOR
							elif(float(self.result_label[index]['text']) <= self.base_window.main_menu.num2*self.base_window.main_menu.num3):
								self.result_label[index]['bg'] = LOW_COPY_COLOR
							else:
								self.result_label[index]['bg'] = POSITIVE_COLOR
						else:
							self.result_label[index]['text'] = "N/A"
							self.result_label[index]['bg'] = NA_COLOR

					self.result_label[index].grid(row=r,column=c, padx=1, pady=1)

					index += 1
							
			self.annotate_result_frame = Frame(self.result_tab, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
			self.annotate_result_frame.grid(row=0, column=1, padx=60)
			
			negative_label = Label(self.annotate_result_frame, bg=NEGATIVE_COLOR, width=4, height=2)
			negative_label.grid(row=0, column=0, padx=20, pady=10)
			negative_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text=Screening3_Language["Note Negative Label"][language], height=2)
			negative_text_label.grid(row=0, column=1, padx=20, pady=10)
			low_copy_label = Label(self.annotate_result_frame, bg=LOW_COPY_COLOR, width=4, height=2)
			low_copy_label.grid(row=1, column=0, padx=20, pady=10)
			low_copy_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text=Screening3_Language["Note LowCopy Label"][language], height=2)
			low_copy_text_label.grid(row=1, column=1, padx=20, pady=10)
			positive_label = Label(self.annotate_result_frame, bg=POSITIVE_COLOR, width=4, height=2)
			positive_label.grid(row=2, column=0, padx=20, pady=10)
			positive_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text=Screening3_Language["Note Positive Label"][language], height=2)
			positive_text_label.grid(row=2, column=1, padx=20, pady=10)
			na_label = Label(self.annotate_result_frame, bg=NA_COLOR, width=4, height=2)
			na_label.grid(row=3, column=0, padx=20, pady=10)
			na_copy_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text=Screening3_Language["Note NoSample Label"][language], height=2)
			na_copy_text_label.grid(row=3, column=1, padx=20, pady=10)
			

			# images tab
			self.img_labelframe_1 = LabelFrame(self.image_tab,
										bg="black",
										text = Screening3_Language["RawImage LabelFrame"][language],
										fg="cyan",
										font = LABELFRAME_TXT_FONT)
			self.img_labelframe_1.pack(fill=BOTH, expand=TRUE, side=LEFT)

			self.img_labelframe_2 = LabelFrame(self.image_tab,
										bg="black",
										text = Screening3_Language["AnalyzedImage LabelFrame"][language],
										fg="cyan",
										font = LABELFRAME_TXT_FONT)
			self.img_labelframe_2.pack(fill=BOTH, expand=TRUE, side=RIGHT)

			a1 = Image.open(self.base_window.qualitative_analysis_0.analysis_result_folder + '/raw.jpg')
			a1_crop = a1.crop((x1-10, y1-10, x2+10, y2+10))
			crop_width, crop_height = a1_crop.size
			scale_percent = 100
			width = int(crop_width * scale_percent / 100)
			height = int(crop_height * scale_percent / 100)
			display_img = a1_crop.resize((width,height))
			a1_display = ImageTk.PhotoImage(display_img)
			a1_label = Label(self.img_labelframe_1, image=a1_display, bg="black")
			a1_label.image = a1_display
			a1_label.pack(fill=BOTH, expand=TRUE)

			a2 = Image.open(self.base_window.qualitative_analysis_0.analysis_result_folder + '/process.jpg')
			a2_crop = a2.crop((x1-10, y1-10, x2+10, y2+10))
			crop_width, crop_height = a2_crop.size
			scale_percent = 100
			width = int(crop_width * scale_percent / 100)
			height = int(crop_height * scale_percent / 100)
			display_img = a2_crop.resize((width,height))
			a2_display = ImageTk.PhotoImage(display_img)
			a2_label = Label(self.img_labelframe_2, image=a2_display, bg="black")
			a2_label.image = a2_display
			a2_label.pack(fill=BOTH, expand=TRUE)


			# In button frame
			self.quantitative_view = 0 
			self.thr_value = IntVar()
			self.thr1_radio_button = Radiobutton(self.button_frame, text = Screening3_Language["Environment RadioButton"][language],  width=12, variable = self.thr_value, value=0, command=self.thr_choose)
			self.thr2_radio_button = Radiobutton(self.button_frame, text = Screening3_Language["Host RadioButton"][language],  width=12, variable = self.thr_value, value=1, command=self.thr_choose)
			self.quantitative_radio_button = Radiobutton(self.button_frame, text = Screening3_Language["Quantitative RadioButton"][language], width=12, variable = self.thr_value, value=2, command=self.quantitative_switch)

			self.thr1_radio_button.pack(ipady=10, side=LEFT)
			# self.thr2_radio_button.pack(ipady=10, side=LEFT)
			self.quantitative_radio_button.pack(ipady=10, side=LEFT)
			
			if(self.base_window.main_menu.threshold_value == 0):
				self.thr_value.set(0)
			else:
				self.thr_value.set(1)
				
			self.finish_button = Button(self.button_frame,
									text = Screening3_Language["Finish Button"][language],
									font = SWITCH_PAGE_BUTTON_FONT,
									width = 10,
									# height = SWITCH_PAGE_BUTTON_HEIGHT,
									bg = 'dodger blue',
									fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.finish_clicked)
			self.finish_button.pack(ipady=10, side=RIGHT)
			
			self.base_window.update_idletasks()

			subprocess.call(["scrot", self.base_window.qualitative_analysis_0.result_folder_path + "/result_capture.jpg"])

			self.base_window.server_setting.check_status()

			wb = Workbook()
			sheet = wb.active

			font0 = Font(bold=False)
			font1 = Font(size='14', bold=True, color='00FF0000')
			font2 = Font(bold=True)
			thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

			for i in range(RESULT_CELL_START, RESULT_CELL_START + SC_VERSION):
				sheet["B"+str(i)].font = font0
				sheet["D"+str(i)].font = font0
				sheet["E"+str(i)].font = font0

			img = Img(working_dir + "/logo.png")
			# img.height = 39
			# img.width = 215
			img.height = 160
			img.width = 160
			img.anchor = 'B2'
			sheet.add_image(img)

			img = Img(self.base_window.qualitative_analysis_0.result_folder_path + "/result_capture.jpg")
			img.anchor = 'I17'
			sheet.add_image(img)

			sheet["C10"] = self.base_window.qualitative_analysis_0.template_name

			sheet.merge_cells(start_row=11, start_column=2, end_row=11, end_column=7)
			sheet["B11"] = Screening3_Language["Result Title Text"][language]
			sheet["B11"].font = font1
			sheet.cell(row=11,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet["B13"] = Screening3_Language["Result ExperimentName Text"][language] + self.base_window.qualitative_analysis_0.experiment_name
			sheet["B13"].font = font2
			sheet['B14'] = Screening3_Language["Result TechnicianName Text"][language] + self.base_window.qualitative_analysis_0.user_name
			sheet["B14"].font = font2
			sheet['B15'] = Screening3_Language["Result Date Text"][language] + self.base_window.qualitative_analysis_0.create_time
			sheet["B15"].font = font2
			# sheet['B60'] = Screening3_Language["Result Note Text"][language]
			# sheet["B60"].font = font2
			# sheet['B61'] = Screening3_Language["Result NoSample Text"][language]
			# sheet['B62'] = Screening3_Language["Result Negative Text"][language]
			# sheet['C61'] = Screening3_Language["Result LowCopy Text"][language]
			# sheet['C62'] = Screening3_Language["Result Positive Text"][language]

			sheet.merge_cells(start_row=45, start_column=5, end_row=45, end_column=7)
			sheet.merge_cells(start_row=46, start_column=5, end_row=46, end_column=7)
			sheet['B45'] = Screening3_Language["Result TechnicianNameSign Text"][language]
			sheet['B46'] = ''
			sheet['E45'] = Screening3_Language["Result HeadOfDivisionSign Text"][language]
			sheet['E46'] = ''
			sheet["B45"].font = font2
			sheet["E45"].font = font2
			sheet["B45"].protection = Protection(locked=False, hidden=False)
			sheet["E45"].protection = Protection(locked=False, hidden=False)
			sheet.cell(row=45,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=46,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=45,column=5).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=46,column=5).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)

			
			sheet.merge_cells(start_row=35, start_column=2, end_row=35, end_column=3)
			sheet.row_dimensions[35].height = 40
			sheet['B35'] = Screening3_Language["Result NoteTable Text"][language]
			sheet["B35"].font = font2
			sheet["B36"].font = font2
			sheet["C36"].font = font2
			sheet["B42"].font = font2
			sheet.cell(row=35,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=35,column=2).border = thin_border
			sheet.cell(row=35,column=3).border = thin_border
			sheet.cell(row=36,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=36,column=2).border = thin_border
			sheet.cell(row=37,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=37,column=2).border = thin_border
			sheet.cell(row=38,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=38,column=2).border = thin_border
			sheet.cell(row=39,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=39,column=2).border = thin_border
			sheet.cell(row=40,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=40,column=2).border = thin_border
			sheet.cell(row=36,column=3).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=36,column=3).border = thin_border
			sheet.cell(row=37,column=3).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=37,column=3).border = thin_border
			sheet.cell(row=38,column=3).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=38,column=3).border = thin_border
			sheet.cell(row=39,column=3).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=39,column=3).border = thin_border
			sheet.cell(row=40,column=3).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=40,column=3).border = thin_border
			sheet['B36'] = 'Spotcheck'
			sheet['C36'] = 'CT'
			sheet['B37'] = 'X ≤ ' + str(self.base_window.main_menu.num1)
			sheet['C37'] = '-'
			sheet['B38'] = str(self.base_window.main_menu.num1) + ' < X ≤ ' + str(self.base_window.main_menu.num2)
			sheet['C38'] = '31-35'
			sheet['B39'] = str(self.base_window.main_menu.num2) + ' < X ≤ ' + str(self.base_window.main_menu.num3)
			sheet['C39'] = '26-30'
			sheet['B40'] = 'X > ' + str(self.base_window.main_menu.num3)
			sheet['C40'] = '< 26'

			sheet['B42'] = Screening3_Language["Result NoteLabel Text"][language]
			sheet.merge_cells(start_row=43, start_column=2, end_row=43, end_column=6)
			sheet['B43'] = Screening3_Language["Result NoteContent Text"][language]

			for r in range(RESULT_CELL_START - 1, RESULT_CELL_START + SC_VERSION):
				for c in range(2,8):
					sheet.cell(row=r,column=c).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
					sheet.cell(row=r,column=c).border = thin_border

			sheet.column_dimensions['B'].width = 24
			sheet.column_dimensions['C'].width = 10
			sheet.column_dimensions['D'].width = 12
			sheet.column_dimensions['E'].width = 10
			sheet.column_dimensions['F'].width = 9
			sheet.column_dimensions['G'].width = 9

			sheet.row_dimensions[17].height = 40

			sheet['B17'] = Screening3_Language["Result SampleName Text"][language]
			sheet["B17"].font = font2
			sheet["B17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['C17'] = Screening3_Language["Result SamplePosition Text"][language]
			sheet["C17"].font = font2
			sheet["C17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['D17'] = Screening3_Language["Result SCResult Text"][language]
			sheet["D17"].font = font2
			sheet["D17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['E17'] = Screening3_Language["Result CTResult Text"][language]
			sheet["E17"].font = font2
			sheet["E17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')

			sheet['F17'] = Screening3_Language["Result GelResult Text"][language]
			sheet["F17"].font = font2
			sheet["F17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['G17'] = Screening3_Language["Result FinalResult Text"][language]
			sheet["G17"].font = font2
			sheet["G17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')

			for r in range(0, WELL_ROW):
				sheet['C' + str(r + RESULT_CELL_START)] = str(chr(65+r)) + '1'
				sheet['C' + str(r + RESULT_CELL_START + WELL_ROW)] = str(chr(65+r)) + '2'
				sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*2)] = str(chr(65+r)) + '3'
				sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*3)] = str(chr(65+r)) + '4'
				if(SC_VERSION == 48):
					sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*4)] = str(chr(65+r)) + '5'
					sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*5)] = str(chr(65+r)) + '6'

			if(SC_VERSION == 48):
				c1 = -6
				c2 = -5
				c3 = -4
				c4 = -3
				c5 = -2
				c6 = -1
			else: 
				c1 = -4
				c2 = -3
				c3 = -2
				c4 = -1
			
			if(self.base_window.main_menu.threshold_value == 0):
				for i in range(0, WELL_ROW):
					c1 = c1 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START)] = self.base_window.qualitative_analysis_2.id_list[c1]
					if(self.base_window.qualitative_analysis_2.id_list[c1]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START)] = 'N/A'
						
						sheet['E'+str(i + RESULT_CELL_START)] = 'N/A'
					else:
						if(float(self.result_label[c1]['text']) <= self.base_window.main_menu.num1):
							sheet['D'+str(i + RESULT_CELL_START)] = self.result_label[c1]['text']
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							
							sheet['E'+str(i + RESULT_CELL_START)] = '-'
							sheet['E'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						elif(float(self.result_label[c1]['text']) <= self.base_window.main_menu.num2):
							sheet['D'+str(i + RESULT_CELL_START)] = self.result_label[c1]['text']
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START)].font = font2
							sheet['B'+str(i + RESULT_CELL_START)].font = font2

							sheet['E'+str(i + RESULT_CELL_START)] = '31-35'
							sheet['E'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START)].font = font2
						elif(float(self.result_label[c1]['text']) <= self.base_window.main_menu.num3):
							sheet['D'+str(i + RESULT_CELL_START)] = self.result_label[c1]['text']
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START)].font = font2
							sheet['B'+str(i + RESULT_CELL_START)].font = font2

							sheet['E'+str(i + RESULT_CELL_START)] = '26-30'
							sheet['E'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START)] = self.result_label[c1]['text']
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START)].font = font2
							sheet['B'+str(i + RESULT_CELL_START)].font = font2

							sheet['E'+str(i + RESULT_CELL_START)] = '< 26'
							sheet['E'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START)].font = font2
						
					sheet['E'+str(i + RESULT_CELL_START)].protection = Protection(locked=False, hidden=False)
					sheet['F'+str(i + RESULT_CELL_START)].protection = Protection(locked=False, hidden=False)

					c2 = c2 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)] = self.base_window.qualitative_analysis_2.id_list[c2]
					if(self.base_window.qualitative_analysis_2.id_list[c2]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'N/A'

						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)] = 'N/A'
					else:
						if(float(self.result_label[c2]['text']) <= self.base_window.main_menu.num1):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = self.result_label[c2]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)] = '-'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						elif(float(self.result_label[c2]['text']) <= self.base_window.main_menu.num2):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = self.result_label[c2]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)] = '31-35'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						elif(float(self.result_label[c2]['text']) <= self.base_window.main_menu.num3):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = self.result_label[c2]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)] = '26-30'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = self.result_label[c2]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2

							sheet['E'+str(i + RESULT_CELL_START+ WELL_ROW)] = '< 26'
							sheet['E'+str(i + RESULT_CELL_START+ WELL_ROW)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START+ WELL_ROW)].font = font2
													
					sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)].protection = Protection(locked=False, hidden=False)
					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW)].protection = Protection(locked=False, hidden=False)

					c3 = c3 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.base_window.qualitative_analysis_2.id_list[c3]
					if(self.base_window.qualitative_analysis_2.id_list[c3]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'N/A'

						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'N/A'
					else:
						if(float(self.result_label[c3]['text']) <= self.base_window.main_menu.num1):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.result_label[c3]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)] = '-'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						elif(float(self.result_label[c3]['text']) <= self.base_window.main_menu.num2):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.result_label[c3]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)] = '31-35'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
						elif(float(self.result_label[c3]['text']) <= self.base_window.main_menu.num3):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.result_label[c3]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)] = '26-30'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.result_label[c3]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)] = '< 26'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							
					sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].protection = Protection(locked=False, hidden=False)
					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*2)].protection = Protection(locked=False, hidden=False)

					c4= c4+ WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.base_window.qualitative_analysis_2.id_list[c4]
					if(self.base_window.qualitative_analysis_2.id_list[c4]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'N/A'

						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'N/A'
					else:
						if(float(self.result_label[c4]['text']) <= self.base_window.main_menu.num1):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.result_label[c4]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = '-'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						elif(float(self.result_label[c4]['text']) <= self.base_window.main_menu.num2):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.result_label[c4]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = '31-35'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						elif(float(self.result_label[c4]['text']) <= self.base_window.main_menu.num3):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.result_label[c4]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = '26-30'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.result_label[c4]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = '< 26'
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
					   
					sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].protection = Protection(locked=False, hidden=False)
					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*3)].protection = Protection(locked=False, hidden=False)
					
					if(SC_VERSION == 48):
						c5 = c5 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.base_window.qualitative_analysis_2.id_list[c5]
						if(self.base_window.qualitative_analysis_2.id_list[c5]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'N/A'
						else:
							if(float(self.result_label[c5]['text']) <= self.base_window.main_menu.num1):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.result_label[c5]['text']
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							elif(float(self.result_label[c5]['text']) <= self.base_window.main_menu.num1 * self.base_window.main_menu.num3):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.result_label[c5]['text']
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.result_label[c5]['text']
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2

						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*4)].protection = Protection(locked=False, hidden=False)

						c6= c6 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.base_window.qualitative_analysis_2.id_list[c6]
						if(self.base_window.qualitative_analysis_2.id_list[c6]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'N/A'
						else:
							if(self.result_label[c6]['text'] != 'B'):
								if(float(self.result_label[c6]['text']) <= self.base_window.main_menu.num1):
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.result_label[c6]['text']
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
								elif(float(self.result_label[c6]['text']) <= self.base_window.main_menu.num1 * self.base_window.main_menu.num3):
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.result_label[c6]['text']
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
									sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
								else:
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.result_label[c6]['text']
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
									sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
						
						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*5)].protection = Protection(locked=False, hidden=False)
			
			else: # if(self.base_window.main_menu.threshold_value == 1)
				for i in range(0, WELL_ROW):
					c1 = c1 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START)] = self.base_window.qualitative_analysis_2.id_list[c1]
					if(self.base_window.qualitative_analysis_2.id_list[c1]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START)] = 'N/A'
					else:
						if(float(self.result_label[c1]['text']) <= self.base_window.main_menu.num2):
							sheet['D'+str(i + RESULT_CELL_START)] = self.result_label[c1]['text']
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							
							sheet['E'+str(i + RESULT_CELL_START)] = '-' 
							sheet['E'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						elif(float(self.result_label[c1]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
								sheet['D'+str(i + RESULT_CELL_START)] = self.result_label[c1]['text']
								sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START)].font = font2
								sheet['B'+str(i + RESULT_CELL_START)].font = font2

								sheet['E'+str(i + RESULT_CELL_START)] = '31-35' 
								sheet['E'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								sheet['E'+str(i + RESULT_CELL_START)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START)] = self.result_label[c1]['text']
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START)].font = font2
							sheet['B'+str(i + RESULT_CELL_START)].font = font2

							sheet['E'+str(i + RESULT_CELL_START)] = '31-35' 
							sheet['E'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['E'+str(i + RESULT_CELL_START)].font = font2
						
					sheet['E'+str(i + RESULT_CELL_START)].protection = Protection(locked=False, hidden=False)
					sheet['F'+str(i + RESULT_CELL_START)].protection = Protection(locked=False, hidden=False)

					c2 = c2 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)] = self.base_window.qualitative_analysis_2.id_list[c2]
					if(self.base_window.qualitative_analysis_2.id_list[c2]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'N/A'
					else:
						if(float(self.result_label[c2]['text']) <= self.base_window.main_menu.num2):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = self.result_label[c2]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						elif(float(self.result_label[c2]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = self.result_label[c2]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = self.result_label[c2]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
													
					sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)].protection = Protection(locked=False, hidden=False)
					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW)].protection = Protection(locked=False, hidden=False)

					c3 = c3 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.base_window.qualitative_analysis_2.id_list[c3]
					if(self.base_window.qualitative_analysis_2.id_list[c3]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'N/A'
					else:
						if(float(self.result_label[c3]['text']) <= self.base_window.main_menu.num2):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.result_label[c3]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						elif(float(self.result_label[c3]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.result_label[c3]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.result_label[c3]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							
					sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].protection = Protection(locked=False, hidden=False)
					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*2)].protection = Protection(locked=False, hidden=False)

					c4 = c4 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.base_window.qualitative_analysis_2.id_list[c4]
					if(self.base_window.qualitative_analysis_2.id_list[c4]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'N/A'
					else:
						if(float(self.result_label[c4]['text']) <= self.base_window.main_menu.num2):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.result_label[c4]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						elif(float(self.result_label[c4]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.result_label[c4]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.result_label[c4]['text']
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
					   
					sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].protection = Protection(locked=False, hidden=False)
					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*3)].protection = Protection(locked=False, hidden=False)

					if(SC_VERSION == 48):
						c5 = c5 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.base_window.qualitative_analysis_2.id_list[c5]
						if(self.base_window.qualitative_analysis_2.id_list[c5]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'N/A'
						else:
							if(float(self.result_label[c5]['text']) <= self.base_window.main_menu.num2):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.result_label[c5]['text']
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							elif(float(self.result_label[c5]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.result_label[c5]['text']
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.result_label[c5]['text']
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2

						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*4)].protection = Protection(locked=False, hidden=False)

						c6 = c6 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.base_window.qualitative_analysis_2.id_list[c6]
						if(self.base_window.qualitative_analysis_2.id_list[c6]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'N/A'
						else:
							if(self.result_label[c6]['text'] != 'B'):
								if(float(self.result_label[c6]['text']) <= self.base_window.main_menu.num2):
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.result_label[c6]['text']
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
								elif(float(self.result_label[c6]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.result_label[c6]['text']
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
									sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
								else:
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.result_label[c6]['text']
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
									sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
							
						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*5)].protection = Protection(locked=False, hidden=False)
				
			sheet.print_area = 'B1:H70'
			wb.save(self.base_window.qualitative_analysis_0.result_folder_path +  '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsx')
						
			# Report tab
			self.report_frame = ScrollableFrame2(self.report_tab)
			self.report_frame.pack(pady=5)
			
			wb = load_workbook(self.base_window.qualitative_analysis_0.result_folder_path +  '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsx')
			sheet = wb.active
			
			sample_button_list = list(range(SC_VERSION + 1))
			result_button_list = list(range(SC_VERSION + 1))
			position_button_list = list(range(SC_VERSION + 1))

			for i in range(0,SC_VERSION + 1):
				sample_pos = 'B' + str(i+(RESULT_CELL_START-1))
				sample_button_list[i] = Button(self.report_frame.scrollable_frame,
						fg = LABEL_TXT_COLOR,
						font = LABEL_TXT_FONT,
						text= sheet[sample_pos].value,
						width=30,
						bg = 'lavender',
						borderwidth = 0)
				sample_button_list[i].grid(row=i, column=0, sticky=EW, padx=1, pady=1)

				position_pos = 'C' + str(i+(RESULT_CELL_START-1))
				position_button_list[i] = Button(self.report_frame.scrollable_frame,
						fg = LABEL_TXT_COLOR,
						font = LABEL_TXT_FONT,
						text= sheet[position_pos].value,
						width=10,
						bg = 'lavender',
						borderwidth = 0)
				position_button_list[i].grid(row=i, column=1, sticky=EW, padx=1, pady=1)

				result_pos = 'D' + str(i+(RESULT_CELL_START-1))
				result_button_list[i] = Button(self.report_frame.scrollable_frame,
						fg = LABEL_TXT_COLOR,
						font = LABEL_TXT_FONT,
						text= sheet[result_pos].value,
						width=22,
						bg = 'lavender',
						borderwidth = 0)
				result_button_list[i].grid(row=i, column=3, sticky=EW, padx=1, pady=1)
				
				# if(i>0 and result_button_list[i]['text'] != 'N/A'):
				# 	if(result_button_list[i]['text'] == 'N'):
				# 		result_button_list[i]['bg'] = NEGATIVE_COLOR
				# 	elif(result_button_list[i]['text'] == 'P_L'):
				# 		result_button_list[i]['bg'] = LOW_COPY_COLOR
				# 	else:
				# 		result_button_list[i]['bg'] = POSITIVE_COLOR
				
			wb.close()
			
			if(os.path.exists(id_path + self.base_window.qualitative_analysis_2.id_file_name_label['text'] + '.xlsx')):
				try:
					shutil.move(id_path + self.base_window.qualitative_analysis_2.id_file_name_label['text'] + '.xlsx', id_old_path)
				except:
					pass

			self.automail_check()
			self.server_check()

			self.title_label['text'] = Screening3_Language["FinalTitle Label"][language]

			msg = messagebox.showinfo("", Screening3_Language["Complete Inform"][language])

		else:
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
			self.base_window.switch_page()
			self.base_window.main_menu.reset()

	def automail_check(self):
		if(self.base_window.email_setting.account_active == 1 and self.base_window.qualitative_analysis_2.automail_is_on == 1):
			# shutil.make_archive(self.base_window.qualitative_analysis_0.result_folder_path,
			# 					format='zip',
			# 					root_dir = self.base_window.qualitative_analysis_0.result_folder_path)
			# ~ try:
			AutoMail(
				self.base_window.email_setting.email_address,
				self.base_window.email_setting.email_password,
				self.base_window.qualitative_analysis_2.recipient_email,
				"Spotcheck Result",
				"This is an automatic email from Spotcheck device.",
				# self.base_window.qualitative_analysis_0.result_folder_path + '.zip',
				self.base_window.qualitative_analysis_0.result_folder_path +  '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsx',
				self.base_window.qualitative_analysis_0.experiment_name).send()
			# ~ except:
				# ~ messagebox.showerror("","ERR 04")
				# ~ pass
				
	def update_result(self):
		self.check_result_frame.destroy()
		try:
			self.save_button.destroy()
		except:
			pass
		try: 
			self.annotate_result_frame.destroy()
		except:
			pass

		if(self.quantitative_view == 0):
			self.check_result_frame = Frame(self.result_tab, bg=RESULT_TABLE_FRAME_BGD_COLOR)
			self.check_result_frame.grid(row=0, column=0, padx=76)
			self.annotate_result_frame = Frame(self.result_tab, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
			self.annotate_result_frame.grid(row=0, column=1, padx=60)
			
			negative_label = Label(self.annotate_result_frame, bg=NEGATIVE_COLOR, width=4, height=2)
			negative_label.grid(row=0, column=0, padx=20, pady=10)
			negative_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="NEGATIVE (N)", height=2)
			negative_text_label.grid(row=0, column=1, padx=20, pady=10)
			low_copy_label = Label(self.annotate_result_frame, bg=LOW_COPY_COLOR, width=4, height=2)
			low_copy_label.grid(row=1, column=0, padx=20, pady=10)
			low_copy_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="LOW COPY (P_L)", height=2)
			low_copy_text_label.grid(row=1, column=1, padx=20, pady=10)
			positive_label = Label(self.annotate_result_frame, bg=POSITIVE_COLOR, width=4, height=2)
			positive_label.grid(row=2, column=0, padx=20, pady=10)
			positive_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="POSITIVE (P_H)", height=2)
			positive_text_label.grid(row=2, column=1, padx=20, pady=10)
			na_label = Label(self.annotate_result_frame, bg=NA_COLOR, width=4, height=2)
			na_label.grid(row=3, column=0, padx=20, pady=10)
			na_copy_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="NO SAMPLE (N/A)", height=2)
			na_copy_text_label.grid(row=3, column=1, padx=20, pady=10)
			
			index = 0 
			for r in range(0, WELL_ROW):
				for c in range(0, WELL_COLUMN):
					self.result_label[index] = Label(self.check_result_frame,
										width=6,
										height=3,
										# ~ bg = RESULT_LABEL_BGD_COLOR,
										font = RESULT_LABEL_TXT_FONT_1)

					# ~ self.tooltip[i] = Pmw.Balloon(self.base_window)
					# ~ self.tooltip[i].bind(result_label[i], self.base_window.qualitative_analysis_2.id_list[i])

					if(self.thr_value.get() == 0): 
						if(self.base_window.qualitative_analysis_2.id_list[index] != 'N/A'):
							self.result_label[index]['text'] = round(self.result[index]/self.base_window.system_check.threshold,2)
							if(float(self.result_label[index]['text']) < self.base_window.main_menu.num1):
								self.result_label[index]['bg'] = NEGATIVE_COLOR
							elif(float(self.result_label[index]['text']) < self.base_window.main_menu.num1 * self.base_window.main_menu.num3):
								self.result_label[index]['bg'] = LOW_COPY_COLOR
							else:
								self.result_label[index]['bg'] = POSITIVE_COLOR

						else:
							self.result_label[index]['text'] = "N/A"
							self.result_label[index]['bg'] = NA_COLOR
						
					else:
						if(self.base_window.qualitative_analysis_2.id_list[index] != 'N/A'):
							self.result_label[index]['text'] = round(self.result[index]/self.base_window.system_check.threshold,2)
							if(float(self.result_label[index]['text']) < self.base_window.main_menu.num2):
								self.result_label[index]['bg'] = NEGATIVE_COLOR
							elif(float(self.result_label[index]['text']) < self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
								self.result_label[index]['bg'] = LOW_COPY_COLOR
							else:
								self.result_label[index]['bg'] = POSITIVE_COLOR

						else:
							self.result_label[index]['text'] = "N/A"
							self.result_label[index]['bg'] = NA_COLOR

					self.result_label[index].grid(row=r, column=c, padx=1, pady=1)

					index += 1
				

			if(self.base_window.main_menu.threshold_value != self.thr_value.get()):
				self.save_button = Button(self.button_frame,
									text = Screening3_Language["Save Button"][language],
									font = SWITCH_PAGE_BUTTON_FONT,
									width = 10,
									# height = SWITCH_PAGE_BUTTON_HEIGHT,
									bg = 'lavender',
									fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.save_clicked)
				self.save_button.pack(ipady=10, side=LEFT, padx=1)
		
			self.base_window.update_idletasks()
			sleep(1)
			if(self.base_window.main_menu.threshold_value == 0 and self.thr_value.get() == 1):
				if(os.path.exists(self.base_window.qualitative_analysis_0.result_folder_path + "/result_capture(high).jpg")):
					pass
				else:
					subprocess.call(["scrot", self.base_window.qualitative_analysis_0.result_folder_path + "/result_capture(high).jpg"])

			elif(self.base_window.main_menu.threshold_value == 1 and self.thr_value.get() == 0):
				if(os.path.exists(self.base_window.qualitative_analysis_0.result_folder_path + "/result_capture(standard).jpg")):
					pass
				else:
					subprocess.call(["scrot", self.base_window.qualitative_analysis_0.result_folder_path + "/result_capture(standard).jpg"])
		
		else: 
			self.x_result_list = list(range(SC_VERSION))
			self.y_result_list = list(range(SC_VERSION))
			self.concen_result_list = list(range(SC_VERSION))
			for i in range(0,SC_VERSION):
				if(self.base_window.qualitative_analysis_2.id_list[i] != 'N/A'):
					self.y_result_list[i] = self.result[i]/self.base_window.system_check.threshold
					self.x_result_list[i] = (self.y_result_list[i] - self.base_window.quantitative_programs_list.b_value)/self.base_window.quantitative_programs_list.a_value
					self.concen_result_list[i] = round(10**self.x_result_list[i])

			self.check_result_frame = Frame(self.result_tab, bg=RESULT_TABLE_FRAME_BGD_COLOR)
			self.check_result_frame.grid(row=0, column=0, padx=0)
			
			index = 0
			for r in range(0, WELL_ROW):
				for c in range(0, WELL_COLUMN):
					self.result_label[index] = Label(self.check_result_frame,
										width=18,
										height=3,
										# ~ bg = RESULT_LABEL_BGD_COLOR,
										font = RESULT_LABEL_TXT_FONT)
					
					if(self.base_window.qualitative_analysis_2.id_list[index] != 'N/A'):
						if(round(self.result[index]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_programs_list.n_base_value):
							self.result_label[index]['bg'] = NEGATIVE_COLOR
							self.result_label[index]['text'] = '0'
						elif(self.concen_result_list[index] < 100):
							self.result_label[index]['bg'] = LOW_COPY_COLOR
							self.result_label[index]['text'] = '< 100'
						elif(self.concen_result_list[index] <= 1000):
							self.result_label[index]['bg'] = LOW_COPY_COLOR
							self.result_label[index]['text'] = "100 - 1000"
						elif(self.concen_result_list[index] <= 5000):
							self.result_label[index]['bg'] = LOW_COPY_COLOR
							self.result_label[index]['text'] = "1001 - 5000"
						else:
							self.result_label[index]['bg'] = LOW_COPY_COLOR
							self.result_label[index]['text'] = "> 5000"
					else:
						self.result_label[index]['text'] = "N/A"
						self.result_label[index]['bg'] = NA_COLOR

					self.result_label[index].grid(row=r,column=c, padx=1, pady=1)

					index += 1

			self.save_button = Button(self.button_frame,
									text = "Save",
									font = SWITCH_PAGE_BUTTON_FONT,
									width = 10,
									# height = SWITCH_PAGE_BUTTON_HEIGHT,
									bg = 'lavender',
									fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.save_clicked)
			self.save_button.pack(ipady=10, side=LEFT, padx=1)
		
			self.base_window.update_idletasks()
			sleep(1)
			if(os.path.exists(self.base_window.qualitative_analysis_0.result_folder_path + "/quantitative.jpg")):
				pass
			else:
				subprocess.call(["scrot", self.base_window.qualitative_analysis_0.result_folder_path + "/quantitative.jpg"])

	def save_clicked(self):
		path = filedialog.asksaveasfilename(initialdir = self.base_window.qualitative_analysis_0.result_folder_path + '/', defaultextension='.xlsx')
		if path is not None:
			wb = Workbook()
			sheet = wb.active

			font0 = Font(bold=False)
			font1 = Font(size='14', bold=True, color='00FF0000')
			font2 = Font(bold=True)
			thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

			# ~ sheet.protection.sheet = True
			# ~ sheet.protection.enable()

			# ~ sheet["B8"].protection = Protection(locked=False, hidden=False)

			for i in range(RESULT_CELL_START, RESULT_CELL_START + SC_VERSION):
				sheet["B"+str(i)].font = font0
				sheet["D"+str(i)].font = font0
				sheet["E"+str(i)].font = font0

			img = Img(working_dir + "/logo.png")
			img.height = 160
			img.width = 160
			img.anchor = 'B2'
			sheet.add_image(img)

			if(self.quantitative_view == 0):
				if(self.base_window.main_menu.threshold_value == 0 and self.thr_value.get() == 1):
					img = Img(self.base_window.qualitative_analysis_0.result_folder_path + "/result_capture(high).jpg")
				elif(self.base_window.main_menu.threshold_value == 1 and self.thr_value.get() == 0):
					img = Img(self.base_window.qualitative_analysis_0.result_folder_path + "/result_capture(standard).jpg")
			else: 
				img = Img(self.base_window.qualitative_analysis_0.result_folder_path + "/quantitative.jpg")

			img.anchor = 'I17'
			sheet.add_image(img)
		
			sheet["C10"] = self.base_window.qualitative_analysis_0.template_name

			sheet.merge_cells(start_row=11, start_column=2, end_row=11, end_column=6)
			sheet["B11"] = Screening3_Language["Result Title Text"][language]
			sheet["B11"].font = font1
			sheet.cell(row=11,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			#global foldername
			sheet["B13"] = Screening3_Language["Result ExperimentName Text"][language] + self.base_window.qualitative_analysis_0.experiment_name
			sheet["B13"].font = font2
			sheet['B14'] = Screening3_Language["Result TechnicianName Text"][language] + self.base_window.qualitative_analysis_0.user_name
			sheet["B14"].font = font2
			#global covid19dir_old
			sheet['B15'] = Screening3_Language["Result Date Text"][language] + self.base_window.qualitative_analysis_0.create_time
			sheet["B15"].font = font2
			# sheet['B60'] = Screening3_Language["Result Note Text"][language]
			# sheet["B60"].font = font2
			# sheet['B61'] = Screening3_Language["Result NoSample Text"][language]
			# sheet['B62'] = Screening3_Language["Result Negative Text"][language]
			# sheet['C61'] = Screening3_Language["Result LowCopy Text"][language]
			# sheet['C62'] = Screening3_Language["Result Positive Text"][language]

			sheet.merge_cells(start_row=45, start_column=4, end_row=45, end_column=6)
			sheet.merge_cells(start_row=46, start_column=4, end_row=46, end_column=6)
			sheet['B45'] = Screening3_Language["Result TechnicianNameSign Text"][language]
			sheet['B46'] = ''
			sheet['D45'] = Screening3_Language["Result HeadOfDivisionSign Text"][language]
			sheet['D46'] = ''
			sheet["B45"].font = font2
			sheet["D45"].font = font2
			sheet["B46"].protection = Protection(locked=False, hidden=False)
			sheet["D46"].protection = Protection(locked=False, hidden=False)
			sheet.cell(row=45,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=46,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=45,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=46,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)

			for r in range(RESULT_CELL_START - 1, RESULT_CELL_START + SC_VERSION):
				for c in range(2,7):
					sheet.cell(row=r,column=c).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
					sheet.cell(row=r,column=c).border = thin_border

			sheet.column_dimensions['B'].width = 26
			sheet.column_dimensions['C'].width = 12
			sheet.column_dimensions['D'].width = 15
			sheet.column_dimensions['E'].width = 12
			sheet.column_dimensions['F'].width = 12

			sheet.row_dimensions[17].height = 40

			sheet['B17'] = Screening3_Language["Result SampleName Text"][language]
			sheet["B17"].font = font2
			sheet["B17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['C17'] = Screening3_Language["Result SamplePosition Text"][language]
			sheet["C17"].font = font2
			sheet["C17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['D17'] = Screening3_Language["Result SCResult Text"][language]
			sheet["D17"].font = font2
			sheet["D17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['E17'] = Screening3_Language["Result GelResult Text"][language]
			sheet["E17"].font = font2
			sheet["E17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['F17'] = Screening3_Language["Result FinalResult Text"][language]
			sheet["F17"].font = font2
			sheet["F17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			
			for r in range(0, WELL_ROW):
				sheet['C' + str(r + RESULT_CELL_START)] = str(chr(65+r)) + '1'
				sheet['C' + str(r + RESULT_CELL_START + WELL_ROW)] = str(chr(65+r)) + '2'
				sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*2)] = str(chr(65+r)) + '3'
				sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*3)] = str(chr(65+r)) + '4'
				if(SC_VERSION == 48):
					sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*4)] = str(chr(65+r)) + '5'
					sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*5)] = str(chr(65+r)) + '6'
			
			if(SC_VERSION == 48):
				c1 = -6
				c2 = -5
				c3 = -4
				c4 = -3
				c5 = -2
				c6 = -1
			else: 
				c1 = -4
				c2 = -3
				c3 = -2
				c4 = -1
				
			if(self.quantitative_view == 0):
				if(self.thr_value.get() == 0):
					for i in range(0, WELL_ROW):
						c1 = c1 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START)] = self.base_window.qualitative_analysis_2.id_list[c1]
						if(self.base_window.qualitative_analysis_2.id_list[c1]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START)] = 'N/A'
						else:
							if(float(self.result_label[c1]['text']) <= self.base_window.main_menu.num1):
								sheet['D'+str(i + RESULT_CELL_START)] = 'N'
								sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							elif(float(self.result_label[c1]['text']) <= self.base_window.main_menu.num1 * self.base_window.main_menu.num3):
									sheet['D'+str(i + RESULT_CELL_START)] = 'P_L'
									sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
									sheet['D'+str(i + RESULT_CELL_START)].font = font2
									sheet['B'+str(i + RESULT_CELL_START)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START)] = 'P_H'
								sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START)].font = font2
								sheet['B'+str(i + RESULT_CELL_START)].font = font2
							
						sheet['E'+str(i + RESULT_CELL_START)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START)].protection = Protection(locked=False, hidden=False)

						c2 = c2 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)] = self.base_window.qualitative_analysis_2.id_list[c2]
						if(self.base_window.qualitative_analysis_2.id_list[c2]=='N/A'):
							sheet['D'+str(i+20)] = 'N/A'
						else:
							if(float(self.result_label[c2]['text']) <= self.base_window.main_menu.num1):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'N'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							elif(float(self.result_label[c2]['text']) <= self.base_window.main_menu.num1 * self.base_window.main_menu.num3):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'P_L'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'P_H'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
														
						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW)].protection = Protection(locked=False, hidden=False)

						c3 = c3 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.base_window.qualitative_analysis_2.id_list[c3]
						if(self.base_window.qualitative_analysis_2.id_list[c3]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'N/A'
						else:
							if(float(self.result_label[c3]['text']) <= self.base_window.main_menu.num1):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'N'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							elif(float(self.result_label[c3]['text']) <= self.base_window.main_menu.num1 * self.base_window.main_menu.num3):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'P_L'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'P_H'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
								
						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*2)].protection = Protection(locked=False, hidden=False)

						c4 = c4 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.base_window.qualitative_analysis_2.id_list[c4]
						if(self.base_window.qualitative_analysis_2.id_list[c4]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'N/A'
						else:
							if(float(self.result_label[c4]['text']) <= self.base_window.main_menu.num1):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'N'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							elif(float(self.result_label[c4]['text']) <= self.base_window.main_menu.num1 * self.base_window.main_menu.num3):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'P_L'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'P_H'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						
						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*3)].protection = Protection(locked=False, hidden=False)

						if(SC_VERSION == 48):
							c5 = c5 + WELL_COLUMN
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.base_window.qualitative_analysis_2.id_list[c5]
							if(self.base_window.qualitative_analysis_2.id_list[c5]=='N/A'):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'N/A'
							else:
								if(float(self.result_label[c5]['text']) <= self.base_window.main_menu.num1):
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'N'
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
								elif(float(self.result_label[c5]['text']) <= self.base_window.main_menu.num1 * self.base_window.main_menu.num3):
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'P_L'
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
									sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
								else:
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'P_H'
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
									sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)].protection = Protection(locked=False, hidden=False)
							sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*4)].protection = Protection(locked=False, hidden=False)

							c6 = c6 + WELL_COLUMN
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.base_window.qualitative_analysis_2.id_list[c6]
							if(self.base_window.qualitative_analysis_2.id_list[c6]=='N/A'):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'N/A'
							else:
								if(self.result_label[c6]['text'] != 'B'):
									if(float(self.result_label[c6]['text']) <= self.base_window.main_menu.num1):
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'N'
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
									elif(float(self.result_label[c6]['text']) <= self.base_window.main_menu.num1 * self.base_window.main_menu.num3):
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'P_L'
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
										sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
									else:
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'P_H'
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
										sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
							
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)].protection = Protection(locked=False, hidden=False)
							sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*5)].protection = Protection(locked=False, hidden=False)
				
				else: # if(self.base_window.main_menu.threshold_value == 1)
					for i in range(0, WELL_ROW):
						c1 = c1 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START)] = self.base_window.qualitative_analysis_2.id_list[c1]
						if(self.base_window.qualitative_analysis_2.id_list[c1]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START)] = 'N/A'
						else:
							if(float(self.result_label[c1]['text']) <= self.base_window.main_menu.num2):
								sheet['D'+str(i + RESULT_CELL_START)] = 'N'
								sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							elif(float(self.result_label[c1]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
									sheet['D'+str(i + RESULT_CELL_START)] = 'P_L'
									sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
									sheet['D'+str(i + RESULT_CELL_START)].font = font2
									sheet['B'+str(i + RESULT_CELL_START)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START)] = 'P_H'
								sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START)].font = font2
								sheet['B'+str(i + RESULT_CELL_START)].font = font2
							
						sheet['E'+str(i + RESULT_CELL_START)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START)].protection = Protection(locked=False, hidden=False)

						c2 = c2 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)] = self.base_window.qualitative_analysis_2.id_list[c2]
						if(self.base_window.qualitative_analysis_2.id_list[c2]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'N/A'
						else:
							if(float(self.result_label[c2]['text']) <= self.base_window.main_menu.num2):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'N'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							elif(float(self.result_label[c2]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'P_L'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'P_H'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
														
						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW)].protection = Protection(locked=False, hidden=False)

						c3 = c3 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.base_window.qualitative_analysis_2.id_list[c3]
						if(self.base_window.qualitative_analysis_2.id_list[c3]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'N/A'
						else:
							if(float(self.result_label[c3]['text']) <= self.base_window.main_menu.num2):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'N'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							elif(float(self.result_label[c3]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'P_L'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'P_H'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
								
						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*2)].protection = Protection(locked=False, hidden=False)

						c4 = c4 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.base_window.qualitative_analysis_2.id_list[c4]
						if(self.base_window.qualitative_analysis_2.id_list[c4]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'N/A'
						else:
							if(float(self.result_label[c4]['text']) <= self.base_window.main_menu.num2):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'N'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							elif(float(self.result_label[c4]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'P_L'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'P_H'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						
						sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)].protection = Protection(locked=False, hidden=False)
						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*3)].protection = Protection(locked=False, hidden=False)

						if(SC_VERSION == 48):
							c5 = c5 + WELL_COLUMN
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.base_window.qualitative_analysis_2.id_list[c5]
							if(self.base_window.qualitative_analysis_2.id_list[c5]=='N/A'):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'N/A'
							else:
								if(float(self.result_label[c5]['text']) <= self.base_window.main_menu.num2):
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'N'
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
								elif(float(self.result_label[c5]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'P_L'
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
									sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
								else:
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'P_H'
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
									sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
									sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2

							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)].protection = Protection(locked=False, hidden=False)
							sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*4)].protection = Protection(locked=False, hidden=False)

							c6 = c6 + WELL_COLUMN
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.base_window.qualitative_analysis_2.id_list[c6]
							if(self.base_window.qualitative_analysis_2.id_list[c6]=='N/A'):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'N/A'
							else:
								if(self.result_label[c6]['text'] != 'B'):
									if(float(self.result_label[c6]['text']) <= self.base_window.main_menu.num2):
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'N'
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
									elif(float(self.result_label[c6]['text']) <= self.base_window.main_menu.num2 * self.base_window.main_menu.num3):
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'P_L'
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
										sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
									else:
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'P_H'
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
										sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
										sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
								
							sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)].protection = Protection(locked=False, hidden=False)
							sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*5)].protection = Protection(locked=False, hidden=False)
			else: 
				for i in range(0, WELL_ROW):
					c1 = c1 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START)] = self.base_window.qualitative_analysis_2.id_list[c1]
					if(self.base_window.qualitative_analysis_2.id_list[c1]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START)] = 'N/A'
					else:
						if(round(self.result[c1]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_programs_list.n_base_value):
							sheet['D'+str(i + RESULT_CELL_START)] = '0'
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START)] = '0'
						elif(self.concen_result_list[c1] < 100):
							sheet['D'+str(i + RESULT_CELL_START)] = '< 100'
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i+12)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START)].font = font2
							sheet['B'+str(i + RESULT_CELL_START)].font = font2
						elif(self.concen_result_list[c1] <= 1000):
							sheet['D'+str(i + RESULT_CELL_START)] = "100 - 1000"
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START)].font = font2
							sheet['B'+str(i + RESULT_CELL_START)].font = font2
						elif(self.concen_result_list[c1] <= 5000):
							sheet['D'+str(i + RESULT_CELL_START)] = "1001 - 5000"
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START)].font = font2
							sheet['B'+str(i + RESULT_CELL_START)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START)] = "> 5000"
							sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START)].font = font2
							sheet['B'+str(i + RESULT_CELL_START)].font = font2
							
					sheet['F'+str(i + RESULT_CELL_START)].protection = Protection(locked=False, hidden=False)

					c2 = c2 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)] = self.base_window.qualitative_analysis_2.id_list[c2]
					if(self.base_window.qualitative_analysis_2.id_list[c2]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'N/A'
					else:
						if(round(self.result[c2]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_programs_list.n_base_value):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = '0'
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START)] = '0'
						elif(self.concen_result_list[c2] < 100):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = '< 100'
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						elif(self.concen_result_list[c2] <= 1000):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = "100 - 1000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						elif(self.concen_result_list[c2] <= 5000):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = "1001 - 5000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = "> 5000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
							
					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW)].protection = Protection(locked=False, hidden=False)

					c3 = c3 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.base_window.qualitative_analysis_2.id_list[c3]
					if(self.base_window.qualitative_analysis_2.id_list[c3]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'N/A'
					else:
						if(round(self.result[c3]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_programs_list.n_base_value):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = '0'
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							# ~ sheet['E'+str(i+20)] = '0'
						elif(self.concen_result_list[c3] < 100):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = '< 100'
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i+20)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
						elif(self.concen_result_list[c3] <= 1000):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = "100 - 1000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i+20)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
						elif(self.concen_result_list[c3] <= 5000):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = "1001 - 5000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i+20)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = "> 5000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i+20)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2

					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*2)].protection = Protection(locked=False, hidden=False)

					c4 = c4 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.base_window.qualitative_analysis_2.id_list[c4]
					if(self.base_window.qualitative_analysis_2.id_list[c4]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'N/A'
					else:
						if(round(self.result[c4]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_programs_list.n_base_value):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = '0'
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = '0'
						elif(self.concen_result_list[c4] < 100):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = '< 100'
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						elif(self.concen_result_list[c4] <= 1000):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = "100 - 1000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						elif(self.concen_result_list[c4] <= 5000):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = "1001 - 5000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = "> 5000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2

					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*3)].protection = Protection(locked=False, hidden=False)

					if(SC_VERSION == 48):
						c5 = c5 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.base_window.qualitative_analysis_2.id_list[c5]
						if(self.base_window.qualitative_analysis_2.id_list[c5]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'N/A'
						else:
							if(round(self.result[c5]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_programs_list.n_base_value):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = '0'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
								# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)] = '0'
							elif(self.concen_result_list[c5] < 100):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = '< 100'
								sheet['D'+str(i+44)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.concen_result_list[c1]
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
							elif(self.concen_result_list[c5] <= 1000):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = "100 - 1000"
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.concen_result_list[c1]
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
							elif(self.concen_result_list[c5] <= 5000):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = "1001 - 5000"
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.concen_result_list[c1]
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = "> 5000"
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.concen_result_list[c1]
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2

						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*4)].protection = Protection(locked=False, hidden=False)

						c6 = c6 + WELL_COLUMN
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.base_window.qualitative_analysis_2.id_list[c6]
						if(self.base_window.qualitative_analysis_2.id_list[c6]=='N/A'):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'N/A'
						else:
							if(round(self.result[c6]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_programs_list.n_base_value):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = '0'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
								# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)] = '0'
							elif(self.concen_result_list[c6] < 100):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = '< 100'
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.concen_result_list[c1]
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
							elif(self.concen_result_list[c6] <= 1000):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = "100 - 1000"
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.concen_result_list[c1]
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
							elif(self.concen_result_list[c6] <= 5000):
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = "1001 - 5000"
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.concen_result_list[c1]
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
							else:
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = "> 5000"
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
								# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.concen_result_list[c1]
								sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
								sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2

						sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*5)].protection = Protection(locked=False, hidden=False)
			# ~ sheet['D59'] = 'B'
			# ~ sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
			sheet.print_area = 'A1:G70'
			# ~ wb.save(self.base_window.qualitative_analysis_1.analysis_result_folder + '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsm')
			wb.save(path)
			messagebox.showinfo("",Screening3_Language["SaveFile Success"][language])
				
	def thr_choose(self):
		print("thr_value = ", self.thr_value.get()) 
		if(self.thr_value.get() != 2):
			self.quantitative_view = 0
		self.base_window.update_frame()


	def quantitative_switch(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_programs_list)
		self.base_window.quantitative_programs_list.load_program()
		self.base_window.switch_page()
		
		
	def server_check(self):
		self.base_window.server_setting.check_status()
		if(self.base_window.server_setting.server_active==1):
			try:
				ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=30)
				ftp.cwd(self.base_window.server_setting.path_set + '/Processed_Data/Screening')
				file = open(self.base_window.qualitative_analysis_0.result_folder_path + '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsx','rb')
				ftp.storbinary('STOR ' + self.base_window.qualitative_analysis_0.experiment_name + ".xlsx", file)
				ftp.quit()
				
				ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=30)
				ftp.cwd(self.base_window.server_setting.path_set + '/Processed_Data/Temp')
				file = open(self.base_window.qualitative_analysis_0.result_folder_path + '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsx','rb')
				ftp.storbinary('STOR ' + self.base_window.qualitative_analysis_0.experiment_name + ".xlsx", file)
				ftp.quit()
			except Exception as e :
				messagebox.showwarning("There was an error while syncing the server",str(e))
				pass
	
	def finish_clicked(self):
		msg = messagebox.askquestion("",Screening3_Language["Finish Question"][language])
		if(msg=="yes"):
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
			self.base_window.switch_page()
			self.base_window.quantitative_programs_list.reset()
			self.base_window.main_menu.reset()


class QuantitativeAnalysisFrame3(QualitativeAnalysisFrame3):
	def __init__(self, container):
		super().__init__(container)
		self.title_label['text'] = Quantitative3_Language["Title Label"][language]

	def serial_handle(self):
		style = ttk.Style()
		style.theme_use('clam')
		style.configure("1.Horizontal.TProgressbar", troughcolor ='grey85', background='green3')
		self.progressbar = ttk.Progressbar(self.work_frame,
									style="1.Horizontal.TProgressbar",
									length = 200,
									mode = 'determinate')
		self.progressbar.pack(ipadx=2, ipady=2)

		self.process_label = Label(self.work_frame,
						bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
						fg = 'black',
						text = Quantitative3_Language["Process Label"][language],
						font = LABEL_TXT_FONT)
		self.process_label.pack(ipadx=2, ipady=2, anchor=N)
		
		if(SERIAL_COMUNICATION):
			ser.flushInput()
			ser.flushOutput()

			self.progressbar['value']=5
			self.base_window.update_idletasks()
			sleep(0.5)

			data_send = 'P'
			print("Data send:", data_send)
			ser.write(data_send.encode())

			receive_data = StringVar()
			count = 0
			bled_ready = 0
			while(receive_data != 'C'):
				if(ser.in_waiting>0):
					receive_data = ser.readline().decode('utf-8').rstrip()
					print("Data received:", receive_data)

					self.progressbar['value']=10
					self.base_window.update_idletasks()
					sleep(0.5)

					if(receive_data == 'C'):
						self.progressbar['value']=20
						self.base_window.update_idletasks()
						sleep(0.5)

						bled_ready = 1
						break;
				else:
					sleep(1)
					count += 1
					if(count > 15):
						msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value),
													ERROR_LIST(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value).name,
													icon = "error")
						break;
		else:
			bled_ready = 1

		if(bled_ready):
			GPIO.output(BLUELIGHT_PIN, GPIO.HIGH)
			
			self.progressbar['value']=45
			self.base_window.update_idletasks()
			try:
				camera_capture(self.base_window.quantitative_analysis_0.analysis_result_folder + '/raw.jpg')

				self.progressbar['value']=65
				self.base_window.update_idletasks()
			except Exception as e :
				msg = messagebox.showerror("ERR "+ str(ERROR_LIST['CAMERA_ERROR'].value),
										ERROR_LIST(ERROR_LIST['CAMERA_ERROR'].value).name,
										icon = "error")
				if(msg=='ok'):
					self.base_window.destroy()

			self.result, self.image = Process_Image(self.base_window.quantitative_analysis_0.analysis_result_folder + '/raw.jpg').process(coefficient, mode=1, well_list=self.base_window.quantitative_analysis_2.id_list)

			self.x_result_list = list(range(SC_VERSION))
			self.y_result_list = list(range(SC_VERSION))
			self.concen_result_list = list(range(SC_VERSION))
			for i in range(0,SC_VERSION):
				if(self.base_window.quantitative_analysis_2.id_list[i] != 'N/A'):
					self.y_result_list[i] = self.result[i]/self.base_window.system_check.threshold
					self.x_result_list[i] = (self.y_result_list[i] - self.base_window.quantitative_analysis_1.b_value)/self.base_window.quantitative_analysis_1.a_value
					self.concen_result_list[i] = round(10**self.x_result_list[i])
					# ~ self.concen_result_list[i] = round((1 + (self.x_result_list[i] - round(self.x_result_list[i])))*(10**round(self.x_result_list[i])))
					# ~ self.concen_result_list[i] = self.x_result_list[i]
			self.progressbar['value']=80
			self.base_window.update_idletasks() 
			sleep(0.5)
			
			GPIO.output(BLUELIGHT_PIN, GPIO.LOW)

			cv2.imwrite(self.base_window.quantitative_analysis_0.analysis_result_folder + '/process.jpg', self.image)

			sum_value = 0
			active_well_number = 0

			# Save analysis file
			wb = Workbook()
			sheet = wb.active

			sheet["A2"] = "A"
			sheet["A3"] = "B"
			sheet["A4"] = "C"
			sheet["A5"] = "D"
			sheet["B1"] = "1"
			sheet["C1"] = "2"
			sheet["D1"] = "3"
			sheet["E1"] = "4"
			if(SC_VERSION == 48):
				sheet["A6"] = "E"
				sheet["A7"] = "F"
				sheet["A8"] = "G"
				sheet["A9"] = "H"
				sheet["F1"] = "5"
				sheet["G1"] = "6"

			index = 0
			for r in range(0, WELL_ROW):
				for c in range(0, WELL_COLUMN):
					pos = str(chr(66+c)) + str(r+2)
				
					if(self.base_window.quantitative_analysis_2.id_list[index] != 'N/A'):
						sheet[pos] = round(self.result[index]/self.base_window.system_check.threshold,2)
					else:
						sheet[pos] = "N/A"

					index += 1 

			wb.save(self.base_window.quantitative_analysis_0.analysis_result_folder + "/analysis_value.xlsx")
			wb.close()

			self.progressbar['value']=95
			self.base_window.update_idletasks()
			sleep(0.5)

			self.progressbar['value']=100
			self.base_window.update_idletasks()
			sleep(0.5)

			self.progressbar.destroy()
			self.process_label.destroy()

			tab_control = ttk.Notebook(self.work_frame)
			result_tab = Frame(tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
			report_tab = Frame(tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
			image_tab = Frame(tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)

			tab_control.add(result_tab, text = Quantitative3_Language["Result Tab"][language])
			tab_control.add(report_tab, text = Quantitative3_Language["Report Tab"][language])
			tab_control.add(image_tab, text = Quantitative3_Language["Images Tab"][language])
			tab_control.grid(row=0, column=0, padx=0, pady=1, sticky=EW)

			self.check_result_frame = Frame(result_tab, bg=RESULT_TABLE_FRAME_BGD_COLOR)
			self.check_result_frame.grid(row=0, column=0, padx=0)

			# ~ Pmw.initialise(self.base_window)
			# ~ self.tooltip = list(range(SC_VERSION))
			
			result_label = list(range(SC_VERSION))
			index = 0 
			for r in range(0, WELL_ROW):
				for c in range(0, WELL_COLUMN):
					result_label[index] = Label(self.check_result_frame,
										width=18,
										height=3,
										# ~ bg = RESULT_LABEL_BGD_COLOR,
										font = RESULT_LABEL_TXT_FONT)

					# ~ self.tooltip[index] = Pmw.Balloon(self.base_window)
					# ~ self.tooltip[index].bind(result_label[index], self.base_window.quantitative_analysis_2.id_list[index])

					if(self.base_window.quantitative_analysis_2.id_list[index] != 'N/A'):
						if(round(self.result[index]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
							result_label[index]['bg'] = NEGATIVE_COLOR
							result_label[index]['text'] = '0'
						elif(self.concen_result_list[index] < 100):
							result_label[index]['bg'] = LOW_COPY_COLOR
							result_label[index]['text'] = '< 100'
						elif(self.concen_result_list[index] <= 1000):
							result_label[index]['bg'] = LOW_COPY_COLOR
							result_label[index]['text'] = "100 - 1000"
						elif(self.concen_result_list[index] <= 5000):
							result_label[index]['bg'] = LOW_COPY_COLOR
							result_label[index]['text'] = "1001 - 5000"
						else:
							result_label[index]['bg'] = LOW_COPY_COLOR
							result_label[index]['text'] = "> 5000"
					else:
						result_label[index]['text'] = "N/A"
						result_label[index]['bg'] = NA_COLOR
					result_label[index].grid(row=r,column=c, padx=1, pady=1)

					index += 1

			# ~ self.annotate_result_frame = Frame(result_tab, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
			# ~ self.annotate_result_frame.grid(row=0, column=1, padx=2)

			# ~ nagative_label = Label(self.annotate_result_frame, bg=NEGATIVE_COLOR, width=4, height=2)
			# ~ nagative_label.grid(row=0, column=0, padx=20, pady=10)
			# ~ nagative_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="NEGATIVE (N)", height=2)
			# ~ nagative_text_label.grid(row=0, column=1, padx=20, pady=10)
			# ~ low_copy_label = Label(self.annotate_result_frame, bg=LOW_COPY_COLOR, width=4, height=2)
			# ~ low_copy_label.grid(row=1, column=0, padx=20, pady=10)
			# ~ low_copy_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="LOW COPY (P_L)", height=2)
			# ~ low_copy_text_label.grid(row=1, column=1, padx=20, pady=10)
			# ~ positive_label = Label(self.annotate_result_frame, bg=POSITIVE_COLOR, width=4, height=2)
			# ~ positive_label.grid(row=2, column=0, padx=20, pady=10)
			# ~ positive_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="POSITIVE (P_H)", height=2)
			# ~ positive_text_label.grid(row=2, column=1, padx=20, pady=10)
			# ~ na_label = Label(self.annotate_result_frame, bg=NA_COLOR, width=4, height=2)
			# ~ na_label.grid(row=3, column=0, padx=20, pady=10)
			# ~ na_copy_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="NO SAMPLE (N/A)", height=2)
			# ~ na_copy_text_label.grid(row=3, column=1, padx=20, pady=10)

			
			#Image Tab
			img_labelframe_1 = LabelFrame(image_tab,
										bg="black",
										text=Quantitative3_Language["RawImage LabelFrame"][language],
										fg="cyan",
										font = LABELFRAME_TXT_FONT)
			img_labelframe_1.pack(fill=BOTH, expand=TRUE, side=LEFT)

			img_labelframe_2 = LabelFrame(image_tab,
										bg="black",
										text=Quantitative3_Language["AnalyzedImage LabelFrame"][language],
										fg="cyan",
										font = LABELFRAME_TXT_FONT)
			img_labelframe_2.pack(fill=BOTH, expand=TRUE, side=RIGHT)

			a1 = Image.open(self.base_window.quantitative_analysis_0.analysis_result_folder + '/raw.jpg')
			a1_crop = a1.crop((x1-10, y1-10, x2+10, y2+10))
			crop_width, crop_height = a1_crop.size
			scale_percent = 100
			width = int(crop_width * scale_percent / 100)
			height = int(crop_height * scale_percent / 100)
			display_img = a1_crop.resize((width,height))
			a1_display = ImageTk.PhotoImage(display_img)
			a1_label = Label(img_labelframe_1, image=a1_display, bg="black")
			a1_label.image = a1_display
			a1_label.pack(fill=BOTH, expand=TRUE)

			a2 = Image.open(self.base_window.quantitative_analysis_0.analysis_result_folder + '/process.jpg')
			a2_crop = a2.crop((x1-10, y1-10, x2+10, y2+10))
			crop_width, crop_height = a2_crop.size
			scale_percent = 100
			width = int(crop_width * scale_percent / 100)
			height = int(crop_height * scale_percent / 100)
			display_img = a2_crop.resize((width,height))
			a2_display = ImageTk.PhotoImage(display_img)
			a2_label = Label(img_labelframe_2, image=a2_display, bg="black")
			a2_label.image = a2_display
			a2_label.pack(fill=BOTH, expand=TRUE)

			# In button frame
			self.finish_button = Button(self.button_frame,
									text = Quantitative3_Language["Finish Button"][language],
									font = SWITCH_PAGE_BUTTON_FONT,
									width = 10,
									# height = SWITCH_PAGE_BUTTON_HEIGHT,
									bg = 'dodger blue',
									fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.finish_clicked)
			self.finish_button.pack(ipady=10)

			self.base_window.update_idletasks()
			self.base_window.server_setting.check_status()
			
			wb = Workbook()
			sheet = wb.active
			
			font0 = Font(bold=False)
			font1 = Font(size='14', bold=True, color='00FF0000')
			font2 = Font(bold=True)
			thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

			# sheet.protection.sheet = True
			# sheet.protection.enable()

			# sheet["B8"].protection = Protection(locked=False, hidden=False)

			for i in range(RESULT_CELL_START, RESULT_CELL_START + SC_VERSION):
				sheet["B"+str(i)].font = font0
				sheet["D"+str(i)].font = font0
				sheet["E"+str(i)].font = font0

			img = Img(working_dir + "/logo.png")
			img.height = 160
			img.width = 160
			img.anchor = 'B2'
			sheet.add_image(img)

			sheet["C10"] = self.base_window.quantitative_analysis_0.template_name

			sheet.merge_cells(start_row=11, start_column=2, end_row=11, end_column=6)
			sheet["B11"] = Quantitative3_Language["Result Title Text"][language]
			sheet["B11"].font = font1
			sheet.cell(row=11,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			#global foldername
			sheet["B13"] = Quantitative3_Language["Result ExperimentName Text"][language] + self.base_window.quantitative_analysis_1.program_name
			sheet["B13"].font = font2
			sheet['B14'] = Quantitative3_Language["Result TechnicianName Text"][language] + self.base_window.quantitative_analysis_0.user_name
			sheet["B14"].font = font2
			#global covid19dir_old
			sheet['B15'] = Quantitative3_Language["Result Date Text"][language] + self.base_window.quantitative_analysis_0.create_time
			sheet["B15"].font = font2
			# ~ sheet['B60'] = 'Note:'
			# ~ sheet["B60"].font = font2
			# ~ sheet['B61'] = '+ N/A: No sample'
			# ~ sheet['B62'] = '+ N: Negative'
			# ~ sheet['C61'] = '+ P_L: Low copy'
			# ~ sheet['C62'] = '+ P_H: Positive'

			sheet.merge_cells(start_row=45, start_column=4, end_row=45, end_column=6)
			sheet.merge_cells(start_row=46, start_column=4, end_row=46, end_column=6)
			sheet['B45'] = Quantitative3_Language["Result TechnicianNameSign Text"][language]
			sheet['B46'] = ''
			sheet['D45'] = Quantitative3_Language["Result HeadOfDivisionSign Text"][language]
			sheet['D46'] = ''
			sheet["B45"].font = font2
			sheet["D45"].font = font2
			sheet["B46"].protection = Protection(locked=False, hidden=False)
			sheet["D46"].protection = Protection(locked=False, hidden=False)
			sheet.cell(row=45,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=46,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=45,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
			sheet.cell(row=46,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)

			for r in range(RESULT_CELL_START - 1, RESULT_CELL_START + SC_VERSION):
				for c in range(2,7):
					sheet.cell(row=r,column=c).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
					sheet.cell(row=r,column=c).border = thin_border

			sheet.column_dimensions['B'].width = 26
			sheet.column_dimensions['C'].width = 12
			sheet.column_dimensions['D'].width = 15
			sheet.column_dimensions['E'].width = 12
			sheet.column_dimensions['F'].width = 12

			sheet.row_dimensions[17].height = 40

			sheet['B17'] = Quantitative3_Language["Result SampleName Text"][language]
			sheet["B17"].font = font2
			sheet["B17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['C17'] = Quantitative3_Language["Result SamplePosition Text"][language]
			sheet["C17"].font = font2
			sheet["C17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['D17'] = Quantitative3_Language["Result SCResult Text"][language]
			sheet["D17"].font = font2
			sheet["D17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['E17'] = Quantitative3_Language["Result GelResult Text"][language]
			sheet["E17"].font = font2
			sheet["E17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
			sheet['F17'] = Quantitative3_Language["Result FinalResult Text"][language]
			sheet["F17"].font = font2
			sheet["F17"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')

			for r in range(0, WELL_ROW):
				sheet['C' + str(r + RESULT_CELL_START)] = str(chr(65+r)) + '1'
				sheet['C' + str(r + RESULT_CELL_START + WELL_ROW)] = str(chr(65+r)) + '2'
				sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*2)] = str(chr(65+r)) + '3'
				sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*3)] = str(chr(65+r)) + '4'
				if(SC_VERSION == 48):
					sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*4)] = str(chr(65+r)) + '5'
					sheet['C' + str(r + RESULT_CELL_START + WELL_ROW*5)] = str(chr(65+r)) + '6'

			if(SC_VERSION == 48):
				c1 = -6
				c2 = -5
				c3 = -4
				c4 = -3
				c5 = -2
				c6 = -1
			else: 
				c1 = -4
				c2 = -3
				c3 = -2
				c4 = -1

			for i in range(0,WELL_ROW):
				c1 = c1 + WELL_COLUMN
				sheet['B'+str(i + RESULT_CELL_START)] = self.base_window.quantitative_analysis_2.id_list[c1]
				if(self.base_window.quantitative_analysis_2.id_list[c1]=='N/A'):
					sheet['D'+str(i + RESULT_CELL_START)] = 'N/A'
				else:
					if(round(self.result[c1]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
						sheet['D'+str(i + RESULT_CELL_START)] = '0'
						sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START)] = '0'
					elif(self.concen_result_list[c1] < 100):
						sheet['D'+str(i + RESULT_CELL_START)] = '< 100'
						sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START)].font = font2
						sheet['B'+str(i + RESULT_CELL_START)].font = font2
					elif(self.concen_result_list[c1] <= 1000):
						sheet['D'+str(i + RESULT_CELL_START)] = "100 - 1000"
						sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START)].font = font2
						sheet['B'+str(i + RESULT_CELL_START)].font = font2
					elif(self.concen_result_list[c1] <= 5000):
						sheet['D'+str(i + RESULT_CELL_START)] = "1001 - 5000"
						sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START)].font = font2
						sheet['B'+str(i + RESULT_CELL_START)].font = font2
					else:
						sheet['D'+str(i + RESULT_CELL_START)] = "> 5000"
						sheet['D'+str(i + RESULT_CELL_START)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START)].font = font2
						sheet['B'+str(i + RESULT_CELL_START)].font = font2
						
				sheet['F'+str(i + RESULT_CELL_START)].protection = Protection(locked=False, hidden=False)

				c2 = c2 + WELL_COLUMN
				sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)] = self.base_window.quantitative_analysis_2.id_list[c2]
				if(self.base_window.quantitative_analysis_2.id_list[c2]=='N/A'):
					sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = 'N/A'
				else:
					if(round(self.result[c2]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = '0'
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)] = '0'
					elif(self.concen_result_list[c2] < 100):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = '< 100'
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
					elif(self.concen_result_list[c2] <= 1000):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = "100 - 1000"
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
					elif(self.concen_result_list[c2] <= 5000):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = "1001 - 5000"
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
					else:
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)] = "> 5000"
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW)].font = font2
						
				sheet['F'+str(i + RESULT_CELL_START + WELL_ROW)].protection = Protection(locked=False, hidden=False)

				c3 = c3 + WELL_COLUMN
				sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.base_window.quantitative_analysis_2.id_list[c3]
				if(self.base_window.quantitative_analysis_2.id_list[c3]=='N/A'):
					sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = 'N/A'
				else:
					if(round(self.result[c3]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = '0'
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)] = '0'
					elif(self.concen_result_list[c3] < 100):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = '< 100'
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i+20)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
					elif(self.concen_result_list[c3] <= 1000):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = "100 - 1000"
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
					elif(self.concen_result_list[c3] <= 5000):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = "1001 - 5000"
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i+20)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
					else:
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)] = "> 5000"
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*2)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*2)].font = font2

				sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*2)].protection = Protection(locked=False, hidden=False)

				c4 = c4 + WELL_COLUMN
				sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.base_window.quantitative_analysis_2.id_list[c4]
				if(self.base_window.quantitative_analysis_2.id_list[c4]=='N/A'):
					sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = 'N/A'
				else:
					if(round(self.result[c4]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = '0'
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = '0'
					elif(self.concen_result_list[c4] < 100):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = '< 100'
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
					elif(self.concen_result_list[c4] <= 1000):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = "100 - 1000"
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
					elif(self.concen_result_list[c4] <= 5000):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = "1001 - 5000"
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
					else:
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)] = "> 5000"
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
						# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*3)] = self.concen_result_list[c1]
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2
						sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*3)].font = font2

				sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*3)].protection = Protection(locked=False, hidden=False)

				if(SC_VERSION == 48):
					c5 = c5 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.base_window.quantitative_analysis_2.id_list[c5]
					if(self.base_window.quantitative_analysis_2.id_list[c5]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = 'N/A'
					else:
						if(round(self.result[c5]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = '0'
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)] = '0'
						elif(self.concen_result_list[c5] < 100):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = '< 100'
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
						elif(self.concen_result_list[c5] <= 1000):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = "100 - 1000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
						elif(self.concen_result_list[c5] <= 5000):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = "1001 - 5000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)] = "> 5000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*4)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*4)].font = font2

					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*4)].protection = Protection(locked=False, hidden=False)

					c6 = c6 + WELL_COLUMN
					sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.base_window.quantitative_analysis_2.id_list[c6]
					if(self.base_window.quantitative_analysis_2.id_list[c6]=='N/A'):
						sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = 'N/A'
					else:
						if(round(self.result[c6]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = '0'
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)] = '0'
						elif(self.concen_result_list[c6] < 100):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = '< 100'
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
						elif(self.concen_result_list[c6] <= 1000):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = "100 - 1000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
						elif(self.concen_result_list[c6] <= 5000):
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = "1001 - 5000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
						else:
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)] = "> 5000"
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
							# ~ sheet['E'+str(i + RESULT_CELL_START + WELL_ROW*5)] = self.concen_result_list[c1]
							sheet['D'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2
							sheet['B'+str(i + RESULT_CELL_START + WELL_ROW*5)].font = font2

					sheet['F'+str(i + RESULT_CELL_START + WELL_ROW*5)].protection = Protection(locked=False, hidden=False)

			sheet.print_area = 'A1:G70'
			
			wb.save(self.base_window.quantitative_analysis_0.result_folder_path + '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsx')

			# Report tab
			self.report_frame = ScrollableFrame2(report_tab)
			self.report_frame.pack(pady=5)
			
			wb = load_workbook(self.base_window.quantitative_analysis_0.result_folder_path +  '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsx')
			sheet = wb.active
			
			sample_button_list = list(range(SC_VERSION + 1))
			result_button_list = list(range(SC_VERSION + 1))
			position_button_list = list(range(SC_VERSION + 1))

			for i in range(0,SC_VERSION + 1):
				sample_pos = 'B' + str(i+(RESULT_CELL_START-1))
				sample_button_list[i] = Button(self.report_frame.scrollable_frame,
						fg = LABEL_TXT_COLOR,
						font = LABEL_TXT_FONT,
						text= sheet[sample_pos].value,
						width=30,
						bg = 'lavender',
						borderwidth = 0)
				sample_button_list[i].grid(row=i, column=0, sticky=EW, padx=1, pady=1)

				position_pos = 'C' + str(i+(RESULT_CELL_START-1))
				position_button_list[i] = Button(self.report_frame.scrollable_frame,
						fg = LABEL_TXT_COLOR,
						font = LABEL_TXT_FONT,
						text= sheet[position_pos].value,
						width=10,
						bg = 'lavender',
						borderwidth = 0)
				position_button_list[i].grid(row=i, column=1, sticky=EW, padx=1, pady=1)

				result_pos = 'D' + str(i+(RESULT_CELL_START-1))
				result_button_list[i] = Button(self.report_frame.scrollable_frame,
						fg = LABEL_TXT_COLOR,
						font = LABEL_TXT_FONT,
						text= sheet[result_pos].value,
						width=22,
						bg = 'lavender',
						borderwidth = 0)
				result_button_list[i].grid(row=i, column=3, sticky=EW, padx=1, pady=1)
				
				# if(i>0 and result_button_list[i]['text'] != 'N/A'):
				# 	if(result_button_list[i]['text'] == '0'):
				# 		result_button_list[i]['bg'] = NEGATIVE_COLOR
				# 	else:
				# 		result_button_list[i]['bg'] = LOW_COPY_COLOR
			
			wb.close()
			
			if(os.path.exists(id_path + self.base_window.quantitative_analysis_2.id_file_name_label['text'] + '.xlsx')):
				try:
					shutil.move(id_path + self.base_window.quantitative_analysis_2.id_file_name_label['text'] + '.xlsx', id_old_path)
				except:
					pass
					
			self.automail_check()
			self.server_check()

			self.title_label['text'] = Quantitative3_Language["FinalTitle Label"][language]

			msg = messagebox.showinfo("", Quantitative3_Language["Complete Inform"][language])				

		else:
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
			self.base_window.switch_page()
			self.base_window.main_menu.reset()

	def finish_clicked(self):
		wb = load_workbook(self.base_window.quantitative_analysis_0.result_folder_path +  '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsx')
		sheet = wb.active
		subprocess.call(["scrot", self.base_window.quantitative_analysis_0.result_folder_path + "/result_capture.jpg"])
		img = Img(self.base_window.quantitative_analysis_0.result_folder_path + "/result_capture.jpg")
		img.anchor = 'I17'
		sheet.add_image(img)
		wb.save(self.base_window.quantitative_analysis_0.result_folder_path + '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsx')
		wb.close()
		
		msg = messagebox.askquestion("",Quantitative3_Language["Finish Question"][language])
		if(msg=="yes"):
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
			self.base_window.switch_page()
			self.base_window.main_menu.reset()

	def automail_check(self):
		if(self.base_window.email_setting.account_active == 1 and self.base_window.quantitative_analysis_2.automail_is_on == 1):
			# shutil.make_archive(self.base_window.qualitative_analysis_0.result_folder_path,
			# 					format='zip',
			# 					root_dir = self.base_window.qualitative_analysis_0.result_folder_path)
			# ~ try:
			AutoMail(
				self.base_window.email_setting.email_address,
				self.base_window.email_setting.email_password,
				self.base_window.quantitative_analysis_2.recipient_email,
				"Spotcheck Result",
				"This is an automatic email from Spotcheck device.",
				# self.base_window.qualitative_analysis_0.result_folder_path + '.zip',
				self.base_window.quantitative_analysis_0.result_folder_path +  '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsx',
				self.base_window.quantitative_analysis_0.experiment_name).send()
			# ~ except:
				# ~ messagebox.showerror("","ERR 04")
				# ~ pass
		
	def server_check(self):
		self.base_window.server_setting.check_status()
		if(self.base_window.server_setting.server_active==1):
			try:
				ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=30)
				ftp.cwd(self.base_window.server_setting.path_set + '/Processed_Data/Quantitative')
				file = open(self.base_window.quantitative_analysis_0.result_folder_path + '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsx','rb')
				ftp.storbinary('STOR ' + self.base_window.quantitative_analysis_0.experiment_name + ".xlsx", file)
				ftp.quit()
				
				ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=30)
				ftp.cwd(self.base_window.server_setting.path_set + '/Processed_Data/Temp')
				file = open(self.base_window.quantitative_analysis_0.result_folder_path + '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsx','rb')
				ftp.storbinary('STOR ' + self.base_window.quantitative_analysis_0.experiment_name + ".xlsx", file)
				ftp.quit()
			except Exception as e :
				messagebox.showwarning("There was an error while syncing the server",str(e))
				pass
		
		# ~ self.base_window.server_setting.check_status()
		# ~ if(self.base_window.server_setting.server_active==1):
			# ~ try:
				# ~ ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=30)
				# ~ ftp.cwd(self.base_window.server_setting.path_set + '/Processed_Data/Quantitative')
				# ~ file = open(self.base_window.quantitative_analysis_1.result_folder_path + '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsx','rb')
				# ~ ftp.storbinary('STOR ' + self.base_window.quantitative_analysis_0.experiment_name + ".xlsx", file)
				# ~ ftp.quit()
			# ~ except Exception as e :
				# ~ messagebox.showwarning("There was an error while syncing the server",str(e))
				# ~ pass
				
class QualitativeAnalysisFrame2(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.id_list = list(range(SC_VERSION))

		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.work_frame.pack_propagate(0)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X, expand=TRUE)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = Screening2_Language["Title Label"][language],
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		self.id_load_frame = LabelFrame(self.work_frame,
									text = Screening2_Language["SamplesFile LabelFrame"][language],
									fg = LABEL_FRAME_TXT_COLOR,
									font = LABEL_FRAME_TXT_FONT,
									bg = LABEL_FRAME_BGD_COLOR)

		self.id_load_frame.grid(row=0, column=0, pady=146, padx=50)

		self.id_pos_frame = Frame(self.work_frame,
							bg = LABEL_FRAME_BGD_COLOR)

		self.id_pos_frame.grid(row=0, column=1, padx=50)

		self.id_file_name_label = Label(self.id_load_frame,
								   bg = LABEL_BGD_COLOR,
								   fg = LABEL_TXT_COLOR,
								   font = ('Helvetica', 13))
		self.id_file_name_label.pack(expand=TRUE, fill=BOTH, pady=10, padx=10)

		self.load_button = Button(self.id_load_frame,
								  text = Screening2_Language["Load Button"][language],
								  bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
								  fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
								  font = MAIN_FUCNTION_BUTTON_FONT,
								  width = 15,
								  height = 3,
								  borderwidth = 0,
								  command = self.load_clicked)
		self.load_button.pack(expand=TRUE, side=LEFT)

		self.create_button = Button(self.id_load_frame,
								  text = Screening2_Language["Create Button"][language],
								  bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
								  fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
								  font = MAIN_FUCNTION_BUTTON_FONT,
								  width = 15,
								  height = 3,
								  borderwidth = 0,
								  command = self.create_clicked)
		self.create_button.pack(expand=TRUE, side=RIGHT)



		# In button frame
		self.back_button = Button(self.button_frame,
								text = Screening2_Language["Back Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(side=LEFT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

		self.next_button = Button(self.button_frame,
								text = Screening2_Language["Next Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.next_clicked)
		self.next_button.pack(side=RIGHT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)
		
	def back_clicked(self):
		self.base_window.frame_list.remove(self.base_window.system_check)
		del self.base_window.system_check
		self.base_window.system_check = SystemCheckFrame(self.base_window)
		self.base_window.frame_list.append(self.base_window.system_check)

		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
		self.base_window.switch_page()

	def next_clicked(self):
		try:
			 self.email_label_frame.place_forget()
		except:
			pass

		if(self.id_file_name_label['text'] == ""):
			messagebox.showwarning("",Screening2_Language["SamplesFile Empty"][language])
		else:
			self.base_window.server_setting.check_status()
			self.base_window.email_setting.check_status()
			if(self.base_window.email_setting.account_active == 1):
				msg = messagebox.askquestion("", Screening2_Language["Email Confirm"][language])
				if(msg=='yes'):
					self.email_label_frame = LabelFrame(self.work_frame,
														width = 200,
														height = 100,
														text = "Recipient email",
														bg = 'grey75')
					self.email_label_frame.place(x=200, y=150)

					self.email_entry = Entry(self.email_label_frame, width=30, justify='right', font=('Courier',14))
					self.email_entry.pack(padx=10, pady=10)
					
					self.email_entry.insert(0, autofill_email.strip('\n'))
					
					self.ok_button = Button(self.email_label_frame,
								text = "OK",
								font = SWITCH_PAGE_BUTTON_FONT,
								width = 5,
								height = 2,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.ok_clicked)
					self.ok_button.pack(side=LEFT, padx=20, pady=10, ipady=10, ipadx=20)

					self.cancel_button = Button(self.email_label_frame,
								text = "Cancel",
								font = SWITCH_PAGE_BUTTON_FONT,
								width = 5,
								height = 2,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.cancel_clicked)
					self.cancel_button.pack(side=RIGHT, padx=20, pady=10, ipady=10, ipadx=20)

				else:
					self.automail_is_on = 0
					self.base_window.forget_page()
					self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_3)
					self.base_window.switch_page()
					self.base_window.qualitative_analysis_3.serial_handle()
			else:
				self.automail_is_on = 0
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_3)
				self.base_window.switch_page()
				self.base_window.qualitative_analysis_3.serial_handle()

	def ok_clicked(self):
		if(self.email_entry.get() == ''):
			messagebox.showwarning("", Screening2_Language["Email Empty"][language])
		else:
			addressToVerify = self.email_entry.get()
			# ~ match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', addressToVerify)
			# ~ if (match == None):
				# ~ messagebox.showerror("","Email syntax error")
			# ~ else:
			
			global autofill_email
			autofill_email = addressToVerify
			fw_info = open('/home/pi/Spotcheck/.oldinfo.txt', 'w')
			fw_info.writelines(addressToVerify + '\n')
			fw_info.writelines(autofill_user + '\n')
			fw_info.close()
			
			self.recipient_email = addressToVerify
			self.automail_is_on = 1
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_3)
			self.base_window.switch_page()
			self.base_window.qualitative_analysis_3.serial_handle()
			
			
	

	def cancel_clicked(self):
		self.email_label_frame.place_forget()

	def load_clicked(self):
		self.server_check()
		path = filedialog.askopenfilename(initialdir=id_path, filetypes=[('Excel file','*.xlsm *.xlsx *.xls')])
		self.id_file_path = path
		if path is not None:
			try:
				wb = load_workbook(path)
				sheet = wb.active
				# ~ for i in range(0,48):
					# ~ pos = 'B' + str(i+12)
					# ~ self.id_list[i] = sheet[pos].value
				
				if(SC_VERSION == 48):
					self.id_list[0] = sheet["B12"].value
					self.id_list[1] = sheet["B20"].value
					self.id_list[2] = sheet["B28"].value
					self.id_list[3] = sheet["B36"].value
					self.id_list[4] = sheet["B44"].value
					self.id_list[5] = sheet["B52"].value

					self.id_list[6] = sheet["B13"].value
					self.id_list[7] = sheet["B21"].value
					self.id_list[8] = sheet["B29"].value
					self.id_list[9] = sheet["B37"].value
					self.id_list[10] = sheet["B45"].value
					self.id_list[11] = sheet["B53"].value
					
					self.id_list[12] = sheet["B14"].value
					self.id_list[13] = sheet["B22"].value
					self.id_list[14] = sheet["B30"].value
					self.id_list[15] = sheet["B38"].value
					self.id_list[16] = sheet["B46"].value
					self.id_list[17] = sheet["B54"].value
					
					self.id_list[18] = sheet["B15"].value
					self.id_list[19] = sheet["B23"].value
					self.id_list[20] = sheet["B31"].value
					self.id_list[21] = sheet["B39"].value
					self.id_list[22] = sheet["B47"].value
					self.id_list[23] = sheet["B55"].value
					
					self.id_list[24] = sheet["B16"].value
					self.id_list[25] = sheet["B24"].value
					self.id_list[26] = sheet["B32"].value
					self.id_list[27] = sheet["B40"].value
					self.id_list[28] = sheet["B48"].value
					self.id_list[29] = sheet["B56"].value
					
					self.id_list[30] = sheet["B17"].value
					self.id_list[31] = sheet["B25"].value
					self.id_list[32] = sheet["B33"].value
					self.id_list[33] = sheet["B41"].value
					self.id_list[34] = sheet["B49"].value
					self.id_list[35] = sheet["B57"].value
					
					self.id_list[36] = sheet["B18"].value
					self.id_list[37] = sheet["B26"].value
					self.id_list[38] = sheet["B34"].value
					self.id_list[39] = sheet["B42"].value
					self.id_list[40] = sheet["B50"].value
					self.id_list[41] = sheet["B58"].value
					
					self.id_list[42] = sheet["B19"].value
					self.id_list[43] = sheet["B27"].value
					self.id_list[44] = sheet["B35"].value
					self.id_list[45] = sheet["B43"].value
					self.id_list[46] = sheet["B51"].value
					self.id_list[47] = sheet["B59"].value
				
				else:
					self.id_list[0] = sheet["B12"].value
					self.id_list[1] = sheet["B16"].value
					self.id_list[2] = sheet["B20"].value
					self.id_list[3] = sheet["B24"].value
					
					self.id_list[4] = sheet["B13"].value
					self.id_list[5] = sheet["B17"].value
					self.id_list[6] = sheet["B21"].value
					self.id_list[7] = sheet["B25"].value

					self.id_list[8] = sheet["B14"].value
					self.id_list[9] = sheet["B18"].value
					self.id_list[10] = sheet["B22"].value
					self.id_list[11] = sheet["B26"].value

					self.id_list[12] = sheet["B15"].value
					self.id_list[13] = sheet["B19"].value
					self.id_list[14] = sheet["B23"].value
					self.id_list[15] = sheet["B27"].value

				tmp = 0
				for i in range(len(path)):
					if(path[i]=='/'):
						tmp = i+1
				file_name = path[tmp:(len(path)-5)]

				self.id_file_name_label['text'] = file_name
				self.id_file_name_label['bg'] = 'lawn green'

				try:
					for i in range(0,SC_VERSION):
						self.id_label[i].destroy()
				except:
					pass

				# ~ Pmw.initialise(self.base_window)
				# ~ self.tooltip = list(range(48))

				self.id_label = list(range(SC_VERSION))
				index = 0
				for r in range(0, WELL_ROW):
					for c in range(0, WELL_COLUMN):
						self.id_label[index] = Label(self.id_pos_frame,
												width=6,
												height=3,
												text = self.id_list[index],
												# ~ bg = RESULT_LABEL_BGD_COLOR,
												font = RESULT_LABEL_TXT_FONT)

						# ~ self.tooltip[index] = Pmw.Balloon(self.base_window)
						# ~ self.tooltip[index].bind(self.id_label[index], self.id_list[index])

						if(self.id_list[index] != 'N/A'):
							self.id_label[index]['bg'] = "lawn green"
						else:
							self.id_label[index]['bg'] = "grey80"

						self.id_label[index].grid(row=r, column=c, padx=1, pady=1)

						index += 1

				msg = messagebox.askokcancel("",Screening2_Language["AllowPutSample Inform"][language])

			except:
				pass

	def create_clicked(self):
		del self.base_window.id_create
		self.base_window.id_create = IDCreateFrame(self.base_window)
		self.base_window.frame_list.append(self.base_window.id_create)

		self.base_window.id_create.direct_create = 1
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.id_create)
		self.base_window.switch_page()
	
	def server_check(self):
		# ~ #Check if the server is connected
		self.base_window.email_setting.check_status()
		if(self.base_window.server_setting.server_active==1):
			try:
				#FTP:
				print("Connect to " + self.base_window.server_setting.ip_set + '/' + self.base_window.server_setting.path_set + '...')
				ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=15)
				print("Done")
				ftp.cwd(self.base_window.server_setting.path_set + '/Unprocessed_Data')
				ftp_files = ftp.nlst()
				for ftp_file in ftp_files:
					if(os.path.exists(id_path + ftp_file)):
						pass
					elif(os.path.exists(id_old_path + ftp_file)):
						pass
					else:
						local_folder = os.path.join(id_path, ftp_file)
						file = open(local_folder,'wb')
						ftp.retrbinary('RETR ' + ftp_file, file.write)
						file.close()
						print(ftp_file, "download done!")
				ftp.quit()
				
			except Exception as e:
				messagebox.showerror(Screening2_Language["ServerCheck Error"][language],str(e))
				pass


class QuantitativeAnalysisFrame2(QualitativeAnalysisFrame2):
	def __init__(self, container):
		super().__init__(container)
		self.title_label['text'] = Quantitative2_Language["Title Label"][language]
		self.id_load_frame['text'] = Quantitative2_Language["SamplesFile LabelFrame"][language]
		self.load_button['text'] = Quantitative2_Language["Load Button"][language]
		self.create_button['text'] = Quantitative2_Language["Create Button"][language]
		self.back_button['text'] = Quantitative2_Language["Back Button"][language]
		self.next_button['text'] = Quantitative2_Language["Next Button"][language]


	def back_clicked(self):
		self.base_window.frame_list.remove(self.base_window.system_check)
		del self.base_window.system_check
		self.base_window.system_check = SystemCheckFrame(self.base_window)
		self.base_window.frame_list.append(self.base_window.system_check)

		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_1)
		self.base_window.switch_page()

	def next_clicked(self):
		try:
			 self.email_label_frame.place_forget()
		except:
			pass

		if(self.id_file_name_label['text'] == ""):
			messagebox.showwarning("", Quantitative2_Language["SamplesFile Empty"][language])
		else:
			self.base_window.email_setting.check_status()
			if(self.base_window.email_setting.account_active == 1):
				msg = messagebox.askquestion("", Quantitative2_Language["Email Confirm"][language])
				if(msg=='yes'):
					self.email_label_frame = LabelFrame(self.work_frame,
														width = 100,
														height = 50,
														text = "Recipient email",
														bg = 'dodger blue')
					self.email_label_frame.place(x=200, y=150)

					self.email_entry = Entry(self.email_label_frame, width=30, justify='right', font=('Courier',14))
					self.email_entry.pack()

					self.ok_button = Button(self.email_label_frame,
								text = Quantitative2_Language["OK Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								width = 5,
								height = 2,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.ok_clicked)
					self.ok_button.pack(side=LEFT, padx=20, pady=2, ipady=10, ipadx=20)

					self.cancel_button = Button(self.email_label_frame,
								text = Quantitative2_Language["Cancel Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								width = 5,
								height = 2,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.cancel_clicked)
					self.cancel_button.pack(side=RIGHT, padx=20, pady=2, ipady=10, ipadx=20)

				else:
					self.automail_is_on = 0
					self.base_window.forget_page()
					self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_3)
					self.base_window.switch_page()
					self.base_window.quantitative_analysis_3.serial_handle()
			else:
				self.automail_is_on = 0
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_3)
				self.base_window.switch_page()
				self.base_window.quantitative_analysis_3.serial_handle()

	def ok_clicked(self):
		if(self.email_entry.get() == ''):
			messagebox.showwarning("",Quantitative2_Language["Email Empty"][language])
		else:
			addressToVerify = self.email_entry.get()
			# ~ match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', addressToVerify)
			# ~ if (match == None):
				# ~ messagebox.showerror("","Email syntax error")
			# ~ else:
			self.recipient_email = addressToVerify
			self.automail_is_on = 1
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_3)
			self.base_window.switch_page()
			self.base_window.quantitative_analysis_3.serial_handle()

	def cancel_clicked(self):
		self.email_label_frame.place_forget()


	def create_clicked(self):
		del self.base_window.id_create
		self.base_window.id_create = IDCreateFrame(self.base_window)
		self.base_window.frame_list.append(self.base_window.id_create)

		self.base_window.id_create.direct_create = 2
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.id_create)
		self.base_window.switch_page()

class QualitativeAnalysisFrame1(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.work_frame.pack_propagate(0)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X, expand=TRUE)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = "ANALYSIS",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		self.program_frame = ScrollableFrame(self.work_frame)
		self.program_frame.grid(row=0, column=0, pady=20)

		self.info_labelframe = LabelFrame(self.work_frame,
								text = "Information",
								font = LABELFRAME_TXT_FONT,
								bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.info_labelframe.grid(row=0, column=1, rowspan=2, ipadx=10, ipady=5, padx=10, pady=28)

		experiment_name_label = Label(self.info_labelframe,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							text = "Kit name:",
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT)
		experiment_name_label.grid(row=0, column=0, padx=10, sticky=E)


		self.experiment_name_text = Text(self.info_labelframe,
							width = 30,
							height = 1,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT,)
							# ~ state = 'disabled')
		self.experiment_name_text.grid(row=0, column=1, padx=5, pady=10)

		user_name_label = Label(self.info_labelframe,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							text = "User name:",
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT)
		user_name_label.grid(row=1, column=0, padx=10, sticky=E)

		self.user_name_text= Text(self.info_labelframe,
							width = 30,
							height = 1,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT,
							state = 'disabled')
		self.user_name_text.grid(row=1, column=1, padx=5, pady=10)

		pfi_label = Label(self.info_labelframe,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							text = "PFi value:",
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT)
		pfi_label.grid(row=2, column=0, padx=10, pady=10, sticky=NE)

		self.pfi_text = Text(self.info_labelframe,
							width = 30,
							height = 1,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT,
							state = 'disabled')
		self.pfi_text.grid(row=2, column=1, padx=5, pady=10)

		comment_label = Label(self.info_labelframe,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							text = "Comment:",
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT)
		comment_label.grid(row=3, column=0, padx=10, pady=10, sticky=NE)

		self.comment_text = Text(self.info_labelframe,
							width = 30,
							height = 12,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT,
							state = 'disabled')
		self.comment_text.grid(row=3, column=1, padx=5, pady=10)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = "Back",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(side=LEFT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

		self.next_button = Button(self.button_frame,
								text = "Next",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.next_clicked)
		self.next_button.pack(side=RIGHT, padx=0, pady=0, ipady=10, ipadx=30, anchor=E)

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
		self.base_window.switch_page()
	def next_clicked(self):
		self.program_name = self.experiment_name_text.get("1.0","end-1c")
		# ~ self.program_name = self.base_window.qualitative_analysis_0.experiment_name
		if(len(self.program_name) != 0):
			# ~ msg = messagebox.askokcancel("","Please make sure no sample is placed in the device !")
			# ~ if(msg == True):
			if os.path.exists(results_qualitative_path + self.program_name):
				self.program_path = results_qualitative_path + self.program_name
			else:
				self.program_path = os.path.join(results_qualitative_path, self.program_name)
				os.mkdir(self.program_path)

			self.create_time = strftime(" %y-%m-%d %H.%M.%S")
			self.result_folder_name = self.base_window.qualitative_analysis_0.experiment_name + self.create_time
			self.result_folder_path = os.path.join(self.program_path + '/', self.result_folder_name)
			os.mkdir(self.result_folder_path)
			print(self.result_folder_path)

			self.system_check_folder = os.path.join(self.result_folder_path, "System_Check")
			os.mkdir(self.system_check_folder)

			self.analysis_result_folder = os.path.join(self.result_folder_path, "Analysis Results")
			os.mkdir(self.analysis_result_folder)

			# ~ self.base_window.system_check.mode_check = 1
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_2)
			self.base_window.switch_page()
			# ~ self.base_window.system_check.serial_handle()

		else:
			messagebox.showwarning("","Please select the kit !")

	def load_program(self):
		try:
			for i in range(len(self.program_button)):
				self.program_button[i].destroy()
		except:
			pass
		self.program_button = list(range(100))
		for file in os.listdir(programs_qualitative_path):
			self.program_button[os.listdir(programs_qualitative_path).index(file)] = Button(self.program_frame.scrollable_frame,
									text=file[:(len(file)-5)],
									font = PROGRAM_BUTTON_TXT_FONT,
									bg = PROGRAM_BUTTON_BGD_COLOR,
									fg = PROGRAM_BUTTON_TXT_COLOR,
									width = 51,
									borderwidth = 0)
			self.program_button[os.listdir(programs_qualitative_path).index(file)]['command'] = partial(self.program_clicked, os.listdir(programs_qualitative_path).index(file))
			self.program_button[os.listdir(programs_qualitative_path).index(file)].pack(pady=2, ipady=5, fill=BOTH, expand=TRUE)

	def program_clicked(self, button_index):
		wb = load_workbook(programs_qualitative_path + self.program_button[button_index]['text'] + ".xlsx")
		sheet = wb.active

		self.program_base_value = float(sheet["E2"].value)

		self.experiment_name_text['state'] = "normal"
		self.experiment_name_text.delete('1.0',END)
		try:
			self.experiment_name_text.insert('1.0', self.program_button[button_index]['text'])
		except:
			pass
		self.experiment_name_text['state'] = "disabled"

		self.user_name_text['state'] = "normal"
		self.user_name_text.delete('1.0',END)
		try:
			self.user_name_text.insert('1.0', sheet["C2"].value)
		except:
			pass
		self.user_name_text['state'] = "disabled"

		self.pfi_text['state'] = "normal"
		self.pfi_text.delete('1.0',END)
		try:
			self.pfi_text.insert('1.0', sheet["E2"].value)
		except:
			pass
		self.pfi_text['state'] = "disabled"

		self.comment_text['state'] = "normal"
		self.comment_text.delete('1.0',END)
		try:
			self.comment_text.insert('1.0', sheet["C3"].value)
		except:
			pass
		self.comment_text['state'] = "disabled"

		wb.close()

class QuantitativeAnalysisFrame1(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.work_frame.pack_propagate(0)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X, expand=TRUE)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = Quantitative1_Language["Title Label"][language],
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		self.program_frame = ScrollableFrame(self.work_frame)
		self.program_frame.grid(row=0, column=0, pady=20)

		self.info_labelframe = LabelFrame(self.work_frame,
								text = Quantitative1_Language["Information LabelFrame"][language],
								font = LABELFRAME_TXT_FONT,
								bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.info_labelframe.grid(row=0, column=1, rowspan=1, ipadx=10, ipady=5, padx=10, pady=69)

		self.experiment_name_label = Label(self.info_labelframe,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							text = Quantitative1_Language["KitName Label"][language],
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT)
		self.experiment_name_label.grid(row=0, column=0, padx=10, sticky=E)


		self.experiment_name_text = Text(self.info_labelframe,
							width = 30,
							height = 1,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT,
							state = 'disabled')
		self.experiment_name_text.grid(row=0, column=1, padx=5, pady=10)

		self.status_label = Label(self.info_labelframe,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							text = Quantitative1_Language["Parameters Label"][language],
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT)
		self.status_label.grid(row=3, column=0, padx=10, pady=10, sticky=NE)

		self.status_text = Text(self.info_labelframe,
							width = 30,
							height = 12,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT,
							state = 'disabled')
		self.status_text.grid(row=3, column=1, padx=5, pady=10)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = Quantitative1_Language["Back Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(side=LEFT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

		self.next_button = Button(self.button_frame,
								text = Quantitative1_Language["Next Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.next_clicked)
		self.next_button.pack(side=RIGHT, padx=0, pady=0, ipady=10, ipadx=30, anchor=E)
		

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_0)
		self.base_window.switch_page()

	def next_clicked(self):
		self.program_name = self.experiment_name_text.get("1.0","end-1c")
		if(len(self.program_name) != 0):
			# ~ msg = messagebox.askokcancel("","Please make sure no sample is placed in the device !")
			# ~ if(msg == True):
				# ~ if os.path.exists(results_quantitative_path + self.program_name):
					# ~ self.program_path = results_quantitative_path + self.program_name
				# ~ else:
					# ~ self.program_path = os.path.join(results_quantitative_path, self.program_name)
					# ~ os.mkdir(self.program_path)

				# ~ self.create_time = strftime(" %y-%m-%d %H.%M.%S")
				# ~ self.result_folder_name = self.program_name + self.create_time
				# ~ self.result_folder_path = os.path.join(self.program_path + '/', self.result_folder_name)
				# ~ os.mkdir(self.result_folder_path)
				# ~ print(self.result_folder_path)

				# ~ self.system_check_folder = os.path.join(self.result_folder_path, "System_Check")
				# ~ os.mkdir(self.system_check_folder)

				# ~ self.analysis_result_folder = os.path.join(self.result_folder_path, "Analysis Results")
				# ~ os.mkdir(self.analysis_result_folder)

				# ~ self.base_window.system_check.mode_check = 3
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_2)
				self.base_window.switch_page()
				# ~ self.base_window.system_check.serial_handle()

		else:
			messagebox.showwarning("",Quantitative1_Language["Kit Empty"][language])

	def load_program(self):
		try:
			for i in range(len(self.program_button)):
				self.program_button[i].destroy()
		except:
			pass
		self.program_button = list(range(100))
		for file in os.listdir(programs_quantitative_path):
			self.program_button[os.listdir(programs_quantitative_path).index(file)] = Button(self.program_frame.scrollable_frame,
									text=file[:(len(file)-5)],
									font = PROGRAM_BUTTON_TXT_FONT,
									bg = PROGRAM_BUTTON_BGD_COLOR,
									fg = PROGRAM_BUTTON_TXT_COLOR,
									width = 51,
									borderwidth = 0)
			self.program_button[os.listdir(programs_quantitative_path).index(file)]['command'] = partial(self.program_clicked, os.listdir(programs_quantitative_path).index(file))
			self.program_button[os.listdir(programs_quantitative_path).index(file)].pack(pady=2, ipady=5, fill=BOTH, expand=TRUE)

	def program_clicked(self, button_index):
		wb = load_workbook(programs_quantitative_path + self.program_button[button_index]['text'] + ".xlsx")
		sheet = wb.active

		# ~ self.n_base_value = float(sheet["I2"].value)
		# ~ self.a_value = float(sheet["E2"].value)
		# ~ self.b_value = float(sheetn["G2"].value)
		
		self.n_base_value = float(sheet["A1"].value)
		
		concen1_enalble = 1
		concen2_enalble = 1
		concen3_enalble = 1
		concen4_enalble = 1
		concen5_enalble = 1
		try:
			self.value1 = float(sheet["C2"].value)
			self.concen1 = round(math.log10(float(sheet["B2"].value)),3)
			concen_1_pt = [self.concen1, self.value1]
		except:
			concen1_enalble = 0
			pass
			
		try:
			self.value2 = float(sheet["C3"].value)
			self.concen2 = round(math.log10(float(sheet["B3"].value)),3)
			concen_2_pt = [self.concen2, self.value2]
		except:
			concen2_enalble = 0
			pass
			
		try:
			self.value3 = float(sheet["C4"].value)
			self.concen3 = round(math.log10(float(sheet["B4"].value)),3)
			concen_3_pt = [self.concen3, self.value3]
		except:
			concen3_enalble = 0
			pass
			
		try:
			self.value4 = float(sheet["C5"].value)
			self.concen4 = round(math.log10(float(sheet["B5"].value)),3)
			concen_4_pt = [self.concen4, self.value4]
		except:
			concen4_enalble = 0
			pass
			
		try:
			self.value5 = float(sheet["C6"].value)
			self.concen5 = round(math.log10(float(sheet["B6"].value)),3)
			concen_5_pt = [self.concen5, self.value5]
		except:
			concen5_enalble = 0
			pass
			
		print("concen1_enalble: ", concen1_enalble)
		print("concen2_enalble: ", concen2_enalble)
		print("concen3_enalble: ", concen3_enalble)
		print("concen4_enalble: ", concen4_enalble)
		print("concen5_enalble: ", concen5_enalble)
		
		# ~ concen_1_pt = [1.3, 1.405]
		# ~ concen_2_pt = [2, 1.46]
		# ~ concen_3_pt = [3, 1.7]
		# ~ concen_4_pt = [4, 1.87]
		
		pts_list = []
		if(concen1_enalble != 0):
			pts_list.append(concen_1_pt)
		if(concen2_enalble != 0):
			pts_list.append(concen_2_pt)
		if(concen3_enalble != 0):
			pts_list.append(concen_3_pt)
		if(concen4_enalble != 0):
			pts_list.append(concen_4_pt)
		if(concen5_enalble != 0):
			pts_list.append(concen_5_pt)
		
		pts_list = np.array(pts_list)
		print("pts_list: ", pts_list)
		x = pts_list[:,0]
		y = pts_list[:,1]
		self.a_value, self.b_value = np.polyfit(x,y,1)
		self.a_value = round(self.a_value, 4)
		self.b_value = round(self.b_value, 4)
		print("a_value: ", self.a_value)
		print("b_value: ", self.b_value)

		self.experiment_name_text['state'] = "normal"
		self.experiment_name_text.delete('1.0',END)
		try:
			self.experiment_name_text.insert('1.0', self.program_button[button_index]['text'])
		except:
			pass
		self.experiment_name_text['state'] = "disabled"

		# ~ self.user_name_text['state'] = "normal"
		# ~ self.user_name_text.delete('1.0',END)
		# ~ try:
			# ~ self.user_name_text.insert('1.0', sheet["C2"].value)
		# ~ except:
			# ~ pass
		# ~ self.user_name_text['state'] = "disabled"

		# ~ self.pfi_text['state'] = "normal"
		# ~ self.pfi_text.delete('1.0',END)
		# ~ try:
			# ~ self.pfi_text.insert('1.0', sheet["I2"].value)
		# ~ except:
			# ~ pass
		# ~ self.pfi_text['state'] = "disabled"

		self.status_text['state'] = "normal"
		self.status_text.delete('1.0',END)
		try:
			if(concen1_enalble != 0):
				self.status_text.insert(END, str(sheet["B2"].value) + ': ')
				self.status_text.insert(END, str(sheet["C2"].value) + '\n')
			if(concen2_enalble != 0):
				self.status_text.insert(END, str(sheet["B3"].value) + ': ')
				self.status_text.insert(END, str(sheet["C3"].value) + '\n')
			if(concen3_enalble != 0):
				self.status_text.insert(END, str(sheet["B4"].value) + ': ')
				self.status_text.insert(END, str(sheet["C4"].value) + '\n')
			if(concen4_enalble != 0):
				self.status_text.insert(END, str(sheet["B5"].value) + ': ')
				self.status_text.insert(END, str(sheet["C5"].value) + '\n')
			if(concen5_enalble != 0):
				self.status_text.insert(END, str(sheet["B6"].value) + ': ')
				self.status_text.insert(END, str(sheet["C6"].value) + '\n')
			self.status_text.insert(END, "a_value: " + str(self.a_value) + '\n')
			self.status_text.insert(END, "b_value: " + str(self.b_value) + '\n')
			self.status_text.insert(END, "n_base_value: " + str(self.n_base_value) + '\n')

		except:
			pass
		self.status_text['state'] = "disabled"

		wb.close()

class QualitativeAnalysisFrame0(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.experiment_name = StringVar()
		self.user_name = StringVar()
		self.comments = StringVar()

		# 3 main frame
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = Screening0_Language["Title Label"][language],
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		setup_labelframe = LabelFrame(self.work_frame,
									font = LABEL_FRAME_TXT_FONT,
									bg = LABEL_BGD_COLOR,
									fg = LABEL_FRAME_TXT_COLOR)
		setup_labelframe.pack(expand=TRUE)

		self.experiment_name_label = Label(setup_labelframe,
									text = Screening0_Language["ExperimentName Label"][language],
									font = ('Helvetica', 10, 'bold'),
									bg = LABEL_BGD_COLOR,
									fg = LABEL_TXT_COLOR,
									justify = LEFT)
		self.experiment_name_label.grid(row=0, column=1, sticky=W, pady=20, padx=10)

		self.user_name_label = Label(setup_labelframe,
								text = Screening0_Language["TechnicianName Label"][language],
								font = ('Helvetica', 10, 'bold'),
								bg = LABEL_BGD_COLOR,
								fg = LABEL_TXT_COLOR,
								justify = LEFT)
		self.user_name_label.grid(row=1, column=1, sticky=W, pady=20, padx=10)

		# ~ self.template_name_label = Label(setup_labelframe,
								# ~ text = Screening0_Language["TemplateName Label"][language],
								# ~ font = ('Helvetica', 10, 'bold'),
								# ~ bg = LABEL_BGD_COLOR,
								# ~ fg = LABEL_TXT_COLOR,
								# ~ justify = LEFT)
		# ~ self.template_name_label.grid(row=2, column=1, sticky=W, pady=20, padx=10)

		self.experiment_name_entry = Entry(setup_labelframe, width=30, font=ENTRY_TXT_FONT)
		self.experiment_name_entry.grid(row=0, column=2, sticky=W, pady=20, padx=20)
		self.user_name_entry = Entry(setup_labelframe, width=30, font=ENTRY_TXT_FONT)
		self.user_name_entry.grid(row=1, column=2, sticky=W, pady=20, padx=20)
		# ~ self.template_name_entry = Entry(setup_labelframe, width=30, font=ENTRY_TXT_FONT)
		# ~ self.template_name_entry.grid(row=2, column=2, sticky=W, pady=20, padx=20)
		
		global autofill_email, autofill_user
		fr_info = open('/home/pi/Spotcheck/.oldinfo.txt')
		autofill_email = fr_info.readline().strip('\n')
		autofill_user = fr_info.readline().strip('\n')
		
		self.user_name_entry.insert(0, autofill_user)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = Screening0_Language["Back Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

		self.next_button = Button(self.button_frame,
								text = Screening0_Language["Next Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.next_clicked)
		self.next_button.pack(ipadx=30, ipady=10, side=RIGHT)

	def back_clicked(self):
		self.base_window.forget_page()
		#self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_option)
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
		
		self.base_window.switch_page()

	def next_clicked(self):
		self.experiment_name = self.experiment_name_entry.get()
		self.user_name = self.user_name_entry.get()
		self.template_name = ''
		# ~ if(self.experiment_name==''):
			# ~ messagebox.showwarning("","Plese enter Folder Name !")
		# ~ elif (self.user_name==''):
			# ~ messagebox.showwarning("","Plese enter User Name !")
		# ~ elif(self.template_name == ''):
			# ~ messagebox.showwarning("","Plese enter Template Name !")
		# ~ else:
		
		global autofill_email, autofill_user
		autofill_user = self.user_name
		fw_info = open('/home/pi/Spotcheck/.oldinfo.txt', 'w')
		fw_info.writelines(autofill_email + '\n')
		fw_info.writelines(self.user_name + '\n')
		fw_info.close()
		
		self.create_time = strftime("%y-%m-%d")
		self.result_folder_name_0 = self.create_time
		self.create_time_1 = strftime("%Hh%Mm%Ss")
		
		if not os.path.exists(results_qualitative_path + self.result_folder_name_0):
			self.result_folder_path_0  = os.path.join(results_qualitative_path , self.result_folder_name_0)
			os.mkdir(self.result_folder_path_0)
		else:
			self.result_folder_path_0 = results_qualitative_path +  self.result_folder_name_0
			 
		if(self.experiment_name != ''):
			self.result_folder_name = self.experiment_name
		else:
			self.result_folder_name = self.create_time_1
			self.experiment_name = self.create_time_1
		
		if not os.path.exists(self.result_folder_path_0 + '/' + self.result_folder_name + '/'):
			self.result_folder_path = os.path.join(self.result_folder_path_0, self.result_folder_name)
			os.mkdir(self.result_folder_path)
			print(self.result_folder_path)
			
			self.system_check_folder = os.path.join(self.result_folder_path, "System_Check")
			os.mkdir(self.system_check_folder)

			self.analysis_result_folder = os.path.join(self.result_folder_path, "Analysis Results")
			os.mkdir(self.analysis_result_folder)
			
			# ~ self.base_window.system_check.mode_check = 1
			
			self.base_window.forget_page()
			#self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_1)
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_2)
			self.base_window.switch_page()
			
		else:
			msg = messagebox.askquestion("",Screening0_Language["Folder Exists"][language])
			if(msg == 'yes'):
				print("self.result_folder_path_0: ", self.result_folder_path_0)
				self.result_folder_path = self.result_folder_path_0 + '/' + self.result_folder_name
				shutil.rmtree(self.result_folder_path_0 + '/' + self.result_folder_name)
				os.mkdir(self.result_folder_path)
				
				self.system_check_folder = os.path.join(self.result_folder_path, "System_Check")
				os.mkdir(self.system_check_folder)

				self.analysis_result_folder = os.path.join(self.result_folder_path, "Analysis Results")
				os.mkdir(self.analysis_result_folder)
				
				# ~ self.base_window.system_check.mode_check = 1
				
				self.base_window.forget_page()
				#self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_1)
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_2)
				self.base_window.switch_page()
				
		
class QuantitativeAnalysisFrame0(QualitativeAnalysisFrame0):
	def __init__(self, container):
		super().__init__(container)
		self.title_label['text'] = Quantitative0_Language['Title Label'][language]
		self.experiment_name_label['text'] = Quantitative0_Language['ExperimentName Label'][language]
		self.user_name_label['text'] = Quantitative0_Language['TechnicianName Label'][language]
		# ~ self.template_name_label['text'] = Quantitative0_Language['TemplateName Label'][language]
		self.back_button['text'] = Quantitative0_Language['Back Button'][language]
		self.next_button['text'] = Quantitative0_Language['Next Button'][language]

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
		self.base_window.switch_page()

	def next_clicked(self):
		self.experiment_name = self.experiment_name_entry.get()
		self.user_name = self.user_name_entry.get()
		self.template_name = ''
		# ~ if(self.experiment_name==''):
			# ~ messagebox.showwarning("","Plese enter Folder Name !")
		# ~ elif (self.user_name==''):
			# ~ messagebox.showwarning("","Plese enter User Name !")
		# ~ elif(self.template_name == ''):
			# ~ messagebox.showwarning("","Plese enter Template Name !")
		# ~ else:
		
		global autofill_email, autofill_user
		autofill_user = self.user_name
		fw_info = open('/home/pi/Spotcheck/.oldinfo.txt', 'w')
		fw_info.writelines(autofill_email + '\n')
		fw_info.writelines(self.user_name + '\n')
		fw_info.close()
		
		self.create_time = strftime("%y-%m-%d")
		self.result_folder_name_0 = self.create_time
		self.create_time_1 = strftime("%Hh%Mm%Ss")
		
		if not os.path.exists(results_quantitative_path + self.result_folder_name_0):
			self.result_folder_path_0  = os.path.join(results_quantitative_path , self.result_folder_name_0)
			os.mkdir(self.result_folder_path_0)
		else:
			self.result_folder_path_0 = results_quantitative_path +  self.result_folder_name_0
			 
		if(self.experiment_name != ''):
			self.result_folder_name = self.experiment_name
		else:
			self.result_folder_name = self.create_time_1
			self.experiment_name = self.create_time_1
		
		if not os.path.exists(self.result_folder_path_0 + '/' + self.result_folder_name + '/'):
			self.result_folder_path = os.path.join(self.result_folder_path_0, self.result_folder_name)
			os.mkdir(self.result_folder_path)
			print(self.result_folder_path)
			
			self.system_check_folder = os.path.join(self.result_folder_path, "System_Check")
			os.mkdir(self.system_check_folder)

			self.analysis_result_folder = os.path.join(self.result_folder_path, "Analysis Results")
			os.mkdir(self.analysis_result_folder)
			
			# ~ self.base_window.system_check.mode_check = 1
			
			self.base_window.forget_page()
			#self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_1)
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_1)
			self.base_window.switch_page()
			self.base_window.quantitative_analysis_1.load_program()
			
		else:
			msg = messagebox.askquestion("",Quantitative0_Language["Folder Exists"][language])
			if(msg == 'yes'):
				print("self.result_folder_path_0: ", self.result_folder_path_0)
				self.result_folder_path = self.result_folder_path_0 + '/' + self.result_folder_name
				shutil.rmtree(self.result_folder_path_0 + '/' + self.result_folder_name)
				os.mkdir(self.result_folder_path)
				
				self.system_check_folder = os.path.join(self.result_folder_path, "System_Check")
				os.mkdir(self.system_check_folder)

				self.analysis_result_folder = os.path.join(self.result_folder_path, "Analysis Results")
				os.mkdir(self.analysis_result_folder)
				
				# ~ self.base_window.system_check.mode_check = 1
				
				self.base_window.forget_page()
				#self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_1)
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_1)
				self.base_window.switch_page()
				self.base_window.quantitative_analysis_1.load_program()


class IDCreateFrame(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.direct_create = 0
		self.first_well_index = 0
		self.last_well_index = 47

		# 3 main frame
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)
		
		# In title frame
		self.title_label = Label(self.title_frame,
								text = CreateFile_Language["Title Label"][language],
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		# Sample button frame
		self.well_button_table_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR)
		self.well_button_table_frame.pack(side=LEFT)

		self.well_button = list(range(SC_VERSION))
		self.well_frame = list(range(SC_VERSION))

		index = 0
		for c in range(0, WELL_COLUMN):
			for r in range(0, WELL_ROW):
				self.well_frame[index] = Frame(self.well_button_table_frame, highlightbackground = SAMPLE_BUTTON_BGD_COLOR,  highlightthickness = 2, bd=0)
				self.well_button[index] = Button(self.well_frame[index],
											bg = SAMPLE_BUTTON_BGD_COLOR,
											fg = SAMPLE_BUTTON_TXT_COLOR,
											activebackground = SAMPLE_BUTTON_ACTIVE_BGD_COLOR,
											justify = 'left',
											borderwidth = 0,
											text = '#',
											width = SAMPLE_BUTTON_WIDTH,
											height = SAMPLE_BUTTON_HEIGHT)
				self.well_button[index]['command'] = partial(self.well_button_clicked, index)
				self.well_frame[index].grid(row=r, column=c, padx=1, pady=1)
				self.well_button[index].pack()

				index += 1

		self.well_frame[0]['highlightbackground'] = QS_FIRSTWELL_COLOR
		self.well_frame[SC_VERSION-1]['highlightbackground'] = QS_LASTWELL_COLOR
		
		# Properties frame
		self.property_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR, width=495)
		self.property_frame.pack(fill=BOTH, expand=TRUE, side=LEFT)

		self.property_labelframe = LabelFrame(self.property_frame,
										text = CreateFile_Language["SampleProperties LabelFrame"][language],
										font  = LABEL_FRAME_TXT_FONT,
										bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
										fg = LABEL_FRAME_TXT_COLOR)
		self.property_labelframe.pack(fill=BOTH, expand=TRUE, padx=10, pady=10)

		self.property_labelframe.rowconfigure(0, weight=1)
		self.property_labelframe.rowconfigure(1, weight=1)
		self.property_labelframe.rowconfigure(2, weight=1)
		self.property_labelframe.rowconfigure(3, weight=4)

		self.well_name_label = Label(self.property_labelframe,
						bg = SAMPLE_BUTTON_CHOOSE_BGD_COLOR,
						fg = LABEL_TXT_COLOR,
						font = SAMPLE_LABEL_TXT_FONT)
		self.well_name_label.grid(row=0, column=0, columnspan=2, sticky=EW)
		
		#quick create frame
		self.quick_create_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR)
		self.quick_create_frame.pack(fill=BOTH, expand=TRUE, side=LEFT)
		
		self.quick_create_labelframe = LabelFrame(self.quick_create_frame,
										text = CreateFile_Language["QuickSetup LabelFrame"][language],
										font  = LABEL_FRAME_TXT_FONT,
										bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
										fg = LABEL_FRAME_TXT_COLOR)
		self.quick_create_labelframe.pack(fill=BOTH, expand=TRUE, padx=10, pady=10)
		
		self.qc1_frame = Frame(self.quick_create_labelframe,bg = SAMPLE_BUTTON_FRAME_BDG_COLOR)
		self.qc1_frame.pack(fill=BOTH, expand=TRUE)
		self.qc2_frame = Frame(self.quick_create_labelframe, bg = SAMPLE_BUTTON_FRAME_BDG_COLOR)
		self.qc2_frame.pack(fill=BOTH, expand=TRUE, side=BOTTOM) 
		
		self.first_well_button = Button(self.qc1_frame,
								text = CreateFile_Language["FirstWell Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = "lawn green",
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.first_well_button_clicked)
		self.first_well_button.grid(row=0, column=0, ipadx=10, ipady=10, pady=20)
		
		self.last_well_button = Button(self.qc1_frame,
								text = CreateFile_Language["LastWell Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.last_well_button_clicked)
		self.last_well_button.grid(row=0, column=1, ipadx=10, ipady=10, pady=20)
		
		self.first_well_label = Label(self.qc1_frame,
						text = "A1",
						bg = QS_FIRSTWELL_COLOR,
						fg = LABEL_TXT_COLOR,
						font = SAMPLE_LABEL_TXT_FONT)
		self.first_well_label.grid(row=1, column=0)
		
		self.last_well_label = Label(self.qc1_frame,
						text = "H6",
						bg = QS_LASTWELL_COLOR,
						fg = LABEL_TXT_COLOR,
						font = SAMPLE_LABEL_TXT_FONT)
		self.last_well_label.grid(row=1, column=1)
	   
		self.quick_create_button = Button(self.qc2_frame,
								text = CreateFile_Language["Set Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.quick_create_button_clicked)
		self.quick_create_button.pack(ipadx=50, ipady=10)
		
		# In button frame
		self.back_button = Button(self.button_frame,
								text = CreateFile_Language["Back Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

		self.create_button = Button(self.button_frame,
								text = CreateFile_Language["Create Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.create_clicked)
		self.create_button.pack(ipadx=30, ipady=10, side=RIGHT)

		self.load_button = Button(self.button_frame,
								text = CreateFile_Language["Load Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.load_clicked)
		self.load_button.pack(ipadx=30, ipady=10, anchor=CENTER)

	def well_button_clicked(self,n):
		if(self.well_button[n]['bg'] == SAMPLE_BUTTON_DONE_BGD_COLOR):
			for k in range (0,SC_VERSION):
				if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
					self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
				if(self.well_button[k]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
					self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
			self.well_button[n]['bg'] = SAMPLE_BUTTON_TMP_BGD_COLOR
		elif(self.well_button[n]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
			pass
		else:
			for k in range (0,SC_VERSION):
				if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
					self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
				else:
					self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
			self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR


		def ok_clicked(event=None):
			if(self.sample_name_entry.get()==''):
				self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
				self.well_button[n]['text'] = '#'
				messagebox.showwarning("",CreateFile_Language["SampleName Empty"][language])
			else:
				self.well_button[n]['text'] = self.sample_name_entry.get()
				self.well_button[n]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
				
				if(n != (SC_VERSION - 1)):
					self.well_button_clicked(n+1)
				else:
					self.well_button_clicked(0)
				
				# try:
				#     if(n==42):
				#         self.well_button_clicked(1)
				#     elif(n==43):
				#         self.well_button_clicked(2)
				#     elif(n==44):
				#         self.well_button_clicked(3)
				#     elif(n==45):
				#         self.well_button_clicked(4)
				#     elif(n==46):
				#         self.well_button_clicked(5)
				#     elif(n==47):
				#         self.well_button_clicked(0)
				#     else:
				#         self.well_button_clicked(n+6)
				# except:
				#     self.well_button_clicked(0)


		self.sample_name_label = Label(self.property_labelframe,
									text = CreateFile_Language["SampleName Label"][language],
									font = LABEL_TXT_FONT,
									bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
									fg = LABEL_TXT_COLOR)
		self.sample_name_label.grid(row=1, column=0, padx=78, pady=2, sticky=SE)

		self.sample_name_entry = Entry(self.property_labelframe, width=20, font=ENTRY_TXT_FONT)
		if(self.well_button[n]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
			self.sample_name_entry.insert(0, self.well_button[n]['text'])
		#id_entry.bind("<Button-1>", enter_entry)
		self.sample_name_entry.bind("<Return>", ok_clicked)
		self.sample_name_entry.grid(row=2, column=0, padx=30, pady=0)
		self.sample_name_entry.focus_set()

		# ~ if(n<6):
			# ~ self.well_name_label['text'] = "A" + str(n+1)
		# ~ elif(n<12):
			# ~ self.well_name_label['text'] = "B" + str(n+1-6)
		# ~ elif(n<18):
			# ~ self.well_name_label['text'] = "C" + str(n+1-12)
		# ~ elif(n<24):
			# ~ self.well_name_label['text'] = "D" + str(n+1-18)
		# ~ elif(n<30):
			# ~ self.well_name_label['text'] = "E" + str(n+1-24)
		# ~ elif(n<36):
			# ~ self.well_name_label['text'] = "F" + str(n+1-30)
		# ~ elif(n<42):
			# ~ self.well_name_label['text'] = "G" + str(n+1-36)
		# ~ else:
			# ~ self.well_name_label['text'] = "H" + str(n+1-42)

		if(SC_VERSION == 48):
			if(n<8):
				self.well_name_label['text'] = str(chr(65+n)) + '1'
			if(n>=8 and n<16):
				self.well_name_label['text'] = str(chr(65+n-8)) + '2'
			if(n>=16 and n<24):
				self.well_name_label['text'] = str(chr(65+n-16)) + '3'
			if(n>=24 and n<32):
				self.well_name_label['text'] = str(chr(65+n-24)) + '4'
			if(n>=32 and n<40):
				self.well_name_label['text'] = str(chr(65+n-32)) + '5'
			if(n>=40):
				self.well_name_label['text'] = str(chr(65+n-40)) + '6'
		else:
			if(n<4):
				self.well_name_label['text'] = str(chr(65+n)) + '1'
			elif(n<8):
				self.well_name_label['text'] = str(chr(65+n-4)) + '2'
			elif(n<12):
				self.well_name_label['text'] = str(chr(65+n-8)) + '3'
			else:
				self.well_name_label['text'] = str(chr(65+n-12)) + '4'
		
		
		if(self.first_well_button['bg'] == 'lawn green' and self.well_frame[n]['highlightbackground'] != QS_LASTWELL_COLOR):
			self.first_well_label['text'] = self.well_name_label['text']
			self.first_well_index = n
			for i in range(0, SC_VERSION):
				if(self.well_frame[i]['highlightbackground'] != QS_LASTWELL_COLOR):
					self.well_frame[i]['highlightbackground'] = SAMPLE_BUTTON_BGD_COLOR
			self.well_frame[n]['highlightbackground'] = QS_FIRSTWELL_COLOR
		elif(self.first_well_button['bg'] != 'lawn green' and self.well_frame[n]['highlightbackground'] != QS_FIRSTWELL_COLOR):
			self.last_well_label['text'] = self.well_name_label['text']
			self.last_well_index = n
			for i in range(0, SC_VERSION):
				if(self.well_frame[i]['highlightbackground'] != QS_FIRSTWELL_COLOR):
					self.well_frame[i]['highlightbackground'] = SAMPLE_BUTTON_BGD_COLOR
			self.well_frame[n]['highlightbackground'] = QS_LASTWELL_COLOR
			
		self.ok_button = Button(self.property_labelframe,
								text = CreateFile_Language["OK Button"][language],
								bg = CONFIRM_BUTTON_BGD_COLOR,
								fg = CONFIRM_BUTTON_TXT_COLOR,
								font = CONFIRM_BUTTON_TXT_FONT,
								borderwidth = 0,
								command = ok_clicked)
		self.ok_button.grid(row=3, column=0, columnspan=2, ipadx=30, ipady=10)

	def load_clicked(self):
		file = filedialog.askopenfilename(initialdir=id_path, filetypes=[('Excel file','*.xlsm *.xlsx *.xls')])
		if file is not None:
			wb = load_workbook(file)
			sheet = wb.active
			
			j=0
			for i in range(0,SC_VERSION):
				if(SC_VERSION == 48):
					if(i==8):
						j=1
					elif(i==16):
						j=2
					elif(i==24):
						j=3
					elif(i==32):
						j=4
					elif(i==40):
						j=5
					# ~ elif(i==47):
						# ~ j=0
				else:
					if(i==4):
						j=1
					elif(i==8):
						j=2
					elif(i==12):
						j=3
					
				pos = 'B' + str(i + 12)
				self.well_button[i]['text'] = sheet[pos].value
				if(self.well_button[i]['text'] == "N/A"):
					self.well_button[i]['bg'] = SAMPLE_BUTTON_BGD_COLOR
				else:
					# ~ if(i!=47):
					self.well_button[i]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
				
				j=j+6
				
	def create_clicked(self):
		file_create_done = 0
		msg = messagebox.askquestion("",CreateFile_Language["CreateFile Confirm"][language])
		if(msg=="yes"):
			self.well_button_change_pos = list(range(SC_VERSION))
			wb = Workbook()
			sheet = wb.active
			
			# ~ self.well_button_change_pos[0] = self.well_button[0]['text']
			# ~ self.well_button_change_pos[1] = self.well_button[6]['text']
			# ~ self.well_button_change_pos[2] = self.well_button[12]['text']
			# ~ self.well_button_change_pos[3] = self.well_button[18]['text']
			# ~ self.well_button_change_pos[4] = self.well_button[24]['text']
			# ~ self.well_button_change_pos[5] = self.well_button[30]['text']
			# ~ self.well_button_change_pos[6] = self.well_button[36]['text']
			# ~ self.well_button_change_pos[7] = self.well_button[42]['text']
			
			# ~ self.well_button_change_pos[8] = self.well_button[1]['text']
			# ~ self.well_button_change_pos[9] = self.well_button[7]['text']
			# ~ self.well_button_change_pos[10] = self.well_button[13]['text']
			# ~ self.well_button_change_pos[11] = self.well_button[19]['text']
			# ~ self.well_button_change_pos[12] = self.well_button[25]['text']
			# ~ self.well_button_change_pos[13] = self.well_button[31]['text']
			# ~ self.well_button_change_pos[14] = self.well_button[37]['text']
			# ~ self.well_button_change_pos[15] = self.well_button[43]['text']
			
			# ~ self.well_button_change_pos[16] = self.well_button[2]['text']
			# ~ self.well_button_change_pos[17] = self.well_button[8]['text']
			# ~ self.well_button_change_pos[18] = self.well_button[14]['text']
			# ~ self.well_button_change_pos[19] = self.well_button[20]['text']
			# ~ self.well_button_change_pos[20] = self.well_button[26]['text']
			# ~ self.well_button_change_pos[21] = self.well_button[32]['text']
			# ~ self.well_button_change_pos[22] = self.well_button[38]['text']
			# ~ self.well_button_change_pos[23] = self.well_button[44]['text']
			
			# ~ self.well_button_change_pos[24] = self.well_button[3]['text']
			# ~ self.well_button_change_pos[25] = self.well_button[9]['text']
			# ~ self.well_button_change_pos[26] = self.well_button[15]['text']
			# ~ self.well_button_change_pos[27] = self.well_button[21]['text']
			# ~ self.well_button_change_pos[28] = self.well_button[27]['text']
			# ~ self.well_button_change_pos[29] = self.well_button[33]['text']
			# ~ self.well_button_change_pos[30] = self.well_button[39]['text']
			# ~ self.well_button_change_pos[31] = self.well_button[45]['text']
			
			# ~ self.well_button_change_pos[32] = self.well_button[4]['text']
			# ~ self.well_button_change_pos[33] = self.well_button[10]['text']
			# ~ self.well_button_change_pos[34] = self.well_button[16]['text']
			# ~ self.well_button_change_pos[35] = self.well_button[22]['text']
			# ~ self.well_button_change_pos[36] = self.well_button[28]['text']
			# ~ self.well_button_change_pos[37] = self.well_button[34]['text']
			# ~ self.well_button_change_pos[38] = self.well_button[40]['text']
			# ~ self.well_button_change_pos[39] = self.well_button[46]['text']
			
			# ~ self.well_button_change_pos[40] = self.well_button[5]['text']
			# ~ self.well_button_change_pos[41] = self.well_button[11]['text']
			# ~ self.well_button_change_pos[42] = self.well_button[17]['text']
			# ~ self.well_button_change_pos[43] = self.well_button[23]['text']
			# ~ self.well_button_change_pos[44] = self.well_button[29]['text']
			# ~ self.well_button_change_pos[45] = self.well_button[35]['text']
			# ~ self.well_button_change_pos[46] = self.well_button[41]['text']
			# ~ self.well_button_change_pos[47] = self.well_button[47]['text']
			
			for i in range(0,SC_VERSION):
				pos = "B" + str(i + 12)
				if(self.well_button[i]['text'] != "#"):
					sheet[pos] = self.well_button[i]['text']
				else:
					sheet[pos] = "N/A"
					
			
			if(self.direct_create == 1):
				wb.save(id_path + '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsx')
				msg = messagebox.askquestion("",CreateFile_Language["CreateFile Done"][language])
				if(msg=="yes"):
					file_name = self.base_window.qualitative_analysis_0.experiment_name
					file_create_done = 1
					self.base_window.qualitative_analysis_2.id_file_path = id_path + '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsx'
			elif(self.direct_create == 2):
				wb.save(id_path + '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsx')
				msg = messagebox.askquestion("",CreateFile_Language["CreateFile Done"][language])
				if(msg=="yes"):
					file_name = self.base_window.quantitative_analysis_0.experiment_name
					file_create_done = 1
			else:
				path = filedialog.asksaveasfilename(initialdir = id_path + '/', defaultextension='.xlsx')
				if path is not None:
					tmp = 0
					for i in range(len(path)):
						if(path[i]=='/'):
							tmp = i+1
					file_name = path[tmp:(len(path)-5)]
					if(len(file_name)<30):
						wb.save(path)
						msg = messagebox.askquestion("",CreateFile_Language["CreateFile Done"][language])
						if(msg=="yes"):
							file_create_done = 1
					else:
						messagebox.showerror("",CreateFile_Language["FileName OverLength"][language])
			
			if(file_create_done):
				if(self.direct_create == 0): #create file with create module from main menu
					self.back_clicked()
				elif(self.direct_create==1): #create file with create module from qualitative_analysis_2
					wb = load_workbook(id_path + '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsx')
					sheet = wb.active			
					if(SC_VERSION == 48):		
						self.base_window.qualitative_analysis_2.id_list[0] = sheet["B12"].value
						self.base_window.qualitative_analysis_2.id_list[1] = sheet["B20"].value
						self.base_window.qualitative_analysis_2.id_list[2] = sheet["B28"].value
						self.base_window.qualitative_analysis_2.id_list[3] = sheet["B36"].value
						self.base_window.qualitative_analysis_2.id_list[4] = sheet["B44"].value
						self.base_window.qualitative_analysis_2.id_list[5] = sheet["B52"].value
						
						self.base_window.qualitative_analysis_2.id_list[6] = sheet["B13"].value
						self.base_window.qualitative_analysis_2.id_list[7] = sheet["B21"].value
						self.base_window.qualitative_analysis_2.id_list[8] = sheet["B29"].value
						self.base_window.qualitative_analysis_2.id_list[9] = sheet["B37"].value
						self.base_window.qualitative_analysis_2.id_list[10] = sheet["B45"].value
						self.base_window.qualitative_analysis_2.id_list[11] = sheet["B53"].value
						
						self.base_window.qualitative_analysis_2.id_list[12] = sheet["B14"].value
						self.base_window.qualitative_analysis_2.id_list[13] = sheet["B22"].value
						self.base_window.qualitative_analysis_2.id_list[14] = sheet["B30"].value
						self.base_window.qualitative_analysis_2.id_list[15] = sheet["B38"].value
						self.base_window.qualitative_analysis_2.id_list[16] = sheet["B46"].value
						self.base_window.qualitative_analysis_2.id_list[17] = sheet["B54"].value
						
						self.base_window.qualitative_analysis_2.id_list[18] = sheet["B15"].value
						self.base_window.qualitative_analysis_2.id_list[19] = sheet["B23"].value
						self.base_window.qualitative_analysis_2.id_list[20] = sheet["B31"].value
						self.base_window.qualitative_analysis_2.id_list[21] = sheet["B39"].value
						self.base_window.qualitative_analysis_2.id_list[22] = sheet["B47"].value
						self.base_window.qualitative_analysis_2.id_list[23] = sheet["B55"].value
						
						self.base_window.qualitative_analysis_2.id_list[24] = sheet["B16"].value
						self.base_window.qualitative_analysis_2.id_list[25] = sheet["B24"].value
						self.base_window.qualitative_analysis_2.id_list[26] = sheet["B32"].value
						self.base_window.qualitative_analysis_2.id_list[27] = sheet["B40"].value
						self.base_window.qualitative_analysis_2.id_list[28] = sheet["B48"].value
						self.base_window.qualitative_analysis_2.id_list[29] = sheet["B56"].value
						
						self.base_window.qualitative_analysis_2.id_list[30] = sheet["B17"].value
						self.base_window.qualitative_analysis_2.id_list[31] = sheet["B25"].value
						self.base_window.qualitative_analysis_2.id_list[32] = sheet["B33"].value
						self.base_window.qualitative_analysis_2.id_list[33] = sheet["B41"].value
						self.base_window.qualitative_analysis_2.id_list[34] = sheet["B49"].value
						self.base_window.qualitative_analysis_2.id_list[35] = sheet["B57"].value
						
						self.base_window.qualitative_analysis_2.id_list[36] = sheet["B18"].value
						self.base_window.qualitative_analysis_2.id_list[37] = sheet["B26"].value
						self.base_window.qualitative_analysis_2.id_list[38] = sheet["B34"].value
						self.base_window.qualitative_analysis_2.id_list[39] = sheet["B42"].value
						self.base_window.qualitative_analysis_2.id_list[40] = sheet["B50"].value
						self.base_window.qualitative_analysis_2.id_list[41] = sheet["B58"].value
						
						self.base_window.qualitative_analysis_2.id_list[42] = sheet["B19"].value
						self.base_window.qualitative_analysis_2.id_list[43] = sheet["B27"].value
						self.base_window.qualitative_analysis_2.id_list[44] = sheet["B35"].value
						self.base_window.qualitative_analysis_2.id_list[45] = sheet["B43"].value
						self.base_window.qualitative_analysis_2.id_list[46] = sheet["B51"].value
						self.base_window.qualitative_analysis_2.id_list[47] = sheet["B59"].value
					else:
						self.base_window.qualitative_analysis_2.id_list[0] = sheet["B12"].value
						self.base_window.qualitative_analysis_2.id_list[1] = sheet["B16"].value
						self.base_window.qualitative_analysis_2.id_list[2] = sheet["B20"].value
						self.base_window.qualitative_analysis_2.id_list[3] = sheet["B24"].value
						
						self.base_window.qualitative_analysis_2.id_list[4] = sheet["B13"].value
						self.base_window.qualitative_analysis_2.id_list[5] = sheet["B17"].value
						self.base_window.qualitative_analysis_2.id_list[6] = sheet["B21"].value
						self.base_window.qualitative_analysis_2.id_list[7] = sheet["B25"].value

						self.base_window.qualitative_analysis_2.id_list[8] = sheet["B14"].value
						self.base_window.qualitative_analysis_2.id_list[9] = sheet["B18"].value
						self.base_window.qualitative_analysis_2.id_list[10] = sheet["B22"].value
						self.base_window.qualitative_analysis_2.id_list[11] = sheet["B26"].value

						self.base_window.qualitative_analysis_2.id_list[12] = sheet["B15"].value
						self.base_window.qualitative_analysis_2.id_list[13] = sheet["B19"].value
						self.base_window.qualitative_analysis_2.id_list[14] = sheet["B23"].value
						self.base_window.qualitative_analysis_2.id_list[15] = sheet["B27"].value

					self.base_window.qualitative_analysis_2.id_file_name_label['text'] = file_name
					self.base_window.qualitative_analysis_2.id_file_name_label['bg'] = 'lawn green'

					try:
						for i in range(0, SC_VERSION):
							self.base_window.qualitative_analysis_2.id_label[i].destroy()
					except:
						pass

					# ~ Pmw.initialise(self.base_window)
					# ~ self.base_window.qualitative_analysis_2.tooltip = list(range(48))

					self.base_window.qualitative_analysis_2.id_label = list(range(SC_VERSION))
					index = 0 
					for r in range(0, WELL_ROW):
						for c in range(0, WELL_COLUMN):
							self.base_window.qualitative_analysis_2.id_label[index] = Label(self.base_window.qualitative_analysis_2.id_pos_frame,
													width=6,
													height=3,
													text = self.base_window.qualitative_analysis_2.id_list[index],
													# ~ bg = RESULT_LABEL_BGD_COLOR,
													font = RESULT_LABEL_TXT_FONT)

							# ~ self.base_window.qualitative_analysis_2.tooltip[i] = Pmw.Balloon(self.base_window)
							# ~ self.base_window.qualitative_analysis_2.tooltip[i].bind(self.base_window.qualitative_analysis_2.id_label[i], self.base_window.qualitative_analysis_2.id_list[i])

							if(self.base_window.qualitative_analysis_2.id_list[index] != '#' and self.base_window.qualitative_analysis_2.id_list[index] != 'N/A'):
								self.base_window.qualitative_analysis_2.id_label[index]['bg'] = "lawn green"
								self.base_window.qualitative_analysis_2.id_label[index]['text'] = self.base_window.qualitative_analysis_2.id_list[index]
							else:
								self.base_window.qualitative_analysis_2.id_label[index]['bg'] = "grey80"
								self.base_window.qualitative_analysis_2.id_label[index]['text'] = "N/A"

							self.base_window.qualitative_analysis_2.id_label[index].grid(row=r, column=c, padx=1, pady=1)

							index += 1

					self.back_clicked()
					msg = messagebox.askokcancel("",CreateFile_Language["AllowPutSample Inform"][language])

				elif(self.direct_create==2):
					wb = load_workbook(id_path + '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsx')
					sheet = wb.active					
					if(SC_VERSION == 48):		
						self.base_window.quantitative_analysis_2.id_list[0] = sheet["B12"].value
						self.base_window.quantitative_analysis_2.id_list[1] = sheet["B20"].value
						self.base_window.quantitative_analysis_2.id_list[2] = sheet["B28"].value
						self.base_window.quantitative_analysis_2.id_list[3] = sheet["B36"].value
						self.base_window.quantitative_analysis_2.id_list[4] = sheet["B44"].value
						self.base_window.quantitative_analysis_2.id_list[5] = sheet["B52"].value
						
						self.base_window.quantitative_analysis_2.id_list[6] = sheet["B13"].value
						self.base_window.quantitative_analysis_2.id_list[7] = sheet["B21"].value
						self.base_window.quantitative_analysis_2.id_list[8] = sheet["B29"].value
						self.base_window.quantitative_analysis_2.id_list[9] = sheet["B37"].value
						self.base_window.quantitative_analysis_2.id_list[10] = sheet["B45"].value
						self.base_window.quantitative_analysis_2.id_list[11] = sheet["B53"].value
						
						self.base_window.quantitative_analysis_2.id_list[12] = sheet["B14"].value
						self.base_window.quantitative_analysis_2.id_list[13] = sheet["B22"].value
						self.base_window.quantitative_analysis_2.id_list[14] = sheet["B30"].value
						self.base_window.quantitative_analysis_2.id_list[15] = sheet["B38"].value
						self.base_window.quantitative_analysis_2.id_list[16] = sheet["B46"].value
						self.base_window.quantitative_analysis_2.id_list[17] = sheet["B54"].value
						
						self.base_window.quantitative_analysis_2.id_list[18] = sheet["B15"].value
						self.base_window.quantitative_analysis_2.id_list[19] = sheet["B23"].value
						self.base_window.quantitative_analysis_2.id_list[20] = sheet["B31"].value
						self.base_window.quantitative_analysis_2.id_list[21] = sheet["B39"].value
						self.base_window.quantitative_analysis_2.id_list[22] = sheet["B47"].value
						self.base_window.quantitative_analysis_2.id_list[23] = sheet["B55"].value
						
						self.base_window.quantitative_analysis_2.id_list[24] = sheet["B16"].value
						self.base_window.quantitative_analysis_2.id_list[25] = sheet["B24"].value
						self.base_window.quantitative_analysis_2.id_list[26] = sheet["B32"].value
						self.base_window.quantitative_analysis_2.id_list[27] = sheet["B40"].value
						self.base_window.quantitative_analysis_2.id_list[28] = sheet["B48"].value
						self.base_window.quantitative_analysis_2.id_list[29] = sheet["B56"].value
						
						self.base_window.quantitative_analysis_2.id_list[30] = sheet["B17"].value
						self.base_window.quantitative_analysis_2.id_list[31] = sheet["B25"].value
						self.base_window.quantitative_analysis_2.id_list[32] = sheet["B33"].value
						self.base_window.quantitative_analysis_2.id_list[33] = sheet["B41"].value
						self.base_window.quantitative_analysis_2.id_list[34] = sheet["B49"].value
						self.base_window.quantitative_analysis_2.id_list[35] = sheet["B57"].value
						
						self.base_window.quantitative_analysis_2.id_list[36] = sheet["B18"].value
						self.base_window.quantitative_analysis_2.id_list[37] = sheet["B26"].value
						self.base_window.quantitative_analysis_2.id_list[38] = sheet["B34"].value
						self.base_window.quantitative_analysis_2.id_list[39] = sheet["B42"].value
						self.base_window.quantitative_analysis_2.id_list[40] = sheet["B50"].value
						self.base_window.quantitative_analysis_2.id_list[41] = sheet["B58"].value
						
						self.base_window.quantitative_analysis_2.id_list[42] = sheet["B19"].value
						self.base_window.quantitative_analysis_2.id_list[43] = sheet["B27"].value
						self.base_window.quantitative_analysis_2.id_list[44] = sheet["B35"].value
						self.base_window.quantitative_analysis_2.id_list[45] = sheet["B43"].value
						self.base_window.quantitative_analysis_2.id_list[46] = sheet["B51"].value
						self.base_window.quantitative_analysis_2.id_list[47] = sheet["B59"].value
					else:
						self.base_window.quantitative_analysis_2.id_list[0] = sheet["B12"].value
						self.base_window.quantitative_analysis_2.id_list[1] = sheet["B16"].value
						self.base_window.quantitative_analysis_2.id_list[2] = sheet["B20"].value
						self.base_window.quantitative_analysis_2.id_list[3] = sheet["B24"].value
						
						self.base_window.quantitative_analysis_2.id_list[4] = sheet["B13"].value
						self.base_window.quantitative_analysis_2.id_list[5] = sheet["B17"].value
						self.base_window.quantitative_analysis_2.id_list[6] = sheet["B21"].value
						self.base_window.quantitative_analysis_2.id_list[7] = sheet["B25"].value

						self.base_window.quantitative_analysis_2.id_list[8] = sheet["B14"].value
						self.base_window.quantitative_analysis_2.id_list[9] = sheet["B18"].value
						self.base_window.quantitative_analysis_2.id_list[10] = sheet["B22"].value
						self.base_window.quantitative_analysis_2.id_list[11] = sheet["B26"].value

						self.base_window.quantitative_analysis_2.id_list[12] = sheet["B15"].value
						self.base_window.quantitative_analysis_2.id_list[13] = sheet["B19"].value
						self.base_window.quantitative_analysis_2.id_list[14] = sheet["B23"].value
						self.base_window.quantitative_analysis_2.id_list[15] = sheet["B27"].value

					self.base_window.quantitative_analysis_2.id_file_name_label['text'] = file_name
					self.base_window.quantitative_analysis_2.id_file_name_label['bg'] = 'lawn green'

					try:
						for i in range(0,SC_VERSION):
							self.base_window.quantitative_analysis_2.id_label[i].destroy()
					except:
						pass

					# ~ Pmw.initialise(self.base_window)
					# ~ self.base_window.qualitative_analysis_2.tooltip = list(range(SC_VERSION))

					self.base_window.quantitative_analysis_2.id_label = list(range(SC_VERSION))
					index = 0 
					for r in range(0, WELL_ROW):
						for c in range(0, WELL_COLUMN):
							self.base_window.quantitative_analysis_2.id_label[index] = Label(self.base_window.quantitative_analysis_2.id_pos_frame,
													width=6,
													height=3,
													text = self.base_window.quantitative_analysis_2.id_list[index],
													# ~ bg = RESULT_LABEL_BGD_COLOR,
													font = RESULT_LABEL_TXT_FONT)

							# ~ self.base_window.qualitative_analysis_2.tooltip[i] = Pmw.Balloon(self.base_window)
							# ~ self.base_window.qualitative_analysis_2.tooltip[i].bind(self.base_window.qualitative_analysis_2.id_label[i], self.base_window.qualitative_analysis_2.id_list[i])

							if(self.base_window.quantitative_analysis_2.id_list[index] != '#' and self.base_window.quantitative_analysis_2.id_list[index] != 'N/A'):
								self.base_window.quantitative_analysis_2.id_label[index]['bg'] = "lawn green"
								self.base_window.quantitative_analysis_2.id_label[index]['text'] = self.base_window.quantitative_analysis_2.id_list[index]
							else:
								self.base_window.quantitative_analysis_2.id_label[index]['bg'] = "grey80"
								self.base_window.quantitative_analysis_2.id_label[index]['text'] = "N/A"

							self.base_window.quantitative_analysis_2.id_label[index].grid(row=r, column=c, padx=1, pady=1)

							index += 1

					self.back_clicked()
					msg = messagebox.askokcancel("",CreateFile_Language["AllowPutSample Inform"][language])
			
	def back_clicked(self):
		try:
			self.well_button_table_frame.destroy()
			self.property_frame.destroy()
			self.quick_create_frame.destroy()
		except:
			pass

		# In work frame
		# Sample button frame
		self.well_button_table_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR)
		self.well_button_table_frame.pack(side=LEFT)
		self.well_button = list(range(SC_VERSION))
		self.well_frame = list(range(SC_VERSION))

		index = 0
		for r in range(0, WELL_ROW):
			for c in range(0, WELL_COLUMN):
				self.well_frame[index] = Frame(self.well_button_table_frame, highlightbackground = SAMPLE_BUTTON_BGD_COLOR,  highlightthickness = 2, bd=0)
				self.well_button[index] = Button(self.well_frame[index],
											bg = SAMPLE_BUTTON_BGD_COLOR,
											fg = SAMPLE_BUTTON_TXT_COLOR,
											activebackground = SAMPLE_BUTTON_ACTIVE_BGD_COLOR,
											justify = 'left',
											borderwidth = 0,
											text = '#',
											width = SAMPLE_BUTTON_WIDTH,
											height = SAMPLE_BUTTON_HEIGHT)
				# ~ if(i!=47):
				self.well_button[index]['command'] = partial(self.well_button_clicked, index)
				self.well_frame[index].grid(row=r, column=c, padx=1, pady=1)
				self.well_button[index].pack()

				index += 1
		
		self.well_frame[0]['highlightbackground'] = QS_FIRSTWELL_COLOR
		self.well_frame[SC_VERSION - 1]['highlightbackground'] = QS_LASTWELL_COLOR
		
		# Properties frame
		self.property_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR, width=495)
		self.property_frame.pack(fill=BOTH, expand=TRUE, side=LEFT)

		self.property_labelframe = LabelFrame(self.property_frame,
										text = "Sample Properties",
										font  = LABEL_FRAME_TXT_FONT,
										bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
										fg = LABEL_FRAME_TXT_COLOR)
		self.property_labelframe.pack(fill=BOTH, expand=TRUE, padx=10, pady=10)

		self.property_labelframe.rowconfigure(0, weight=1)
		self.property_labelframe.rowconfigure(1, weight=1)
		self.property_labelframe.rowconfigure(2, weight=1)
		self.property_labelframe.rowconfigure(3, weight=4)

		self.well_name_label = Label(self.property_labelframe,
						bg = SAMPLE_BUTTON_CHOOSE_BGD_COLOR,
						fg = LABEL_TXT_COLOR,
						font = SAMPLE_LABEL_TXT_FONT)
		self.well_name_label.grid(row=0, column=0, columnspan=2, sticky=EW)
		
		#quick create frame
		self.quick_create_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR)
		self.quick_create_frame.pack(fill=BOTH, expand=TRUE, side=LEFT)
		
		self.quick_create_labelframe = LabelFrame(self.quick_create_frame,
										text = "Quick Setup",
										font  = LABEL_FRAME_TXT_FONT,
										bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
										fg = LABEL_FRAME_TXT_COLOR)
		self.quick_create_labelframe.pack(fill=BOTH, expand=TRUE, padx=10, pady=10)
		
		self.qc1_frame = Frame(self.quick_create_labelframe,bg = SAMPLE_BUTTON_FRAME_BDG_COLOR)
		self.qc1_frame.pack(fill=BOTH, expand=TRUE)
		self.qc2_frame = Frame(self.quick_create_labelframe, bg = SAMPLE_BUTTON_FRAME_BDG_COLOR)
		self.qc2_frame.pack(fill=BOTH, expand=TRUE, side=BOTTOM) 
		
		self.first_well_button = Button(self.qc1_frame,
								text = "First well",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = "lawn green",
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.first_well_button_clicked)
		self.first_well_button.grid(row=0, column=0, ipadx=10, ipady=10, pady=20)
		
		self.last_well_button = Button(self.qc1_frame,
								text = "Last well",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.last_well_button_clicked)
		self.last_well_button.grid(row=0, column=1, ipadx=10, ipady=10, pady=20)
		
		self.first_well_label = Label(self.qc1_frame,
						text = "A1",
						bg = QS_FIRSTWELL_COLOR,
						fg = LABEL_TXT_COLOR,
						font = SAMPLE_LABEL_TXT_FONT)
		self.first_well_label.grid(row=1, column=0)
		
		self.last_well_label = Label(self.qc1_frame,
						text = "H6",
						bg = QS_LASTWELL_COLOR,
						fg = LABEL_TXT_COLOR,
						font = SAMPLE_LABEL_TXT_FONT)
		self.last_well_label.grid(row=1, column=1)
	   
		self.quick_create_button = Button(self.qc2_frame,
								text = "Set",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.quick_create_button_clicked)
		self.quick_create_button.pack(ipadx=50, ipady=10)


		self.base_window.forget_page()
		if(self.direct_create==0):
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
		elif(self.direct_create==1):
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_2)
		else:
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_2)
		self.base_window.switch_page()
	
	def first_well_button_clicked(self):
		self.first_well_button['bg'] = "lawn green"
		self.last_well_button['bg'] = "grey75"
	def last_well_button_clicked(self):
		self.first_well_button['bg'] = "grey75"
		self.last_well_button['bg'] = "lawn green"	
		
	def quick_create_button_clicked(self):
		msg = messagebox.askquestion("",CreateFile_Language["QuickSetup Confirm"][language])
		if(msg == 'yes'):
			auto_sample_name = 0
			for i in range(0,SC_VERSION):
				self.well_button[i]['text'] = "#"
				self.well_button[i]['bg'] = SAMPLE_BUTTON_BGD_COLOR
			for i in  range(self.first_well_index, self.last_well_index + 1):
				auto_sample_name += 1
				self.well_button[i]['text'] = "Sample " + str(auto_sample_name) 
				self.well_button[i]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
					

class QualitativeSavedFrame(QualitativeAnalysisFrame1):
	def __init__(self, container):
		super().__init__(container)
		self.title_label['text'] = "Qualitative Program"

		self.next_button.destroy()

		self.delete_button = Button(self.button_frame,
								text = "Delete",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.delete_clicked)
		self.delete_button.pack(side=RIGHT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

	def delete_clicked(self):
		msg = messagebox.askquestion("","Are you sure you want to delete this program ?")
		if(msg == 'yes'):
			os.remove(programs_qualitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
			self.load_program()
			self.experiment_name_text['state'] = "normal"
			self.user_name_text['state'] = "normal"
			self.comment_text['state'] = "normal"
			self.experiment_name_text.delete('1.0',END)
			self.user_name_text.delete('1.0',END)
			self.comment_text.delete('1.0',END)
			self.experiment_name_text['state'] = "disabled"
			self.user_name_text['state'] = "disabled"
			self.comment_text['state'] = "disabled"
			messagebox.showinfo("","Deleted")


	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.saved_program)
		self.base_window.switch_page()

class QuantitativeSavedFrame(QuantitativeAnalysisFrame1):
	def __init__(self, container):
		super().__init__(container)
		self.title_label['text'] = "Quantitative Program"

		self.next_button.destroy()

		self.delete_button = Button(self.button_frame,
								text = "Delete",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.delete_clicked)
		self.delete_button.pack(side=RIGHT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

	def delete_clicked(self):
		msg = messagebox.askquestion("","Are you sure you want to delete this program ?")
		if(msg == 'yes'):
			os.remove(programs_quantitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
			self.load_program()
			self.experiment_name_text['state'] = "normal"
			self.user_name_text['state'] = "normal"
			self.comment_text['state'] = "normal"
			self.experiment_name_text.delete('1.0',END)
			self.user_name_text.delete('1.0',END)
			self.comment_text.delete('1.0',END)
			self.experiment_name_text['state'] = "disabled"
			self.user_name_text['state'] = "disabled"
			self.comment_text['state'] = "disabled"
			messagebox.showinfo("","Deleted")

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.saved_program)
		self.base_window.switch_page()

class AnalysisFrame(Frame):
	def __init__(self, container, base_window):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = base_window

		qualitative_button = Button(self,
									text = "Qualitative",
									font = MAIN_FUCNTION_BUTTON_FONT,
									# ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
									# ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
									bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
									fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.qualitative_clicked)
		qualitative_button.pack(fill=BOTH, expand=TRUE, side=LEFT, padx=90, pady=210)
		quantitative_button = Button(self,
									text = "Quantitative",
									font = MAIN_FUCNTION_BUTTON_FONT,
									# ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
									# ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
									bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
									fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.quantitative_clicked)
		quantitative_button.pack(fill=BOTH, expand=TRUE, side=RIGHT, padx=90, pady=210)

	def qualitative_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_1)
		self.base_window.switch_page()
		self.base_window.qualitative_analysis_1.load_program()
	def quantitative_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_1)
		self.base_window.switch_page()
		self.base_window.quantitative_analysis_1.load_program()


class NewProgramFrame(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		self.title_label = Label(self.title_frame,
								text = "NEW PROGRAM",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		qualitative_button = Button(self.work_frame,
									text = "Qualitative",
									font = MAIN_FUCNTION_BUTTON_FONT,
									# ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
									# ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
									bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
									fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.qualitative_clicked)
		qualitative_button.pack(fill=BOTH, expand=TRUE, side=LEFT, padx=90, pady=100, ipadx=25, ipady=20)

		quantitative_button = Button(self.work_frame,
									text = "Quantitative",
									font = MAIN_FUCNTION_BUTTON_FONT,
									# ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
									# ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
									bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
									fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.quantitative_clicked)
		quantitative_button.pack(fill=BOTH, expand=TRUE, side=LEFT, padx=90, pady=100, ipadx=25, ipady=20)

		self.back_button = Button(self.button_frame,
								text = "Back",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

	def qualitative_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_1)
		self.base_window.switch_page()
		
	def quantitative_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_1)
		self.base_window.switch_page()

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
		self.base_window.switch_page()
	
class SavedProgramFrame(NewProgramFrame):
	def __init__(self, container):
		super().__init__(container)
		self.title_label['text'] = "Saved Program"

	def qualitative_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_saved_program)
		self.base_window.switch_page()
		self.base_window.qualitative_saved_program.load_program()
	def quantitative_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_saved_program)
		self.base_window.switch_page()
		self.base_window.quantitative_saved_program.load_program()

class ServerSettingFrame(Frame):
	def __init__(self, container):
		super().__init__(container)
		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container
		self.account_active = 0

		# 3 main frame
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = "Server",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		self.login_label = Label(self.work_frame,text ="LOGIN",
								bg=LABEL_BGD_COLOR,
								fg=LABEL_TXT_COLOR,
								font=LOGIN_LABEL_TXT_FONT)
		self.login_label.grid(row=0, column=0, columnspan=3, pady=30, ipadx=1, ipady=1)

		ip_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="IP address :")
		ip_label.grid(row=1, column=0, sticky=E, pady=15, padx=30)
		user_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="User :")
		user_label.grid(row=2, column=0, sticky=E, pady=15, padx=30)
		password_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="Password :")
		password_label.grid(row=3, column=0, sticky=E, pady=15, padx=30)
		path_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="Folder path :")
		path_label.grid(row=4, column=0, sticky=E, pady=15, padx=30)

		self.ip_entry = Entry(self.work_frame, width=30, justify='left', font=('Courier',14))
		self.ip_entry.grid(row=1, column=1, sticky=W)
		self.user_entry = Entry(self.work_frame, width=30, justify='left', font=('Courier',14))
		self.user_entry.grid(row=2, column=1, sticky=W)
		self.pass_entry = Entry(self.work_frame, width=30, show='◼', justify='left', font=('Courier',14))
		self.pass_entry.grid(row=3, column=1, sticky=W)
		self.path_entry = Entry(self.work_frame, width=30, justify='left', font=('Courier',14))
		self.path_entry.grid(row=4, column=1, sticky=W)

		hide_var = IntVar()
		def hide_charaters():
			if(hide_var.get()==0):
				self.pass_entry['show']=""
			else:
				self.pass_entry['show']="◼"
		self.hidepass_checkbutton = Checkbutton(self.work_frame, variable=hide_var, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text = "Hide characters",onvalue=1, offvalue=0, command=hide_charaters)
		self.hidepass_checkbutton.select()
		self.hidepass_checkbutton.grid(row=3, column=2, sticky=W, padx=10)

		self.connect_button = Button(self.work_frame,
								text = "Connect",
								bg = LOGIN_BUTTON_BGD_COLOR,
								fg = LOGIN_BUTTON_TXT_COLOR,
								font = LOGIN_BUTTON_TXT_FONT,
								borderwidth = 0,
								command = self.connect_clicked)
		self.connect_button.grid(row=5, column=0, columnspan=3, ipadx=30, ipady=10, pady=20)


		# In button frame
		self.back_button = Button(self.button_frame,
								text = "Back",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

		# ~ self.save_button = Button(self.button_frame,
								# ~ text = "Back",
								# ~ font = SWITCH_PAGE_BUTTON_FONT,
								# ~ # width = SWITCH_PAGE_BUTTON_WIDTH,
								# ~ # height = SWITCH_PAGE_BUTTON_HEIGHT,
								# ~ bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								# ~ fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								# ~ borderwidth = 0,
								# ~ command = self.save_clicked)
		# ~ self.save_button.pack(ipadx=30, ipady=10, side=LEFT, sticky=W)

		self.check_status()

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.connect)
		self.base_window.switch_page()

	def connect_clicked(self):
		if(self.connect_button['text']=='Connect'):
			self.ip_set = self.ip_entry.get().strip()
			self.user_set = self.user_entry.get().strip()
			self.password_set = self.pass_entry.get().strip()
			self.path_set = self.path_entry.get().strip()
			if(self.ip_set==''):
				messagebox.showwarning("","Please enter yours IP address.")
			elif(self.user_set==''):
				messagebox.showwarning("","Please enter yours user name.")
			elif(self.password_set==''):
				messagebox.showwarning("","Please enter yours password.")
			# ~ elif(self.path_set==''):
				# ~ messagebox.showwarning("","Please enter yours folder path.")
			else:
				try:
					#FTP:
					ftp = FTP(self.ip_set, self.user_set, self.password_set)
					ftp.cwd(self.path_set + '/')
					ftp.quit()
					
					
					#LAN:
					# ~ os.system('sudo mount -t cifs -o username=' + self.user_set + ',password=' + self.password_set + ' //' + self.ip_set + '/' + self.path_set + ' /home/pi/Server')
					
					tc= open(working_dir + '/.server.txt',"w")
					tc.writelines('1\n')
					tc.writelines(self.ip_set+"\n")
					tc.writelines(self.user_set+"\n")
					tc.writelines(self.password_set+"\n")
					tc.writelines(self.path_set+"\n")
					tc.close()

					self.check_status()

					messagebox.showinfo("", "Connection successful.")
				except Exception as e :
					error = messagebox.showerror("Could not connect to server.",str(e))
					if(error=='ok'):
						pass
		else:
			msg = messagebox.askquestion("","Do you want to disconnect from the server ?")
			if(msg=='yes'):
				tc= open(working_dir + '/.server.txt',"w")
				tc.writelines('0\n')
				tc.writelines("\n")
				tc.writelines("\n")
				tc.writelines("\n")
				tc.writelines("\n")
				tc.close()
				
				#LAN:
				# ~ try:
					# ~ os.system('sudo umount //' + self.ip_set + '/' + self.path_set)
				# ~ except:
					# ~ pass
				
				self.check_status()
				
	def check_status(self):
		fr = open(working_dir + "/.server.txt","r")
		self.server_active = int(fr.readline().strip())
		self.ip_set = fr.readline().strip()
		self.user_set = fr.readline().strip()
		self.password_set = fr.readline().strip()
		self.path_set = fr.readline().strip()
		fr.close()

		if(self.server_active == 0):
			self.connect_button['text'] = 'Connect'
			self.login_label['text'] = 'Server'
			self.login_label['fg'] = 'black'

			self.ip_entry['state'] = 'normal'
			self.user_entry['state'] = 'normal'
			self.pass_entry['state'] = 'normal'
			self.path_entry['state'] = 'normal'
			self.ip_entry.delete(0,END)
			self.user_entry.delete(0,END)
			self.pass_entry.delete(0,END)
			self.path_entry.delete(0,END)

			self.pass_entry['show']="◼"
			self.hidepass_checkbutton['state'] = 'normal'
			self.hidepass_checkbutton.select()
		else:
			self.connect_button['text'] = 'Disconnect'
			self.login_label['text'] = 'You are already connected to the server ✔️'
			self.login_label['fg'] = 'green3'

			self.ip_entry['state'] = 'normal'
			self.user_entry['state'] = 'normal'
			self.pass_entry['state'] = 'normal'
			self.path_entry['state'] = 'normal'
			self.ip_entry.delete(0,END)
			self.user_entry.delete(0,END)
			self.pass_entry.delete(0,END)
			self.path_entry.delete(0,END)
			self.ip_entry.insert(END, self.ip_set)
			self.user_entry.insert(END, self.user_set)
			self.pass_entry.insert(END, self.password_set)
			self.path_entry.insert(END, self.path_set)
			self.ip_entry['state'] = 'disabled'
			self.user_entry['state'] = 'disabled'
			self.pass_entry['state'] = 'disabled'
			self.path_entry['state'] = 'disabled'
			
			self.pass_entry['show']="◼"
			self.hidepass_checkbutton.select()
			self.hidepass_checkbutton['state'] = 'disable'
			

class EmailSettingFrame(Frame):
	def __init__(self, container):
		super().__init__(container)
		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container
		self.account_active = 0

		# 3 main frame
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = "EMAIL",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		self.login_label = Label(self.work_frame,text ="LOGIN",
								bg=LABEL_BGD_COLOR,
								fg=LABEL_TXT_COLOR,
								font=LOGIN_LABEL_TXT_FONT)
		self.login_label.grid(row=0, column=0, columnspan=2, pady=30, ipadx=1, ipady=1)

		user_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="Email :")
		user_label.grid(row=1, column=0, sticky=E, pady=20, padx=30)
		pass_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="Device Password :")
		pass_label.grid(row=2, column=0, sticky=NE, pady=3, padx=30)

		self.user_entry = Entry(self.work_frame, width=30, justify='right', font=('Courier',14))
		self.user_entry.grid(row=1, column=1, sticky=W)
		self.pass_entry = Entry(self.work_frame, width=30, show='◼', justify='right', font=('Courier',14))
		self.pass_entry.grid(row=2, column=1, sticky=NW)

		hide_var = IntVar()
		def hide_charaters():
			if(hide_var.get()==0):
				self.pass_entry['show']=""
			else:
				self.pass_entry['show']="◼"
		self.hidepass_checkbutton = Checkbutton(self.work_frame, variable=hide_var, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text = "Hide characters",onvalue=1, offvalue=0, command=hide_charaters)
		self.hidepass_checkbutton.select()
		self.hidepass_checkbutton.grid(row=2, column=1, pady=32, sticky=SE)

		self.login_button = Button(self.work_frame,
								text = "Login",
								bg = LOGIN_BUTTON_BGD_COLOR,
								fg = LOGIN_BUTTON_TXT_COLOR,
								font = LOGIN_BUTTON_TXT_FONT,
								borderwidth = 0,
								command = self.login_clicked)
		self.login_button.grid(row=3, column=0, columnspan=2, ipadx=30, ipady=10, pady=30)


		# In button frame
		self.back_button = Button(self.button_frame,
								text = "Back",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(ipadx=30, ipady=10, side=LEFT)


	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.connect)
		self.base_window.switch_page()

	def check_status(self):
		fr = open(working_dir + "/.email.txt","r")
		self.account_active = int(fr.readline())
		self.email_address = fr.readline()
		self.email_password = fr.readline()
		fr.close()

		if(self.account_active==1):
			self.login_label['text'] = "You are already logged in ✔️"
			self.login_label['fg'] = 'green2'

			self.user_entry['state'] = 'normal'
			self.pass_entry['state'] = 'normal'
			self.user_entry.delete(0,END)
			self.pass_entry.delete(0,END)
			self.user_entry.insert(END, self.email_address)
			self.pass_entry.insert(END, self.email_password)
			self.user_entry['state'] = 'disabled'
			self.pass_entry['state'] = 'disabled'

			self.pass_entry['show']="◼"
			self.hidepass_checkbutton.select()
			self.hidepass_checkbutton['state'] = 'disabled'

			self.login_button['text'] = 'Logout'
		else:
			self.login_label['text'] = "LOGIN"
			self.login_label['fg'] = LABEL_TXT_COLOR

			self.user_entry['state'] = 'normal'
			self.pass_entry['state'] = 'normal'
			self.user_entry.delete(0,END)
			self.pass_entry.delete(0,END)

			self.login_button['text'] = 'Login'

			self.hidepass_checkbutton['state'] = 'normal'

	def login_clicked(self):
		if(self.login_button['text'] == 'Login'):
			if(self.user_entry.get()==''):
				messagebox.showwarning("","Please enter the email address!")
			elif(self.pass_entry.get()==''):
				messagebox.showwarning("","Please enter the password !")
			else:
				self.email_address = self.user_entry.get()
				self.email_password = self.pass_entry.get()

				addressToVerify = self.email_address
				match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', addressToVerify)
				if(match == None):
					messagebox.showerror("","Email syntax error")
				else:
					domain_name = self.email_address.split('@')[1]
					records = dns.resolver.query(domain_name, 'MX')
					mxRecord = records[0].exchange
					mxRecord = str(mxRecord)

					host = socket.gethostname()

					server = smtplib.SMTP()
					server.set_debuglevel(0)

					server.connect(mxRecord)
					server.helo(host)
					server.mail('me@domain.com')
					code, message = server.rcpt(str(addressToVerify))
					server.quit()

					if(code==250):
						server=smtplib.SMTP('smtp.gmail.com:587')
						server.starttls()
						try:
							server.login(self.email_address,self.email_password)
							save_file = open(working_dir + '/.email.txt',"w")
							save_file.writelines('1' + "\n")
							save_file.writelines(self.email_address + "\n")
							save_file.writelines(self.email_password + "\n")
							save_file.close()
							messagebox.showinfo("", "Login Success !")
							self.check_status()
						except:
							messagebox.showerror("","Your password was incorrect\rPlease try again !")
						server.quit()
					else:
						messagebox.showerror("","Your email address was incorrect\rPlease try again !")
		else:
			msg = messagebox.askquestion("","Do you want to logout ?")
			if(msg=='yes'):
				save_file = open(working_dir + '/.email.txt',"w")
				save_file.writelines('0' + "\n")
				save_file.writelines("\n")
				save_file.writelines("\n")
				save_file.close()

				self.account_active = 0

				self.login_label['text'] = "LOGIN"
				self.login_label['fg'] = LABEL_TXT_COLOR

				self.user_entry['state'] = 'normal'
				self.pass_entry['state'] = 'normal'
				self.user_entry.delete(0,END)
				self.pass_entry.delete(0,END)

				self.login_button['text'] = 'Login'
				self.hidepass_checkbutton['state'] = 'normal'

class ProgramFrame(Frame):
	def __init__(self, container, base_window):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = base_window

		new_program_button = Button(self,
									text = "New Program",
									font = MAIN_FUCNTION_BUTTON_FONT,
									# ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
									# ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
									bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
									fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.new_program_clicked)
		new_program_button.pack(fill=BOTH, expand=TRUE, side=LEFT, padx=90, pady=210)
		saved_program_button = Button(self,
									text = "Saved Program",
									font = MAIN_FUCNTION_BUTTON_FONT,
									# ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
									# ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
									bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
									fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.saved_program_clicked)
		saved_program_button.pack(fill=BOTH, expand=TRUE, side=RIGHT, padx=90, pady=210)

	def new_program_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_program)
		self.base_window.switch_page()
	def saved_program_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.saved_program)
		self.base_window.switch_page()

class SetIdFrame(Frame):
	def __init__(self, container, base_window):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = base_window

		set_button = Button(self,
									text = "Create ID file",
									font = MAIN_FUCNTION_BUTTON_FONT,
									# ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
									# ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
									bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
									fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.create_clicked)
		set_button.pack(ipady=20, ipadx=20, expand=TRUE)

	def create_clicked(self):
		del self.base_window.id_create
		self.base_window.id_create = IDCreateFrame(self.base_window)
		self.base_window.frame_list.append(self.base_window.id_create)

		self.base_window.id_create.direct_create = 0
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.id_create)
		self.base_window.switch_page()

class ViewResultFrame(Frame):
	def __init__(self, container):
		super().__init__(container)
		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container
		self.account_active = 0

		# 3 main frame
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = ViewResult_Language['Title Label'][language],
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		#In work frame
		self.infor_frame = LabelFrame(self.work_frame,
								bg= LABEL_FRAME_BGD_COLOR,
								font = LABELFRAME_TXT_FONT,
								fg = LABEL_FRAME_TXT_COLOR,
								width=150,
								text = ViewResult_Language['Information LabelFrame'][language])
		self.infor_frame.pack(side=LEFT, anchor=W, pady=5, padx=5)

		self.report_frame = ScrollableFrame2(self.work_frame)
		self.report_frame.pack(side=RIGHT, pady=5)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = ViewResult_Language['Back Button'][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(ipadx=30, ipady=10,side=LEFT)

		self.open_button = Button(self.button_frame,
								text = ViewResult_Language['Open Button'][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.open_clicked)
		self.open_button.pack(ipadx=30, ipady=10, side=LEFT, padx=225)

		# ~ self.delete_button = Button(self.button_frame,
								# ~ text = "Delete",
								# ~ font = SWITCH_PAGE_BUTTON_FONT,
								# ~ # width = SWITCH_PAGE_BUTTON_WIDTH,
								# ~ # height = SWITCH_PAGE_BUTTON_HEIGHT,
								# ~ bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								# ~ fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								# ~ borderwidth = 0,
								# ~ state = 'disabled',
								# ~ command = self.delete_clicked)
		# ~ self.delete_button.pack(ipadx=30, ipady=10, side=LEFT)

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
		self.base_window.switch_page()

	def open_clicked(self):
		sample_button_list = list(range(SC_VERSION + 1))
		result_button_list = list(range(SC_VERSION + 1))
		position_button_list = list(range(SC_VERSION + 1))

		self.path = filedialog.askopenfilename(initialdir=results_path, filetypes=[('Excel file','.xlsx')])
		if self.path is not None:
			try:
				self.infor_frame.destroy()
				self.infor_frame = LabelFrame(self.work_frame,
								bg= LABEL_FRAME_BGD_COLOR,
								font = LABELFRAME_TXT_FONT,
								fg = LABEL_FRAME_TXT_COLOR,
								width=150,
								text = ViewResult_Language['Information LabelFrame'][language])
				self.infor_frame.pack(side=LEFT, anchor=N, pady=5, padx=5)
			except:
				pass

			try:
				self.report_frame.destroy()
				self.report_frame = ScrollableFrame2(self.work_frame)
				self.report_frame.pack(side=RIGHT, pady=5)
			except:
				pass

			wb = load_workbook(self.path)
			sheet = wb.active

			info1_button_list = Button(self.infor_frame,
					fg = LABEL_TXT_COLOR,
					font = LABEL_TXT_FONT,
					text= sheet['B13'].value,
					width=30,
					bg = 'lavender',
					borderwidth = 0)
			info1_button_list.grid(row=0, column=0, sticky=EW, padx=1, pady=1)
			info2_button_list = Button(self.infor_frame,
					fg = LABEL_TXT_COLOR,
					font = LABEL_TXT_FONT,
					text= sheet['B14'].value,
					width=30,
					bg = 'lavender',
					borderwidth = 0)
			info2_button_list.grid(row=1, column=0, sticky=EW, padx=1, pady=1)
			info3_button_list = Button(self.infor_frame,
					fg = LABEL_TXT_COLOR,
					font = LABEL_TXT_FONT,
					text= sheet['B15'].value,
					width=30,
					bg = 'lavender',
					borderwidth = 0)
			info3_button_list.grid(row=2, column=0, sticky=EW, padx=1, pady=1)

			for i in range(0,SC_VERSION + 1):
				sample_pos = 'B' + str(i+17)
				sample_button_list[i] = Button(self.report_frame.scrollable_frame,
						fg = LABEL_TXT_COLOR,
						font = LABEL_TXT_FONT,
						text= sheet[sample_pos].value,
						width=30,
						bg = 'lavender',
						borderwidth = 0)
				sample_button_list[i].grid(row=i, column=0, sticky=EW, padx=1, pady=1)

				position_pos = 'C' + str(i+17)
				position_button_list[i] = Button(self.report_frame.scrollable_frame,
						fg = LABEL_TXT_COLOR,
						font = LABEL_TXT_FONT,
						text= sheet[position_pos].value,
						width=10,
						bg = 'lavender',
						borderwidth = 0)
				position_button_list[i].grid(row=i, column=1, sticky=EW, padx=1, pady=1)

				result_pos = 'D' + str(i+17)
				result_button_list[i] = Button(self.report_frame.scrollable_frame,
						fg = LABEL_TXT_COLOR,
						font = LABEL_TXT_FONT,
						text= sheet[result_pos].value,
						width=22,
						bg = 'lavender',
						borderwidth = 0)
				result_button_list[i].grid(row=i, column=3, sticky=EW, padx=1, pady=1)
			wb.close()


class QualitativeCalibListFrame(QualitativeAnalysisFrame1):
	def __init__(self, container):
		super().__init__(container)
		self.title_label['text'] = "CALIBRATION"

		self.edit_frame = Frame(self.work_frame,
								bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.edit_frame.grid(row=1, column=0, sticky=EW)

		new_button = Button(self.edit_frame,
								text = "New",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.new_clicked)
		new_button.grid(row=0, column=0, ipadx=35, ipady=10)

		recalibration_button = Button(self.edit_frame,
								text = "Recalibration",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.recalibration_clicked)
		recalibration_button.grid(row=0, column=1, padx=10, ipadx=10, ipady=10)

		delete_button = Button(self.edit_frame,
								text = "Delete",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.delete_clicked)
		delete_button.grid(row=0, column=2, ipadx=30, ipady=10)

		self.next_button.pack_forget()

	def delete_clicked(self):
		self.program_name = self.experiment_name_text.get("1.0","end-1c")
		if(len(self.program_name) != 0):
			msg = messagebox.askquestion("","Are you sure you want to delete this kit ?")
			if(msg == 'yes'):
				os.remove(programs_qualitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
				self.load_program()
				self.experiment_name_text['state'] = "normal"
				self.user_name_text['state'] = "normal"
				self.pfi_text['state'] = "normal"
				self.comment_text['state'] = "normal"
				self.experiment_name_text.delete('1.0',END)
				self.user_name_text.delete('1.0',END)
				self.pfi_text.delete('1.0', END)
				self.comment_text.delete('1.0',END)
				self.experiment_name_text['state'] = "disabled"
				self.user_name_text['state'] = "disabled"
				self.pfi_text['state'] = "disabled"
				self.comment_text['state'] = "disabled"
				messagebox.showinfo("","Deleted")
		else:
			messagebox.showwarning("","Please select the kit !")


	def recalibration_clicked(self):
		self.program_name = self.experiment_name_text.get("1.0","end-1c")
		if(len(self.program_name) != 0):
			msg = messagebox.askquestion("","Are you sure you want to recalibrate this kit ?")
			if(msg == 'yes'):
				os.remove(programs_qualitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
				self.load_program()

				self.base_window.new_qualitative_1.experiment_name_entry.delete(0,END)
				self.base_window.new_qualitative_1.user_name_entry.delete(0,END)
				self.base_window.new_qualitative_1.comments_text.delete('1.0',END)

				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_1)
				self.base_window.switch_page()

				self.base_window.new_qualitative_1.experiment_name_entry.insert(0, self.experiment_name_text.get("1.0",'end-1c'))
				self.base_window.new_qualitative_1.user_name_entry.insert(0, self.user_name_text.get("1.0",'end-1c'))
				self.base_window.new_qualitative_1.comments_text.insert('1.0', self.comment_text.get("1.0",'end-1c'))

				self.experiment_name_text['state'] = "normal"
				self.user_name_text['state'] = "normal"
				self.pfi_text['state'] = "normal"
				self.comment_text['state'] = "normal"
				self.experiment_name_text.delete('1.0',END)
				self.user_name_text.delete('1.0',END)
				self.pfi_text.delete('1.0', END)
				self.comment_text.delete('1.0',END)
				self.experiment_name_text['state'] = "disabled"
				self.user_name_text['state'] = "disabled"
				self.pfi_text['state'] = "disabled"
				self.comment_text['state'] = "disabled"

		else:
			messagebox.showwarning("","Please select the kit !")

	def new_clicked(self):
		self.base_window.new_qualitative_1.experiment_name_entry.delete(0,END)
		self.base_window.new_qualitative_1.user_name_entry.delete(0,END)
		self.base_window.new_qualitative_1.comments_text.delete('1.0',END)

		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_1)
		self.base_window.switch_page()

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_option)
		self.base_window.switch_page()


class QuantitativeCalibListFrame(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container
		
		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.work_frame.pack_propagate(0)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X, expand=TRUE)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = QuantitativeKit_Language["Title Label"][language],
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		self.program_frame = ScrollableFrame(self.work_frame)
		self.program_frame.grid(row=0, column=0, pady=17, sticky=S)

		self.info_labelframe = LabelFrame(self.work_frame,
								text = QuantitativeKit_Language["Title Label"][language],
								font = LABELFRAME_TXT_FONT,
								bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.info_labelframe.grid(row=0, column=1, rowspan=3, ipadx=10, ipady=5, padx=10, pady=28)

		experiment_name_label = Label(self.info_labelframe,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							text = QuantitativeKit_Language["Title Label"][language],
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT)
		experiment_name_label.grid(row=0, column=0, padx=10, sticky=E)


		self.experiment_name_text = Text(self.info_labelframe,
							width = 30,
							height = 1,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT,)
							# ~ state = 'disabled')
		self.experiment_name_text.grid(row=0, column=1, padx=5, pady=30)

		properties_label = Label(self.info_labelframe,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							text = QuantitativeKit_Language["Title Label"][language],
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT)
		properties_label.grid(row=3, column=0, padx=10, pady=30, sticky=NE)

		self.properties_text = Text(self.info_labelframe,
							width = 30,
							height = 12,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							fg = LABEL_TXT_COLOR,
							font = LABEL_TXT_FONT,
							state = 'disabled')
		self.properties_text.grid(row=3, column=1, padx=5, pady=30)
		
		self.edit_frame = Frame(self.work_frame,
								bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.edit_frame.grid(row=1, column=0)

		new_button = Button(self.edit_frame,
								text = QuantitativeKit_Language["Title Label"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.new_clicked)
		new_button.grid(row=0, column=0, ipadx=35, ipady=10, padx=10)

		# ~ modify_button = Button(self.edit_frame,
								# ~ text = "Modify",
								# ~ font = SWITCH_PAGE_BUTTON_FONT,
								# ~ # width = SWITCH_PAGE_BUTTON_WIDTH,
								# ~ # height = SWITCH_PAGE_BUTTON_HEIGHT,
								# ~ bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								# ~ fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								# ~ borderwidth = 0,
								# ~ command = self.modify_clicked)
		# ~ modify_button.grid(row=0, column=1, padx=10, ipadx=30, ipady=10)

		delete_button = Button(self.edit_frame,
								text = QuantitativeKit_Language["Title Label"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.delete_clicked)
		delete_button.grid(row=0, column=1, ipadx=30, ipady=10, padx=10)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = "Back",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(side=LEFT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

	def delete_clicked(self):
		self.program_name = self.experiment_name_text.get("1.0","end-1c")
		if(len(self.program_name) != 0):
			msg = messagebox.askquestion("",QuantitativeCalibList_Language["DeleteKit Confirm"][language])
			if(msg == 'yes'):
				os.remove(programs_quantitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
				self.load_program()
				self.experiment_name_text['state'] = "normal"
				self.properties_text['state'] = "normal"
				self.experiment_name_text.delete('1.0',END)
				self.properties_text.delete('1.0',END)
				self.experiment_name_text['state'] = "disabled"
				self.properties_text['state'] = "disabled"
				messagebox.showinfo("","Deleted")
		else:
			messagebox.showwarning("",QuantitativeCalibList_Language["Kit Empty"][language])


	# ~ def modify_clicked(self):
		# ~ self.program_name = self.experiment_name_text.get("1.0","end-1c")
		# ~ if(len(self.program_name) != 0):
			# ~ os.remove(programs_quantitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
			# ~ self.load_program()

			# ~ self.base_window.new_quantitative_1.experiment_name_entry.delete(0,END)
			# ~ self.base_window.new_quantitative_1.properties_text.delete('1.0',END)

			# ~ self.base_window.forget_page()
			# ~ self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_1)
			# ~ self.base_window.switch_page()

			# ~ self.base_window.new_quantitative_1.experiment_name_entry.insert(0, self.experiment_name_text.get("1.0",'end-1c'))
			# ~ self.base_window.new_quantitative_1.properties_text.insert('1.0', self.comment_text.get("1.0",'end-1c'))

			# ~ os.remove(programs_quantitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
			# ~ self.load_program()
			# ~ self.experiment_name_text['state'] = "normal"
			# ~ self.properties_text['state'] = "normal"
			# ~ self.experiment_name_text.delete('1.0',END)
			# ~ self.comment_text.delete('1.0',END)
			# ~ self.experiment_name_text['state'] = "disabled"
			# ~ self.properties_text['state'] = "disabled"
		# ~ else:
			# ~ messagebox.showwarning("","Please select the kit !")

	def new_clicked(self):
		self.base_window.new_quantitative_1.experiment_name_entry.delete(0,END)
		self.base_window.new_quantitative_1.properties_text.delete('1.0',END)
		
		# ~ create_kit_frame = Frame(self.work_frame,
								 # ~ width = 200,
								 # ~ height = 300,
								 # ~ bg = 'grey80')
		# ~ create_kit_frame.place(x=400, y=30)
		try: 
			self.info_labelframe.grid_forget()
		except: 
			pass
			
		self.create_kit_labelframe = LabelFrame(self.work_frame,
								text = QuantitativeCalibList_Language["CreateKit LabelFrame"][language],
								font = LABELFRAME_TXT_FONT,
								bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.create_kit_labelframe.grid(row=0, column=1, rowspan=3, ipadx=10, ipady=5, padx=10, pady=28)
		# ~ self.base_window.forget_page()
		# ~ self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_1)
		# ~ self.base_window.switch_page()

	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_option)
		self.base_window.switch_page()
		
	def load_program(self):
		try:
			for i in range(len(self.program_button)):
				self.program_button[i].destroy()
		except:
			pass
		self.program_button = list(range(100))
		for file in os.listdir(programs_quantitative_path):
			self.program_button[os.listdir(programs_qualitative_path).index(file)] = Button(self.program_frame.scrollable_frame,
									text=file[:(len(file)-5)],
									font = PROGRAM_BUTTON_TXT_FONT,
									bg = PROGRAM_BUTTON_BGD_COLOR,
									fg = PROGRAM_BUTTON_TXT_COLOR,
									width = 51,
									borderwidth = 0)
			self.program_button[os.listdir(programs_qualitative_path).index(file)]['command'] = partial(self.program_clicked, os.listdir(programs_qualitative_path).index(file))
			self.program_button[os.listdir(programs_qualitative_path).index(file)].pack(pady=2, ipady=5, fill=BOTH, expand=TRUE)


class QualitativeOptionFrame(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.work_frame.pack_propagate(0)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X, expand=TRUE)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = "SCREENING MODE",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		self.calibration_button = Button(self.work_frame,
									text = "Create kit",
									font = MAIN_FUCNTION_BUTTON_FONT,
									bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
									fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.calibration_clicked)
		self.calibration_button.grid(row=0, column=0, ipadx=30, ipady=20, padx=80, pady=168)

		self.analysis_button = Button(self.work_frame,
									text = "Analysis",
									font = MAIN_FUCNTION_BUTTON_FONT,
									bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
									fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.analysis_clicked)
		self.analysis_button.grid(row=0, column=1, ipadx=30, ipady=20, padx=80, pady=168)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = "Back",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(side=LEFT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

	def calibration_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_calib_list)
		self.base_window.switch_page()
		self.base_window.qualitative_calib_list.load_program()
	def analysis_clicked(self):
		self.dye_label_frame = LabelFrame(self.work_frame,
											width = 450,
											height = 50,
											text = "NUMBER OF DYES",
											bg = 'dodger blue')
		self.dye_label_frame.place(x=20, y=100)

		self.one_dye_button = Button(self.dye_label_frame,
					text = "1 Dye",
					font = SWITCH_PAGE_BUTTON_FONT,

					bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
					fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
					borderwidth = 0,
					command = self.one_dye_clicked)
		self.one_dye_button.pack(side=LEFT, padx=85, pady=50, ipady=10, ipadx=20)

		self.two_dye_button = Button(self.dye_label_frame,
					text = "2 Dyes",
					font = SWITCH_PAGE_BUTTON_FONT,
					bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
					fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
					borderwidth = 0,
					command = self.two_dye_clicked)
		self.two_dye_button.pack(side=LEFT, padx=85, pady=50, ipady=10, ipadx=20, anchor=W)
		
		self.cancel_button = Button(self.dye_label_frame,
					text = "X",
					font = SWITCH_PAGE_BUTTON_FONT,
					width = 1,
					height = 1,
					bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
					fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
					borderwidth = 0,
					command = self.cancel_clicked)
		self.cancel_button.place(x=516,y=-8)
		
	def back_clicked(self):
		self.cancel_clicked()
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
		self.base_window.switch_page()
	
	def one_dye_clicked(self):
		self.number_of_dyes = 1 
		self.cancel_clicked()
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
		self.base_window.switch_page()
	def two_dye_clicked(self):
		self.number_of_dyes = 2
		self.cancel_clicked()
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
		self.base_window.switch_page()
		
	def cancel_clicked(self):
		try:
			self.dye_label_frame.place_forget()
		except:
			pass

class QuantitativeOptionFrame(QualitativeOptionFrame):
	def __init__(self, container):
		super().__init__(container)

		self.title_label['text'] = "QUANTITATIVE MODE"

	def calibration_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_calib_list)
		self.base_window.switch_page()
		self.base_window.quantitative_calib_list.load_program()
	def analysis_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_0)
		self.base_window.switch_page()
	def back_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
		self.base_window.switch_page()


class ConnectFrame(QualitativeOptionFrame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.title_label['text'] = Connect_Language["Title Label"][language]

		self.calibration_button['text'] = "Email"
		self.analysis_button['text'] = "Server"

		self.calibration_button['command'] = self.email_clicked
		self.analysis_button['command'] = self.server_clicked

		self.back_button['text'] = Connect_Language["Back Button"][language]


	def email_clicked(self):
		self.base_window.forget_page()
		self.base_window.email_setting.check_status()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.email_setting)
		self.base_window.switch_page()
	def server_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.server_setting)
		self.base_window.switch_page()


class LanguageFrame(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.work_frame.pack_propagate(0)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X, expand=TRUE)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = "LANGUAGE",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		# In work frame
		self.language_list = StringVar()
		self.language_combobox = ttk.Combobox(self.work_frame,
												state = "readonly",
												width = 8,
												font = ("Courier", 13),
												textvariable = self.language_list)
		self.language_combobox['values'] = ('English',
											'Tiếng Việt')
		self.language_combobox.current(language)

		# ~ self.language_combobox.pack(padx=20, pady=50, ipadx=20, ipady=20, side=LEFT)
		self.language_combobox.grid(row=0, column=0, padx=20, pady=20)

		self.confirm_button = Button(self.work_frame,
									text = "Confirm",
									font = MAIN_FUCNTION_BUTTON_FONT,
									bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
									fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
									borderwidth = 0,)
		# ~ self.confirm_button.pack(ipadx=30, ipady=20, padx=80, pady=100, side=LEFT)
		self.confirm_button.grid(row=1, column=0, padx=20, pady=20)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = "Back",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command=self.back_clicked)
		self.back_button.pack(side=LEFT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

		
	def back_clicked(self):
		self.cancel_clicked()
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
		self.base_window.switch_page()

class SettingFrame(Frame):
	def __init__(self, container):
		super().__init__(container)

		self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
		self.base_window = container

		self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.work_frame.pack_propagate(0)
		self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X, expand=TRUE)

		# In title frame
		self.title_label = Label(self.title_frame,
								text = QuantitativeKit_Language["Title Label"][language],
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE)

		#In work frame
		self.program_frame = ScrollableFrame(self.work_frame)
		self.program_frame.grid(row=0, column=1, pady=63, rowspan=4)

		self.file_name_frame = Frame(self.work_frame,
								bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.file_name_frame.grid(row=0, column=0, ipadx=10, ipady=0, padx=0, pady=0)

		self.info_labelframe = LabelFrame(self.work_frame,
								font = LABELFRAME_TXT_FONT,
								bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.info_labelframe.grid(row=1, column=0, ipadx=10, ipady=5, padx=0, pady=0)

		self.base_value_labelframe = LabelFrame(self.work_frame,
								font = LABELFRAME_TXT_FONT,
								bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.base_value_labelframe.grid(row=2, column=0, ipadx=28, ipady=5, padx=0, pady=0)

		self.control_frame = Frame(self.work_frame,
								bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.control_frame.grid(row=3, column=0, ipadx=0, ipady=0, padx=0, pady=0)
		
		# In file_name_frame
		self.experiment_name_label = Label(self.file_name_frame,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							text = QuantitativeKit_Language["KitName Label"][language],
							fg = LABEL_TXT_COLOR,
							font = ('Helvetica', 10),
							anchor = 'e')
		self.experiment_name_label.grid(row=0, column=0, padx=17, sticky=E)

		self.experiment_name_text = Text(self.file_name_frame,
							width = 22,
							height = 1,
							bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
							fg = LABEL_TXT_COLOR,
							font = ('Helvetica', 12))
		self.experiment_name_text.grid(row=0, column=1, padx=2, pady=10)

		# In info_labelframe
		self.concentration_label = Label(self.info_labelframe,
									bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
									text = QuantitativeKit_Language["Concentration Label"][language],
									font = LABEL_TXT_FONT,
									fg = LABEL_TXT_COLOR)
		self.concentration_label.grid(row=0, column=1, padx=5, pady=5)

		self.value_label = Label(self.info_labelframe,
									bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
									text = QuantitativeKit_Language["Value Label"][language],
									font = LABEL_TXT_FONT,
									fg = LABEL_TXT_COLOR)
		self.value_label.grid(row=0, column=2, padx=5, pady=5)

		numeric_label_list = list(range(0,5))
		self.concentration_entry_list = list(range(0,5))
		self.value_entry_list = list(range(0,5))
		for i in range(0,5):
			numeric_label_list[i] = Label(self.info_labelframe,
									bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
									text = str(i+1),
									font = LABEL_TXT_FONT,
									fg = LABEL_TXT_COLOR)
			numeric_label_list[i].grid(row=i+1, column=0, padx=10, pady=10)

			self.concentration_entry_list[i] = Entry(self.info_labelframe,
												width=10, 
												font=('Courier',14))
			self.concentration_entry_list[i].grid(row=i+1, column=1)

			self.value_entry_list[i] = Entry(self.info_labelframe,
												width=10, 
												font=('Courier',14))
			self.value_entry_list[i].grid(row=i+1, column=2)

		# In base_value_labelframe
		shift_label = Label(self.base_value_labelframe,
									bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
									text = "       ",
									font = LABEL_TXT_FONT,
									fg = LABEL_TXT_COLOR)
		shift_label.grid(row=0, column=0, padx=5, pady=5)

		self.base_value_label = Label(self.base_value_labelframe,
									bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
									text = QuantitativeKit_Language["NValue Label"][language],
									font = LABEL_TXT_FONT,
									fg = LABEL_TXT_COLOR)
		self.base_value_label.grid(row=0, column=1, padx=5, pady=5)

		self.base_value_entry = Entry(self.base_value_labelframe,
								width=10, 
								font=('Courier',14))
		self.base_value_entry.grid(row=0, column=2, padx=5, pady=5)
								
		
		# In control_frame
		self.save_button = Button(self.control_frame,
								text = QuantitativeKit_Language["Save Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.save_clicked)
		self.save_button.pack(side=LEFT, padx=2, pady=10, ipadx=18, ipady=5)

		self.delete_button = Button(self.control_frame,
								text = QuantitativeKit_Language["Delete Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.delete_clicked)
		self.delete_button.pack(side=LEFT, padx=2, pady=10, ipadx=18, ipady=5)

		self.clear_button = Button(self.control_frame,
								text = QuantitativeKit_Language["Clear Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.clear_clicked)
		self.clear_button.pack(side=LEFT, padx=2, pady=10, ipadx=19, ipady=5)

		# In button frame
		self.back_button = Button(self.button_frame,
								text = QuantitativeKit_Language["Back Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.back_clicked)
		self.back_button.pack(side=LEFT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

		self.next_button = Button(self.button_frame,
								text = QuantitativeKit_Language["Analysis Button"][language],
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.next_clicked)
		self.next_button.pack(side=RIGHT, padx=0, pady=0, ipady=10, ipadx=30, anchor=E)
	
	def save_clicked(self):
		program_name = self.experiment_name_text.get("1.0","end-1c")
		if(len(program_name) != 0):
			count = 0
			for i in range(0,5):
				if(self.concentration_entry_list[i].get() != "" and self.value_entry_list[i] != ""):
					count += 1
			if(count >= 3):
				if(self.base_value_entry.get() != ""):
					if os.path.exists(programs_quantitative_path + program_name + '.xlsx'):
						msg = messagebox.askquestion(QuantitativeKit_Language["FileExists Error"][language], QuantitativeKit_Language["FileExists Ask"][language])
						if(msg == 'yes'):
							wb = Workbook()
							sheet = wb.active

							sheet["A1"] = self.base_value_entry.get()
							sheet["B2"] = self.concentration_entry_list[0].get()
							sheet["C2"] = self.value_entry_list[0].get()
							sheet["B3"] = self.concentration_entry_list[1].get()
							sheet["C3"] = self.value_entry_list[1].get()
							sheet["B4"] = self.concentration_entry_list[2].get()
							sheet["C4"] = self.value_entry_list[2].get()
							sheet["B5"] = self.concentration_entry_list[3].get()
							sheet["C5"] = self.value_entry_list[3].get()
							sheet["B6"] = self.concentration_entry_list[4].get()
							sheet["C6"] = self.value_entry_list[4].get()
							
							# shutil.rmtree(programs_quantitative_path + program_name + '.xlsx')
							wb.save(programs_quantitative_path + program_name + '.xlsx')
							wb.close()

							messagebox.showinfo("",QuantitativeKit_Language["SaveFile Success"][language])
					else:
						wb = Workbook()
						sheet = wb.active

						sheet["A1"] = self.base_value_entry.get()
						sheet["B2"] = self.concentration_entry_list[0].get()
						sheet["C2"] = self.value_entry_list[0].get()
						sheet["B3"] = self.concentration_entry_list[1].get()
						sheet["C3"] = self.value_entry_list[1].get()
						sheet["B4"] = self.concentration_entry_list[2].get()
						sheet["C4"] = self.value_entry_list[2].get()
						sheet["B5"] = self.concentration_entry_list[3].get()
						sheet["C5"] = self.value_entry_list[3].get()
						sheet["B6"] = self.concentration_entry_list[4].get()
						sheet["C6"] = self.value_entry_list[4].get()

						wb.save(programs_quantitative_path + program_name + '.xlsx')
						wb.close()

						messagebox.showinfo("",QuantitativeKit_Language["SaveFile Success"][language])

						self.experiment_name_text.delete("1.0","end")
						self.base_value_entry.delete(0, END)
						for i in range(0,5):
							self.concentration_entry_list[i].delete(0, END)
							self.value_entry_list[i].delete(0, END)
						self.load_program()
				else:
					messagebox.showwarning("", QuantitativeKit_Language["NValue Empty"][language])
			else:
				messagebox.showwarning("", QuantitativeKit_Language["Concentration Empty"][language])
		else:
			messagebox.showwarning("", QuantitativeKit_Language["Kit Empty"][language])
			
	def delete_clicked(self):
		program_name = self.experiment_name_text.get("1.0","end-1c")
		if os.path.exists(programs_quantitative_path + program_name + '.xlsx'):
			msg = messagebox.askquestion("", QuantitativeKit_Language["Delete Confirm"][language])
			if(msg == 'yes'):
				os.remove(programs_quantitative_path + program_name + '.xlsx')

				self.experiment_name_text.delete("1.0","end")
				self.base_value_entry.delete(0, END)
				for i in range(0,5):
					self.concentration_entry_list[i].delete(0, END)
					self.value_entry_list[i].delete(0, END)
				self.load_program()
				messagebox.showinfo("", QuantitativeKit_Language["Delete Done"][language])
		else:
			messagebox.showerror("",  QuantitativeKit_Language["FileNotExists Error"][language])

	def clear_clicked(self):
		self.experiment_name_text.delete("1.0","end")
		self.base_value_entry.delete(0, END)
		for i in range(0,5):
			self.concentration_entry_list[i].delete(0, END)
			self.value_entry_list[i].delete(0, END)

	def back_clicked(self):
		self.clear_clicked()
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
		self.base_window.switch_page()

	def next_clicked(self):
		self.clear_clicked()
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_0)
		self.base_window.switch_page()

	def load_program(self):
		try:
			for i in range(len(self.program_button)):
				self.program_button[i].destroy()
		except:
			pass
		self.program_button = list(range(100))
		for file in os.listdir(programs_quantitative_path):
			self.program_button[os.listdir(programs_quantitative_path).index(file)] = Button(self.program_frame.scrollable_frame,
									text=file[:(len(file)-5)],
									font = PROGRAM_BUTTON_TXT_FONT,
									bg = PROGRAM_BUTTON_BGD_COLOR,
									fg = PROGRAM_BUTTON_TXT_COLOR,
									width = 51,
									borderwidth = 0)
			self.program_button[os.listdir(programs_quantitative_path).index(file)]['command'] = partial(self.program_clicked, os.listdir(programs_quantitative_path).index(file))
			self.program_button[os.listdir(programs_quantitative_path).index(file)].pack(pady=2, ipady=5, fill=BOTH, expand=TRUE)

	def program_clicked(self, button_index):
		wb = load_workbook(programs_quantitative_path + self.program_button[button_index]['text'] + ".xlsx")
		sheet = wb.active

		# ~ self.n_base_value = float(sheet["I2"].value)
		# ~ self.a_value = float(sheet["E2"].value)
		# ~ self.b_value = float(sheetn["G2"].value)
		
		self.n_base_value = float(sheet["A1"].value)
		
		concen1_enalble = 1
		concen2_enalble = 1
		concen3_enalble = 1
		concen4_enalble = 1
		concen5_enalble = 1
		try:
			self.value1 = float(sheet["C2"].value)
			self.concen1 = float(sheet["B2"].value)
			concen_1_pt = [self.concen1, self.value1]
		except:
			concen1_enalble = 0
			pass
			
		try:
			self.value2 = float(sheet["C3"].value)
			self.concen2 = float(sheet["B3"].value)
			concen_2_pt = [self.concen2, self.value2]
		except:
			concen2_enalble = 0
			pass
			
		try:
			self.value3 = float(sheet["C4"].value)
			self.concen3 = float(sheet["B4"].value)
			concen_3_pt = [self.concen3, self.value3]
		except:
			concen3_enalble = 0
			pass
			
		try:
			self.value4 = float(sheet["C5"].value)
			self.concen4 = float(sheet["B5"].value)
			concen_4_pt = [self.concen4, self.value4]
		except:
			concen4_enalble = 0
			pass
			
		try:
			self.value5 = float(sheet["C6"].value)
			self.concen5 = float(sheet["B6"].value)
			concen_5_pt = [self.concen5, self.value5]
		except:
			concen5_enalble = 0
			pass
		
		self.experiment_name_text.delete('1.0',END)
		self.experiment_name_text.insert('1.0', self.program_button[button_index]['text'])

		self.base_value_entry.delete(0, END)
		self.base_value_entry.insert(0, self.n_base_value)

		self.concentration_entry_list[0].delete(0,END)
		self.value_entry_list[0].delete(0,END)
		self.concentration_entry_list[1].delete(0,END)
		self.value_entry_list[1].delete(0,END)
		self.concentration_entry_list[2].delete(0,END)
		self.value_entry_list[2].delete(0,END)
		self.concentration_entry_list[3].delete(0,END)
		self.value_entry_list[3].delete(0,END)
		self.concentration_entry_list[4].delete(0,END)
		self.value_entry_list[4].delete(0,END)

		if(concen1_enalble):
			self.concentration_entry_list[0].insert(0, self.concen1)
			self.value_entry_list[0].insert(0, self.value1)
		if(concen2_enalble):
			self.concentration_entry_list[1].insert(0, self.concen2)
			self.value_entry_list[1].insert(0, self.value2)
		if(concen3_enalble):
			self.concentration_entry_list[2].insert(0, self.concen3)
			self.value_entry_list[2].insert(0, self.value3)
		if(concen4_enalble):
			self.concentration_entry_list[3].insert(0, self.concen4)
			self.value_entry_list[3].insert(0, self.value4)
		if(concen5_enalble):
			self.concentration_entry_list[4].insert(0, self.concen5)
			self.value_entry_list[4].insert(0, self.value5)

		wb.close()

class TrialExpiredFrame(Frame):
	def __init__(self, master):
		super().__init__(master)
		self.base_window = master

		# ~ self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
		# ~ self.title_frame.pack(ipadx=0, ipady=5, fill=X)
		self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.work_frame.pack_propagate(0)
		# ~ self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		# ~ self.button_frame.pack(fill=X, expand=TRUE)
		
		# In title frame
		# ~ self.title_label = Label(self.title_frame,
								# ~ text = "...",
								# ~ font = TITLE_TXT_FONT,
								# ~ bg = TITILE_FRAME_BGD_COLOR,
								# ~ fg = TITILE_FRAME_TXT_COLOR)
		# ~ self.title_label.pack(padx=0, pady=0, ipady=10, ipadx=30)
		
		# In work frame
		print("self.base_window.trial_days: ", self.base_window.trial_days)
		self.expire_info1_label = Label(self.work_frame,
								text = "Your " + str(self.base_window.trial_days) + "-day trial has expired",
								font = ('Courier',15),
								bg = LABEL_FRAME_BGD_COLOR,
								fg = 'red')
		self.expire_info1_label.grid(row=0, column=0, pady=30, sticky=EW)
		
		self.expire_info2_label = Label(self.work_frame,
								text = " Please enter the activation code to continue using the application",
								font = TITLE_TXT_FONT,
								bg = LABEL_FRAME_BGD_COLOR,
								fg = 'grey35')
		self.expire_info2_label.grid(row=2, column=0, pady=10, padx=30, sticky=W)
		
		self.active_code_entry = Entry(self.work_frame, width=30, font=('Courier',14))
		self.active_code_entry.grid(row=3, column=0, pady=10, padx=30, sticky=EW)
		
		self.activate_button = Button(self.work_frame,
								text = "Activate",
								font = SWITCH_PAGE_BUTTON_FONT,
								# width = SWITCH_PAGE_BUTTON_WIDTH,
								# height = SWITCH_PAGE_BUTTON_HEIGHT,
								bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
								fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
								borderwidth = 0,
								command = self.activate_clicked)
		self.activate_button.grid(row=4, column=0, ipady=10, pady=30, padx=150, sticky=EW)
	
	def activate_clicked(self):
		self.active_code_enter = self.active_code_entry.get()
		if(self.active_code_enter != ''):
			if(self.active_code_enter == trial_30days_extend_code):
				if(active_code != trial_30days_extend_code):
					fw = open(working_dir + "/active_code.txt",'w')	
					fw.writelines(self.active_code_enter + '\n')
					messagebox.showinfo("","Your trial package has been extended to 30 days.")
					self.base_window.forget_page()
					self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
					self.base_window.switch_page()
					self.base_window.system_check_light()
				else:
					messagebox.showerror("","Your code is invalid, please try again.")
			elif(self.active_code_enter == trial_full_active_code):
				fw = open(working_dir + "/active_code.txt",'w')	
				fw.writelines(self.active_code_enter + '\n')
				messagebox.showinfo("","Successful activation.")
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
				self.base_window.switch_page()
				self.base_window.system_check_light()
			else:
				messagebox.showerror("","Your code is invalid, please try again.")
		else:
			messagebox.showwarning("","Please enter activation code.")


class QuantitativeProgramsList(QuantitativeAnalysisFrame1):
	def __init__(self, container):
		super().__init__(container)
		self.title_label['text'] = QuantitativeProgramList_Language["Title Label"][language]
		self.next_button['text'] = QuantitativeProgramList_Language["Choose Button"][language]
		self.info_labelframe['text'] = QuantitativeProgramList_Language["Information LabelFrame"][language]
		self.experiment_name_label['text'] = QuantitativeProgramList_Language["KitName Label"][language]
		self.status_label['text'] = QuantitativeProgramList_Language["Parameters Label"][language]
		self.back_button['text'] = QuantitativeProgramList_Language["Back Button"][language]
		

	def back_clicked(self):
		if(self.base_window.qualitative_analysis_3.quantitative_view == 0):
			self.base_window.qualitative_analysis_3.thr_value.set(self.base_window.main_menu.threshold_value)
			self.base_window.update_frame()
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_3)
		self.base_window.switch_page()

	def next_clicked(self):
		self.program_name = self.experiment_name_text.get("1.0","end-1c")
		if(len(self.program_name) != 0):
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_3)
			self.base_window.switch_page()
			self.base_window.qualitative_analysis_3.quantitative_view = 1
			self.base_window.update_frame()
			# self.base_window.qualitative_analysis_3.update_result()
		else:
			messagebox.showwarning("",QuantitativeProgramList_Language["Kit Empty"][language])

	def reset(self):
		self.base_window.qualitative_analysis_3.quantitative_view = 0
		self.experiment_name_text['state'] = "normal"
		self.experiment_name_text.delete('1.0',END)
		self.experiment_name_text['state'] = "disabled"
		self.status_text['state'] = "normal"
		self.status_text.delete('1.0',END)
		self.status_text['state'] = "disabled"


class MainMenu(Frame):
	def __init__(self, master):
		super().__init__(master)
		self.base_window = master

		# Base frame create
		self.base_frame = Frame(self,bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.base_frame.pack(fill=BOTH, expand=TRUE)

		self.title_frame = Frame(self.base_frame, bg = TITILE_FRAME_BGD_COLOR)
		self.title_frame.pack(ipadx=0, ipady=20, fill=X)
		self.work_frame = Frame(self.base_frame, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
		self.work_frame.pack(expand=TRUE)
		self.button_frame = Frame(self.base_frame, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
		self.button_frame.pack(fill=X)

		self.title_label = Label(self.title_frame,
								text = "SPOTCHECK",
								font = TITLE_TXT_FONT,
								bg = TITILE_FRAME_BGD_COLOR,
								fg = TITILE_FRAME_TXT_COLOR)
		self.title_label.pack(expand=TRUE, ipady=5)
		
		self.language_list = StringVar()
		self.language_combobox = ttk.Combobox(self.title_frame,
												state = "readonly",
												width = 10,
												height = 30,
												font = ("Arial", 11),
												textvariable = self.language_list)
		self.language_combobox['values'] = ('English',
											'Tiếng Việt')
		self.language_combobox.current(language)
		self.language_combobox.place(x=700, y=1)
		self.language_combobox.bind("<<ComboboxSelected>>", self.language_check_option)
		
	
		self.screening_button = Button(self.work_frame,
									text = MainScreen_Language["Screening Button"][language],
									# ~ text = "ANALYSIS",
									font = MAIN_MENU_BUTTON_FONT,
									width = MAIN_MENU_BUTTON_WIDTH,
									height = MAIN_MENU_BUTTON_HEIGHT,
									bg = MAIN_MENU_BUTTON_BGD_COLOR,
									fg = MAIN_MENU_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.screening_clicked)
		self.screening_button.grid(row=0, column=0, ipadx=20, ipady=10, padx=100, pady=130)

		self.quantitative_button = Button(self.work_frame,
									text = MainScreen_Language["Quantitative Button"][language],
									font = MAIN_MENU_BUTTON_FONT,
									width = MAIN_MENU_BUTTON_WIDTH,
									height = MAIN_MENU_BUTTON_HEIGHT,
									bg = MAIN_MENU_BUTTON_BGD_COLOR,
									fg = MAIN_MENU_BUTTON_TXT_COLOR,
									borderwidth = 0,
									command = self.quantitative_clicked)
		self.quantitative_button.grid(row=0, column=1, ipadx=20, ipady=10, padx=100, pady=130)

		self.view_result_button = Button(self.button_frame,
									text = MainScreen_Language["ViewResult Button"][language],
									font = MAIN_MENU_BUTTON_FONT,
									bg = MAIN_MENU_BUTTON_BGD_COLOR,
									fg = MAIN_MENU_BUTTON_TXT_COLOR,
									# ~ width = 16,
									# ~ height = 2,
									borderwidth = 0,
									command = self.view_result_clicked)
		self.view_result_button.pack(side=LEFT, fill=BOTH, expand=TRUE, ipady=5)

		self.create_file_button = Button(self.button_frame,
									text = MainScreen_Language["CreateFile Button"][language],
									font = MAIN_MENU_BUTTON_FONT,
									bg = MAIN_MENU_BUTTON_BGD_COLOR,
									fg = MAIN_MENU_BUTTON_TXT_COLOR,
									# ~ width = 16,
									# ~ height = 2,
									borderwidth = 0,
									command = self.set_id_clicked)
		self.create_file_button.pack(side=LEFT, fill=BOTH, expand=TRUE)

		self.connect_button = Button(self.button_frame,
									text = MainScreen_Language["Setting Button 1"][language],
									font = MAIN_MENU_BUTTON_FONT,
									bg = MAIN_MENU_BUTTON_BGD_COLOR,
									fg = MAIN_MENU_BUTTON_TXT_COLOR,
									# ~ width = 16,
									# ~ height = 2,
									borderwidth = 0,
									command = self.connect_clicked)
		self.connect_button.pack(side=LEFT, fill=BOTH, expand=TRUE)
		
		# self.language_button = Button(self.button_frame,
		# 							text = "Language",
		# 							font = MAIN_MENU_BUTTON_FONT,
		# 							bg = MAIN_MENU_BUTTON_BGD_COLOR,
		# 							fg = MAIN_MENU_BUTTON_TXT_COLOR,
		# 							borderwidth = 0,
		# 							command = self.language_clicked)
		# self.language_button.pack(side=LEFT, fill=BOTH, expand=TRUE)

		self.exit_button = Button(self.button_frame,
									text = MainScreen_Language["Exit Button"][language],
									font = MAIN_MENU_BUTTON_FONT,
									bg = MAIN_MENU_BUTTON_BGD_COLOR,
									fg = MAIN_MENU_BUTTON_TXT_COLOR,
									# ~ width = 16,
									# ~ height = 2,
									borderwidth = 0,
									command = self.exit_clicked)
		self.exit_button.pack(side=LEFT, fill=BOTH, expand=TRUE)

	def language_check_option(self, eventObject):
		global language
		if(self.language_combobox.current() == 0): 
			language = 0
			fw = open(working_dir + "/language.txt",'w')
			fw.writelines(["0\n"])
			fw.close()
		else:
			language = 1
			fw = open(working_dir + "/language.txt",'w')
			fw.writelines(["1\n"])
			fw.close()
		### Main Sceen ###
		self.screening_button['text'] = MainScreen_Language['Screening Button'][language]
		self.quantitative_button['text'] = MainScreen_Language['Quantitative Button'][language]
		self.view_result_button['text'] = MainScreen_Language['ViewResult Button'][language]
		self.create_file_button['text'] = MainScreen_Language['CreateFile Button'][language]
		self.connect_button['text'] = MainScreen_Language['Setting Button 1'][language]
		self.exit_button['text'] = MainScreen_Language['Exit Button'][language]

		try:
			self.thr1_button['text'] = MainScreen_Language['Environment Button'][language]
			self.thr2_button['text'] = MainScreen_Language['Host Button'][language]
			self.threshold_label_frame['text'] = MainScreen_Language['Screening LabelFrame'][language]
		except:
			pass

		try:
			self.setting_button['text'] = MainScreen_Language['Setting Button 2'][language]
			self.analysis_button['text'] = MainScreen_Language['Analysis Button'][language]
			self.threshold_label_frame['text'] = MainScreen_Language['Quantitative LabelFrame'][language]
		except:
			pass

		### View Result ###
		self.base_window.view_results.title_label['text'] = ViewResult_Language['Title Label'][language]
		self.base_window.view_results.back_button['text'] = ViewResult_Language['Back Button'][language]
		self.base_window.view_results.open_button['text'] = ViewResult_Language['Open Button'][language]
		self.base_window.view_results.infor_frame['text'] = ViewResult_Language['Information LabelFrame'][language]

		### Create File ###
		self.base_window.id_create.title_label['text'] = CreateFile_Language["Title Label"][language]
		self.base_window.id_create.property_labelframe['text'] = CreateFile_Language["SampleProperties LabelFrame"][language]
		self.base_window.id_create.quick_create_labelframe['text'] = CreateFile_Language["QuickSetup LabelFrame"][language]
		self.base_window.id_create.first_well_button['text'] = CreateFile_Language["FirstWell Button"][language]
		self.base_window.id_create.last_well_button['text'] = CreateFile_Language["LastWell Button"][language]
		self.base_window.id_create.quick_create_button['text'] = CreateFile_Language["Set Button"][language]
		self.base_window.id_create.back_button['text'] = CreateFile_Language["Back Button"][language]
		self.base_window.id_create.create_button['text'] = CreateFile_Language["Create Button"][language]
		self.base_window.id_create.load_button['text'] = CreateFile_Language["Load Button"][language]

		try:
			self.base_window.id_create.sample_name_label['text'] = CreateFile_Language["SampleName Label"][language]
			self.base_window.id_create.ok_button['text'] = CreateFile_Language["OK Button"][language]
		except:
			pass

		### Setting ###
		self.base_window.connect.back_button['text'] = Connect_Language["Back Button"][language]
		self.base_window.connect.title_label['text'] = Connect_Language["Title Label"][language]
		
		### Screening 0 ###
		self.base_window.qualitative_analysis_0.title_label['text'] = Screening0_Language["Title Label"][language]
		self.base_window.qualitative_analysis_0.experiment_name_label['text'] = Screening0_Language["ExperimentName Label"][language]
		self.base_window.qualitative_analysis_0.user_name_label['text'] = Screening0_Language["TechnicianName Label"][language]
		self.base_window.qualitative_analysis_0.template_name_label['text'] = Screening0_Language["TemplateName Label"][language]
		self.base_window.qualitative_analysis_0.back_button['text'] = Screening0_Language["Back Button"][language]
		self.base_window.qualitative_analysis_0.next_button['text'] = Screening0_Language["Next Button"][language]

		### Screening 2 ###
		self.base_window.qualitative_analysis_2.title_label['text'] = Screening2_Language["Title Label"][language]
		self.base_window.qualitative_analysis_2.id_load_frame['text'] = Screening2_Language["SamplesFile LabelFrame"][language]
		self.base_window.qualitative_analysis_2.load_button['text'] = Screening2_Language["Load Button"][language]
		self.base_window.qualitative_analysis_2.create_button['text'] = Screening2_Language["Create Button"][language]
		self.base_window.qualitative_analysis_2.back_button['text'] = Screening2_Language["Back Button"][language]
		self.base_window.qualitative_analysis_2.next_button['text'] = Screening2_Language["Next Button"][language]
		# self.base_window.qualitative_analysis_2.ok_button['text'] = Screening2_Language["OK Button"][language]
		# self.base_window.qualitative_analysis_2.cancel_button['text'] = Screening2_Language["Cancel Button"][language]
		
		### Screening 3 ###
		self.base_window.qualitative_analysis_3.title_label['text'] = Screening3_Language["Title Label"][language]

		### Quantitative Program List ###
		self.base_window.quantitative_programs_list.title_label['text'] = QuantitativeProgramList_Language["Title Label"][language]
		self.base_window.quantitative_programs_list.next_button['text'] = QuantitativeProgramList_Language["Choose Button"][language]
		self.base_window.quantitative_programs_list.info_labelframe['text'] = QuantitativeProgramList_Language["Information LabelFrame"][language]
		self.base_window.quantitative_programs_list.experiment_name_label['text'] = QuantitativeProgramList_Language["KitName Label"][language]
		self.base_window.quantitative_programs_list.status_label['text'] = QuantitativeProgramList_Language["Parameters Label"][language]
		self.base_window.quantitative_programs_list.back_button['text'] = QuantitativeProgramList_Language["Back Button"][language]

		### Quantitative 0 ###
		self.base_window.quantitative_analysis_0.experiment_name_label['text'] = Quantitative0_Language['ExperimentName Label'][language]
		self.base_window.quantitative_analysis_0.user_name_label['text'] = Quantitative0_Language['TechnicianName Label'][language]
		self.base_window.quantitative_analysis_0.template_name_label['text'] = Quantitative0_Language['TemplateName Label'][language]
		self.base_window.quantitative_analysis_0.back_button['text'] = Quantitative0_Language['Back Button'][language]
		self.base_window.quantitative_analysis_0.next_button['text'] = Quantitative0_Language['Next Button'][language]

		### Quantitative 1 ###
		self.base_window.quantitative_analysis_1.title_label['text'] = Quantitative1_Language["Title Label"][language]
		self.base_window.quantitative_analysis_1.next_button['text'] = Quantitative1_Language["Next Button"][language]
		self.base_window.quantitative_analysis_1.info_labelframe['text'] = Quantitative1_Language["Information LabelFrame"][language]
		self.base_window.quantitative_analysis_1.experiment_name_label['text'] = Quantitative1_Language["KitName Label"][language]
		self.base_window.quantitative_analysis_1.status_label['text'] = Quantitative1_Language["Parameters Label"][language]
		self.base_window.quantitative_analysis_1.back_button['text'] = Quantitative1_Language["Back Button"][language]

		### Quantitative 2 ###
		self.base_window.quantitative_analysis_2.title_label['text'] = Quantitative2_Language["Title Label"][language]
		self.base_window.quantitative_analysis_2.id_load_frame['text'] = Quantitative2_Language["SamplesFile LabelFrame"][language]
		self.base_window.quantitative_analysis_2.load_button['text'] = Quantitative2_Language["Load Button"][language]
		self.base_window.quantitative_analysis_2.create_button['text'] = Quantitative2_Language["Create Button"][language]
		self.base_window.quantitative_analysis_2.back_button['text'] = Quantitative2_Language["Back Button"][language]
		self.base_window.quantitative_analysis_2.next_button['text'] = Quantitative2_Language["Next Button"][language]
		
		### Quantitative 3 ###
		self.base_window.quantitative_analysis_3.title_label['text'] = Quantitative3_Language["Title Label"][language]
		
		### Quantitative Kit ###
		self.base_window.setting.title_label['text'] = QuantitativeKit_Language["Title Label"][language]
		self.base_window.setting.experiment_name_label['text'] = QuantitativeKit_Language["KitName Label"][language]
		self.base_window.setting.concentration_label['text'] = QuantitativeKit_Language["Concentration Label"][language]
		self.base_window.setting.value_label['text'] = QuantitativeKit_Language["Value Label"][language]
		self.base_window.setting.base_value_label['text'] = QuantitativeKit_Language["NValue Label"][language]

		self.base_window.setting.save_button['text'] = QuantitativeKit_Language["Save Button"][language]
		self.base_window.setting.delete_button['text'] = QuantitativeKit_Language["Delete Button"][language]
		self.base_window.setting.clear_button['text'] = QuantitativeKit_Language["Clear Button"][language]
		self.base_window.setting.back_button['text'] = QuantitativeKit_Language["Back Button"][language]
		self.base_window.setting.next_button['text'] = QuantitativeKit_Language["Analysis Button"][language]

		### System Check ###
		self.base_window.system_check.title_label['text'] = SystemCheck_Language["Title Label"][language]

	def screening_clicked(self):		
		# self.threshold_label_frame = LabelFrame(self.work_frame,
		# 									# ~ width = 600,
		# 									# ~ height = 300,
		# 									text = MainScreen_Language['Screening LabelFrame'][language],
		# 									bg = 'grey70')
		# self.threshold_label_frame.place(x=35, y=95)

		# self.thr1_button = Button(self.threshold_label_frame,
		# 			text = MainScreen_Language['Environment Button'][language],
		# 			font = SWITCH_PAGE_BUTTON_FONT,
		# 			bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
		# 			fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
		# 			borderwidth = 0,
		# 			command = self.thr1_clicked)
		# self.thr1_button.pack(side=LEFT, padx=30, pady=50, ipady=10, ipadx=2)

		# self.thr2_button = Button(self.threshold_label_frame,
		# 			text = MainScreen_Language['Host Button'][language],
		# 			font = SWITCH_PAGE_BUTTON_FONT,
		# 			bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
		# 			fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
		# 			borderwidth = 0,
		# 			command = self.thr2_clicked)
		# self.thr2_button.pack(side=LEFT, padx=30, pady=50, ipady=10, ipadx=21)
		
		# self.cancel_button = Button(self.threshold_label_frame,
		# 			text = "X",
		# 			font = SWITCH_PAGE_BUTTON_FONT,
		# 			width = 1,
		# 			height = 1,
		# 			bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
		# 			fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
		# 			borderwidth = 0,
		# 			command = self.cancel_clicked)
		# self.cancel_button.place(x=297,y=-18)
		
		# self.quantitative_button['state'] = 'disabled'
		# self.view_result_button['state'] = 'disabled'
		# self.create_file_button['state'] = 'disabled'
		# self.connect_button['state'] = 'disabled'


		# self.language_button['state'] = 'disable'
	
		# ~ self.base_window.forget_page()
		# ~ #self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_option)
		# ~ self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
		# ~ self.base_window.switch_page()

		self.threshold_value = 0; # su dung gia tri thu 1 trong file multiplier de so sanh
		fr = open(working_dir + "/multiplier.txt","r")
		self.num1 = float(fr.readline())
		self.num2 = float(fr.readline())
		self.num3 = float(fr.readline())
		
		####### SYSTEM CHECK ########
		global last_checkDate, last_checkMonth, last_checkYear, last_checkHour, last_checkMinute, last_checkSecond, last_checkValue
		fr = open(working_dir + "/.system.txt","r")
		last_checkDay = int(fr.readline())
		last_checkMonth = int(fr.readline())
		last_checkYear = int(fr.readline())
		last_checkHour = int(fr.readline())
		last_checkMinute = int(fr.readline())
		last_checkSecond = int(fr.readline())
		last_checkValue = float(fr.readline())
		
		now = datetime.now()
		
		time1 = last_checkYear*31536000 + last_checkMonth*2419200 + last_checkDay*86400 + last_checkHour*3600 + last_checkMinute*60 + last_checkSecond
		time2 = now.year*31536000 + now.month*2419200 + now.day*86400 + now.hour*3600 + now.minute*60 + now.second         
		number_of_hours = round(abs((time2 - time1)/3600),1)
		
		print("number_of_hours: ", number_of_hours)
		
		if(number_of_hours >= 1):
			msg = messagebox.askquestion("","It's been a while since the last system check, would you like to check again ?")
			if(msg == 'yes'):
				self.base_window.system_check.mode_check = 1
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
				self.base_window.switch_page()
				self.base_window.update_idletasks()
				self.base_window.system_check.serial_handle()
			else:
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
				self.base_window.switch_page()
		else:
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
			self.base_window.switch_page()
			
		self.quantitative_button['state'] = 'normal'
		self.view_result_button['state'] = 'normal'
		self.create_file_button['state'] = 'normal'
		self.connect_button['state'] = 'normal'
		
	def thr1_clicked(self):
		self.threshold_label_frame.place_forget()
		
		self.threshold_value = 0; # su dung gia tri thu 1 trong file multiplier de so sanh
		fr = open(working_dir + "/multiplier.txt","r")
		self.num1 = float(fr.readline())
		self.num2 = float(fr.readline())
		self.num3 = float(fr.readline())
		
		####### SYSTEM CHECK ########
		global last_checkDate, last_checkMonth, last_checkYear, last_checkHour, last_checkMinute, last_checkSecond, last_checkValue
		fr = open(working_dir + "/.system.txt","r")
		last_checkDay = int(fr.readline())
		last_checkMonth = int(fr.readline())
		last_checkYear = int(fr.readline())
		last_checkHour = int(fr.readline())
		last_checkMinute = int(fr.readline())
		last_checkSecond = int(fr.readline())
		last_checkValue = float(fr.readline())
		
		now = datetime.now()
		
		time1 = last_checkYear*31536000 + last_checkMonth*2419200 + last_checkDay*86400 + last_checkHour*3600 + last_checkMinute*60 + last_checkSecond
		time2 = now.year*31536000 + now.month*2419200 + now.day*86400 + now.hour*3600 + now.minute*60 + now.second         
		number_of_hours = round(abs((time2 - time1)/3600),1)
		
		print("number_of_hours: ", number_of_hours)
		
		if(number_of_hours >= 1):
			msg = messagebox.askquestion("","It's been a while since the last system check, would you like to check again ?")
			if(msg == 'yes'):
				self.base_window.system_check.mode_check = 1
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
				self.base_window.switch_page()
				self.base_window.update_idletasks()
				self.base_window.system_check.serial_handle()
			else:
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
				self.base_window.switch_page()
		else:
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
			self.base_window.switch_page()
			
		self.quantitative_button['state'] = 'normal'
		self.view_result_button['state'] = 'normal'
		self.create_file_button['state'] = 'normal'
		self.connect_button['state'] = 'normal'
		# self.language_button['state'] = 'normal'
	
			
	def thr2_clicked(self):
		self.threshold_label_frame.place_forget()
		
		self.threshold_value = 1; # su dung gia tri thu 2 trong file multiplier de so sanh
		fr = open(working_dir + "/multiplier.txt","r")
		self.num1 = float(fr.readline())
		self.num2 = float(fr.readline())
		self.num3 = float(fr.readline())
		
		####### SYSTEM CHECK ########
		global last_checkDate, last_checkMonth, last_checkYear, last_checkHour, last_checkMinute, last_checkSecond, last_checkValue
		fr = open(working_dir + "/.system.txt","r")
		last_checkDay = int(fr.readline())
		last_checkMonth = int(fr.readline())
		last_checkYear = int(fr.readline())
		last_checkHour = int(fr.readline())
		last_checkMinute = int(fr.readline())
		last_checkSecond = int(fr.readline())
		last_checkValue = float(fr.readline())
		
		now = datetime.now()
		
		time1 = last_checkYear*31536000 + last_checkMonth*2419200 + last_checkDay*86400 + last_checkHour*3600 + last_checkMinute*60 + last_checkSecond
		time2 = now.year*31536000 + now.month*2419200 + now.day*86400 + now.hour*3600 + now.minute*60 + now.second  
		number_of_hours = round(abs((time2 - time1)/3600),1)
		
		print("number_of_hours: ", number_of_hours)
		
		if(number_of_hours >= 1):
			msg = messagebox.askquestion("",MainScreen_Language["SystemCheck Ask"][language])
			if(msg == 'yes'):
				self.base_window.system_check.mode_check = 1
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
				self.base_window.switch_page()
				self.base_window.update_idletasks()
				self.base_window.system_check.serial_handle()
			else:
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
				self.base_window.switch_page()
		else:
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
			self.base_window.switch_page()
		
		self.quantitative_button['state'] = 'normal'
		self.view_result_button['state'] = 'normal'
		self.create_file_button['state'] = 'normal'
		self.connect_button['state'] = 'normal'
		# self.language_button['state'] = 'normal'
	
		
	def cancel_clicked(self):
		try:
			self.threshold_label_frame.place_forget()
		except:
			pass
		
		self.quantitative_button['state'] = 'normal'
		self.view_result_button['state'] = 'normal'
		self.create_file_button['state'] = 'normal'
		self.connect_button['state'] = 'normal'
		# self.language_button['state'] = 'normal'
	
	
	def quantitative_clicked(self):
		self.quanti_labelframe = LabelFrame(self.work_frame,
											text = MainScreen_Language['Quantitative LabelFrame'][language],
											bg = 'grey70')
		self.quanti_labelframe.place(x=380, y=95)

		
		self.analysis_button = Button(self.quanti_labelframe,
					text = MainScreen_Language['Analysis Button'][language],
					font = SWITCH_PAGE_BUTTON_FONT,
					bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
					fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
					borderwidth = 0,
					command = self.analysis_clicked)
		self.analysis_button.pack(side=LEFT, padx=30, pady=50, ipady=10, ipadx=20)

		self.setting_button = Button(self.quanti_labelframe,
					text = MainScreen_Language['Setting Button 2'][language],
					font = SWITCH_PAGE_BUTTON_FONT,
					bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
					fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
					borderwidth = 0,
					command = self.setting_clicked)
		self.setting_button.pack(side=LEFT, padx=30, pady=50, ipady=10, ipadx=21)
		
		self.cancel2_button = Button(self.quanti_labelframe,
					text = "X",
					font = SWITCH_PAGE_BUTTON_FONT,
					width = 1,
					height = 1,
					bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
					fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
					borderwidth = 0,
					command = self.cancel2_clicked)
		self.cancel2_button.place(x=312,y=-18)
		
		self.screening_button['state'] = 'disabled'
		self.view_result_button['state'] = 'disabled'
		self.create_file_button['state'] = 'disabled'
		self.connect_button['state'] = 'disabled'
		# self.language_button['state'] = 'disable'
	


	def analysis_clicked(self):
		self.quanti_labelframe.place_forget()

		self.screening_button['state'] = 'normal'
		self.view_result_button['state'] = 'normal'
		self.create_file_button['state'] = 'normal'
		self.connect_button['state'] = 'normal'
		# self.language_button['state'] = 'normal'
	

		####### SYSTEM CHECK ########
		global last_checkDate, last_checkMonth, last_checkYear, last_checkHour, last_checkMinute, last_checkSecond, last_checkValue
		fr = open(working_dir + "/.system.txt","r")
		last_checkDay = int(fr.readline())
		last_checkMonth = int(fr.readline())
		last_checkYear = int(fr.readline())
		last_checkHour = int(fr.readline())
		last_checkMinute = int(fr.readline())
		last_checkSecond = int(fr.readline())
		last_checkValue = float(fr.readline())
		
		now = datetime.now()
		
		time1 = last_checkYear*31536000 + last_checkMonth*2419200 + last_checkDay*86400 + last_checkHour*3600 + last_checkMinute*60 + last_checkSecond
		time2 = now.year*31536000 + now.month*2419200 + now.day*86400 + now.hour*3600 + now.minute*60 + now.second  
		number_of_hours = round(abs((time2 - time1)/3600),1)
		
		print("number_of_hours: ", number_of_hours)
		
		if(number_of_hours >= 1):
			msg = messagebox.askquestion("",MainScreen_Language["SystemCheck Ask"][language])
			if(msg == 'yes'):
				self.base_window.system_check.mode_check = 2
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
				self.base_window.switch_page()
				self.base_window.update_idletasks()
				self.base_window.system_check.serial_handle()
			else:
				self.base_window.forget_page()
				self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_0)
				self.base_window.switch_page()
		else:
			self.base_window.forget_page()
			self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_0)
			self.base_window.switch_page()

	def setting_clicked(self):
		try:
			self.quanti_labelframe.place_forget()
		except:
			pass

		self.screening_button['state'] = 'normal'
		self.view_result_button['state'] = 'normal'
		self.create_file_button['state'] = 'normal'
		self.connect_button['state'] = 'normal'
		# self.language_button['state'] = 'normal'
	

		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.setting)
		self.base_window.setting.load_program()
		self.base_window.switch_page()

	def cancel2_clicked(self):
		try:
			self.quanti_labelframe.place_forget()
		except:
			pass
		
		self.screening_button['state'] = 'normal'
		self.view_result_button['state'] = 'normal'
		self.create_file_button['state'] = 'normal'
		self.connect_button['state'] = 'normal'
		# self.language_button['state'] = 'normal'
	

	def language_clicked(self):
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.language)
		self.base_window.switch_page()

	def view_result_clicked(self):
		try:
			self.threshold_label_frame.place_forget()
		except:
			pass
		try:
			self.quanti_labelframe.place_forget()
		except:
			pass
		
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.view_results)
		self.base_window.switch_page()

	def set_id_clicked(self):
		try:
			self.threshold_label_frame.place_forget()
		except:
			pass
		try:
			self.quanti_labelframe.place_forget()
		except:
			pass

		self.base_window.forget_page()
		self.base_window.id_create.first_well_index = 0
		self.base_window.id_create.last_well_index = 47
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.id_create)
		self.base_window.switch_page()

	def connect_clicked(self):
		try:
			self.threshold_label_frame.place_forget()
		except:
			pass 
		try:
			self.quanti_labelframe.place_forget()
		except:
			pass
			
		self.base_window.forget_page()
		self.base_window.page_num = self.base_window.frame_list.index(self.base_window.connect)
		self.base_window.switch_page() 

	def exit_clicked(self):
		msg = messagebox.askquestion("", MainScreen_Language["Exit Confirm"][language])
		if(msg == "yes"):
			os._exit(0)
			self.base_window.destroy()

	def reset(self):
		self.base_window.frame_list.remove(self.base_window.new_qualitative_1)
		self.base_window.frame_list.remove(self.base_window.new_qualitative_2)
		self.base_window.frame_list.remove(self.base_window.new_qualitative_3)
		# ~ self.base_window.frame_list.remove(self.base_window.system_check)
		self.base_window.frame_list.remove(self.base_window.qualitative_analysis_1)
		self.base_window.frame_list.remove(self.base_window.qualitative_analysis_2)
		self.base_window.frame_list.remove(self.base_window.qualitative_analysis_3)
		self.base_window.frame_list.remove(self.base_window.id_create)
		self.base_window.frame_list.remove(self.base_window.new_quantitative_1)
		self.base_window.frame_list.remove(self.base_window.new_quantitative_2)
		self.base_window.frame_list.remove(self.base_window.new_quantitative_3)
		self.base_window.frame_list.remove(self.base_window.quantitative_analysis_1)
		self.base_window.frame_list.remove(self.base_window.quantitative_analysis_2)
		self.base_window.frame_list.remove(self.base_window.quantitative_analysis_3)
		self.base_window.frame_list.remove(self.base_window.qualitative_calib_list)
		self.base_window.frame_list.remove(self.base_window.qualitative_analysis_0)
		self.base_window.frame_list.remove(self.base_window.quantitative_calib_list)
		self.base_window.frame_list.remove(self.base_window.quantitative_analysis_0)
		# self.base_window.frame_list.remove(self.base_window.quantitative_programs_list)

		del self.base_window.new_qualitative_1
		del self.base_window.new_qualitative_2
		del self.base_window.new_qualitative_3
		# ~ del self.base_window.system_check
		del self.base_window.qualitative_analysis_1
		del self.base_window.qualitative_analysis_2
		del self.base_window.qualitative_analysis_3
		del self.base_window.id_create
		del self.base_window.new_quantitative_1
		del self.base_window.new_quantitative_2
		del self.base_window.new_quantitative_3
		del self.base_window.quantitative_analysis_1
		del self.base_window.quantitative_analysis_2
		del self.base_window.quantitative_analysis_3
		del self.base_window.qualitative_calib_list
		del self.base_window.qualitative_analysis_0
		del self.base_window.quantitative_calib_list
		del self.base_window.quantitative_analysis_0
		# del self.base_window.quantitative_programs_list

		self.base_window.new_qualitative_1 = NewQualitativeFrame1(self.base_window)
		self.base_window.new_qualitative_2 = NewQualitativeFrame2(self.base_window)
		self.base_window.new_qualitative_3 = NewQualitativeFrame3(self.base_window)
		# ~ self.base_window.system_check = SystemCheckFrame(self.base_window)
		self.base_window.qualitative_analysis_1 = QualitativeAnalysisFrame1(self.base_window)
		self.base_window.qualitative_analysis_2 = QualitativeAnalysisFrame2(self.base_window)
		self.base_window.qualitative_analysis_3 = QualitativeAnalysisFrame3(self.base_window)
		self.base_window.id_create = IDCreateFrame(self.base_window)
		self.base_window.new_quantitative_1 = NewQuantitativeFrame1(self.base_window)
		self.base_window.new_quantitative_2 = NewQuantitativeFrame2(self.base_window)
		self.base_window.new_quantitative_3 = NewQuantitativeFrame3(self.base_window)
		self.base_window.quantitative_analysis_1 = QuantitativeAnalysisFrame1(self.base_window)
		self.base_window.quantitative_analysis_2 = QuantitativeAnalysisFrame2(self.base_window)
		self.base_window.quantitative_analysis_3 = QuantitativeAnalysisFrame3(self.base_window)
		self.base_window.qualitative_calib_list = QualitativeCalibListFrame(self.base_window)
		self.base_window.qualitative_analysis_0 = QualitativeAnalysisFrame0(self.base_window)
		self.base_window.quantitative_calib_list = QuantitativeCalibListFrame(self.base_window)
		self.base_window.quantitative_analysis_0 = QuantitativeAnalysisFrame0(self.base_window)
		# self.base_window.quantitative_programs_list = QuantitativeProgramsList(self.base_window)

		self.base_window.frame_list.append(self.base_window.new_qualitative_1)
		self.base_window.frame_list.append(self.base_window.new_qualitative_2)
		self.base_window.frame_list.append(self.base_window.new_qualitative_3)
		self.base_window.frame_list.append(self.base_window.system_check)
		self.base_window.frame_list.append(self.base_window.qualitative_analysis_1)
		self.base_window.frame_list.append(self.base_window.qualitative_analysis_2)
		self.base_window.frame_list.append(self.base_window.qualitative_analysis_3)
		self.base_window.frame_list.append(self.base_window.id_create)
		self.base_window.frame_list.append(self.base_window.new_quantitative_1)
		self.base_window.frame_list.append(self.base_window.new_quantitative_2)
		self.base_window.frame_list.append(self.base_window.new_quantitative_3)
		self.base_window.frame_list.append(self.base_window.quantitative_analysis_1)
		self.base_window.frame_list.append(self.base_window.quantitative_analysis_2)
		self.base_window.frame_list.append(self.base_window.quantitative_analysis_3)
		self.base_window.frame_list.append(self.base_window.qualitative_calib_list)
		self.base_window.frame_list.append(self.base_window.qualitative_analysis_0)
		self.base_window.frame_list.append(self.base_window.quantitative_calib_list)
		self.base_window.frame_list.append(self.base_window.quantitative_analysis_0)
		# self.base_window.frame_list.append(self.quantitative_programs_list)


class MainWindow(Tk):
	def __init__(self):
		Tk.__init__(self)
		self.title('Spotcheck')
		self.geometry('800x480')
		self.configure(background = APP_BGD_COLOR)
		self.resizable(FALSE, FALSE)
		self.attributes('-fullscreen', True)

		self.page_num = 0
		self.frame_list = []
		self.trial_days = 0

		self.main_menu = MainMenu(self)
		self.qualitative_option = QualitativeOptionFrame(self)
		self.quantitative_option = QuantitativeOptionFrame(self)
		self.new_program = NewProgramFrame(self)
		self.view_results = ViewResultFrame(self)
		self.new_qualitative_1 = NewQualitativeFrame1(self)
		self.new_qualitative_2 = NewQualitativeFrame2(self)
		self.new_qualitative_3 = NewQualitativeFrame3(self)
		self.qualitative_analysis_1 = QualitativeAnalysisFrame1(self)
		self.qualitative_analysis_2 = QualitativeAnalysisFrame2(self)
		self.qualitative_analysis_3 = QualitativeAnalysisFrame3(self)
		self.new_quantitative_1 = NewQuantitativeFrame1(self)
		self.new_quantitative_2 = NewQuantitativeFrame2(self)
		self.new_quantitative_3 = NewQuantitativeFrame3(self)
		self.quantitative_analysis_1 = QuantitativeAnalysisFrame1(self)
		self.quantitative_analysis_2 = QuantitativeAnalysisFrame2(self)
		self.quantitative_analysis_3 = QuantitativeAnalysisFrame3(self)
		self.system_check = SystemCheckFrame(self)
		self.id_create = IDCreateFrame(self)
		self.saved_program = SavedProgramFrame(self)
		self.qualitative_saved_program = QualitativeSavedFrame(self)
		self.quantitative_saved_program = QuantitativeSavedFrame(self)
		self.email_setting = EmailSettingFrame(self)
		self.server_setting = ServerSettingFrame(self)
		self.qualitative_calib_list = QualitativeCalibListFrame(self)
		self.qualitative_analysis_0 = QualitativeAnalysisFrame0(self)
		self.quantitative_calib_list = QuantitativeCalibListFrame(self)
		self.quantitative_analysis_0 = QuantitativeAnalysisFrame0(self)
		self.connect = ConnectFrame(self)
		self.setting = SettingFrame(self)
		self.language = LanguageFrame(self)
		self.trial_expried = TrialExpiredFrame(self)
		self.quantitative_programs_list = QuantitativeProgramsList(self)

		self.frame_list.append(self.main_menu)
		self.frame_list.append(self.qualitative_option)
		self.frame_list.append(self.quantitative_option)
		self.frame_list.append(self.new_program)
		self.frame_list.append(self.view_results)
		self.frame_list.append(self.new_qualitative_1)
		self.frame_list.append(self.new_qualitative_2)
		self.frame_list.append(self.new_qualitative_3)
		self.frame_list.append(self.qualitative_analysis_1)
		self.frame_list.append(self.qualitative_analysis_2)
		self.frame_list.append(self.qualitative_analysis_3)
		self.frame_list.append(self.new_quantitative_1)
		self.frame_list.append(self.new_quantitative_2)
		self.frame_list.append(self.new_quantitative_3)
		self.frame_list.append(self.quantitative_analysis_1)
		self.frame_list.append(self.quantitative_analysis_2)
		self.frame_list.append(self.quantitative_analysis_3)
		self.frame_list.append(self.system_check)
		self.frame_list.append(self.id_create)
		self.frame_list.append(self.saved_program)
		self.frame_list.append(self.qualitative_saved_program)
		self.frame_list.append(self.quantitative_saved_program)
		self.frame_list.append(self.email_setting)
		self.frame_list.append(self.server_setting)
		self.frame_list.append(self.qualitative_calib_list)
		self.frame_list.append(self.qualitative_analysis_0)
		self.frame_list.append(self.quantitative_calib_list)
		self.frame_list.append(self.quantitative_analysis_0)
		self.frame_list.append(self.connect)
		self.frame_list.append(self.setting) 
		self.frame_list.append(self.language)
		self.frame_list.append(self.trial_expried)
		self.frame_list.append(self.quantitative_programs_list)

		self.switch_page()

		# Check trial period
		if(active_code == trial_full_active_code):
			self.system_check_light()
		elif(active_code == trial_30days_extend_code):
			self.trial_30days_extend()
		else:
			self.trial_7days()
			
		global number_of_instance
		number_of_instance = 0
		fw_instance = open('/home/pi/Spotcheck/.instance.txt', 'w')
		fw_instance.writelines('0\n')
		
			
	def forget_page(self):
		self.frame_list[self.page_num].forget()
	def switch_page(self):
		self.frame_list[self.page_num].tkraise()
		self.frame_list[self.page_num].pack(expand=TRUE, fill=BOTH)
	def update_frame(self):
		self.qualitative_analysis_3.update_result()
		self.qualitative_analysis_3.update()
		self.qualitative_analysis_3.update_idletasks()
		# ~ self.update()
		# ~ self.update_idletasks()
		
	def system_check_light(self):
		self.system_check.mode_check = 0
		self.forget_page()
		self.page_num = self.frame_list.index(self.system_check)
		self.switch_page()
		self.update_idletasks()
		self.system_check.serial_handle()
	
	def trial_30days_extend(self):
		self.dt = rtc.datetime
		self.recent_date = self.dt.tm_mday
		self.recent_month = self.dt.tm_mon
		self.recent_year = self.dt.tm_year
		
		time1 = trial_year*365 + trial_month*30 + trial_date
		time2 = self.recent_year*365 + self.recent_month*30 + self.recent_date
		number_of_days = time2 - time1
		print("Trial days left: ", 30 - number_of_days, '/30')
		
		if(number_of_days > 30):
			self.trial_days = 30
			self.forget_page()
			self.page_num = self.frame_list.index(self.trial_expried)
			self.switch_page()
			self.update_idletasks()
		else:
			self.system_check_light()
			
	def trial_7days(self):
		self.dt = rtc.datetime
		self.recent_date = self.dt.tm_mday
		self.recent_month = self.dt.tm_mon
		self.recent_year = self.dt.tm_year
		
		time1 = trial_year*365 + trial_month*30 + trial_date
		time2 = self.recent_year*365 + self.recent_month*30 + self.recent_date
		number_of_days = time2 - time1
		print("Trial days left: ", 7 - number_of_days, '/7')
		
		if(number_of_days > 7):
			self.trial_days = 7
			self.forget_page()
			self.page_num = self.frame_list.index(self.trial_expried)
			self.switch_page()
			self.update_idletasks()
		else:
			self.system_check_light()
			 
############################################### GUI DESIGN _ END #################################################
if __name__ == "__main__":
	app = MainWindow()
	app.mainloop()
